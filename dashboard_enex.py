"""
Dashboard Pricing Intelligence · ENEX S.A.
Gerencia Comercial · Análisis de ciclo y promociones
"""
import os, glob, gc, io
import pandas as pd
import numpy as np
import duckdb
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# ══════════════════════════════════════════════════════════════════
#  PATHS
# ══════════════════════════════════════════════════════════════════
DIR    = os.path.dirname(os.path.abspath(__file__))
FOLDER = os.path.join(DIR, "data_csv")  # optional local CSV folder
CACHE_MAIN      = os.path.join(DIR, "cache_ventas.parquet")
CACHE_PROD      = os.path.join(DIR, "cache_productos.parquet")
CACHE_TKT       = os.path.join(DIR, "cache_tickets.parquet")
CACHE_TIENDAS   = os.path.join(DIR, "cache_tiendas.parquet")
CACHE_PROMO_PROD= os.path.join(DIR, "cache_promo_prod.parquet")
EXCEL_MODEL     = os.path.join(DIR, "modelo 2026.xlsx")
CATALOGO_PROMOS = os.path.join(DIR, "catalogo_promos.csv")

# ══════════════════════════════════════════════════════════════════
#  CATEGORÍAS EXCLUIDAS DEL MODELO
# ══════════════════════════════════════════════════════════════════
CATS_EXCLUIDAS = {
    "ACCESORIOS AUTOMOTRIZ",
    "COMBOS COMIDA",
    "COMBUSTIBLES OTROS",
    "COMISIONES",
    "CUPONES SC - DESCUENTOS",
    "DONACIONES",
    "Descuento MiCopiloto",
    "IMPRESOS",
    "LUBRICANTES",
    "PROMOCIONES COMBUSTIBLE",
    "PROMOCIONES TIENDA",
    "RINCON ENEX",
    "SERVICIO AUTOMOTRIZ",
    "SERVICIOS",
    "SIN CATEGORIA",
    "TARJETA BIP",
    "TELEFONIA MOVIL",
}

# ══════════════════════════════════════════════════════════════════
#  ENEX BRAND COLORS
# ══════════════════════════════════════════════════════════════════
C_NAVY  = "#003B7A"
C_RED   = "#C8102E"
C_AMBER = "#F5A623"
C_GREEN = "#1A9E5C"
C_ALERT = "#D32F2F"
C_BLUE2 = "#0066CC"
C_GRAY  = "#64748B"
C_LIGHT = "#F1F5F9"
C_WHITE = "#FFFFFF"

CAT_COLORS = ["#003B7A","#C8102E","#F5A623","#1A9E5C","#0066CC",
              "#7B2D8B","#00838F","#E64A19","#558B2F","#AD1457",
              "#1565C0","#F57F17","#2E7D32","#6A1B9A","#00695C"]
SEQ_RG = ["#C8102E","#F5A623","#1A9E5C"]

# ══════════════════════════════════════════════════════════════════
#  CICLOS ENEX  (C1-C6 por año, fechas exactas)
# ══════════════════════════════════════════════════════════════════
_CICLOS_RAW = [
    ("C5","2023","06-09-2023","02-11-2023"),
    ("C6","2023","03-11-2023","02-01-2024"),
    ("C1","2024","03-01-2024","05-03-2024"),
    ("C2","2024","06-03-2024","07-05-2024"),
    ("C3","2024","08-05-2024","02-07-2024"),
    ("C4","2024","03-07-2024","03-09-2024"),
    ("C5","2024","04-09-2024","05-11-2024"),
    ("C6","2024","06-11-2024","07-01-2025"),
    ("C1","2025","08-01-2025","04-03-2025"),
    ("C2","2025","05-03-2025","06-05-2025"),
    ("C3","2025","07-05-2025","01-07-2025"),
    ("C4","2025","02-07-2025","02-09-2025"),
    ("C5","2025","03-09-2025","04-11-2025"),
    ("C6","2025","05-11-2025","06-01-2026"),
    ("C1","2026","07-01-2026","03-03-2026"),
    ("C2","2026","04-03-2026","05-05-2026"),
    ("C3","2026","06-05-2026","30-06-2026"),
    ("C4","2026","01-07-2026","01-09-2026"),
    ("C5","2026","02-09-2026","03-11-2026"),
    ("C6","2026","04-11-2026","05-01-2027"),
]

def _build_ciclo_maps():
    """Construye dos dicts:
       MES_CICLO  : "YYYY-MM" → "CX-YYYY"   (primario, por regla del punto medio día 15)
       CICLO_MESES: "CX-YYYY" → set("YYYY-MM")
    """
    import datetime
    mes_ciclo  = {}
    ciclo_meses = {}
    for cn, ano, fi, ff in _CICLOS_RAW:
        ca = f"{cn}-{ano}"
        d_ini = datetime.datetime.strptime(fi, "%d-%m-%Y").date()
        d_fin = datetime.datetime.strptime(ff, "%d-%m-%Y").date()
        ciclo_meses[ca] = set()
        cur = d_ini.replace(day=1)
        while cur <= d_fin:
            ym = cur.strftime("%Y-%m")
            mid = cur.replace(day=15)
            if d_ini <= mid <= d_fin:
                if ym not in mes_ciclo:
                    mes_ciclo[ym] = ca
                ciclo_meses[ca].add(ym)
            # avanzar al siguiente mes
            if cur.month == 12:
                cur = cur.replace(year=cur.year+1, month=1, day=1)
            else:
                cur = cur.replace(month=cur.month+1, day=1)
    return mes_ciclo, ciclo_meses

MES_CICLO, CICLO_MESES = _build_ciclo_maps()

# Tipos de promo conocidos (para label en filtros)
TIPOS_PROMO_CONOCIDOS = ["CIRM","CIRG","CA","UM","CRM","CRG"]

def parse_cod_promo(cod):
    """'C4-2024-001-CIRM' → (CicloNum='C4', CicloAño='C4-2024',
                              PromoAño='2024', PromoNum='001', TipoPromo='CIRM')
    Devuelve tupla de 5 strings vacíos si no parseable."""
    if not cod or not isinstance(cod, str) or not str(cod).strip():
        return ("","","","","")
    parts = str(cod).strip().upper().split("-")
    cn  = parts[0] if len(parts) >= 1 else ""
    ano = parts[1] if len(parts) >= 2 else ""
    num = parts[2] if len(parts) >= 3 else ""
    tip = parts[-1] if len(parts) >= 4 else (parts[-1] if len(parts) >= 2 else "")
    ca  = f"{cn}-{ano}" if cn and ano else ""
    return (cn, ca, ano, num, tip)

# ══════════════════════════════════════════════════════════════════
#  PAGE CONFIG & CSS
# ══════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="ENEX Pricing Intelligence",
    page_icon="⛽", layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown(f"""<style>
html,body,[class*="css"]{{font-family:'Segoe UI',Arial,sans-serif!important}}
.block-container{{padding-top:3.5rem;padding-bottom:.5rem}}

/* Ocultar toolbar de Streamlit */
[data-testid="stToolbar"]{{display:none!important}}
header[data-testid="stHeader"]{{background:transparent!important;height:0!important}}
#MainMenu{{visibility:hidden}}
footer{{visibility:hidden}}

/* Eliminar efecto desvanecido al recalcular */
[data-stale="true"]{{opacity:1!important;transition:none!important}}
.stApp[data-teststate="running"] .main .block-container{{opacity:1!important}}
div[data-testid="stAppViewContainer"]{{opacity:1!important}}
.element-container{{opacity:1!important}}
.stSpinner{{display:none!important}}

/* Sidebar */
[data-testid="stSidebar"]{{
  background:#FFFFFF;border-right:1px solid #E2E8F0;
  min-width:200px!important;max-width:600px!important
}}
[data-testid="stSidebarUserContent"]{{white-space:normal;word-break:break-word}}
.sb-logo{{
  background:linear-gradient(135deg,{C_NAVY} 0%,#0055A8 100%);
  border-radius:10px;padding:.7rem .9rem;margin-bottom:.5rem
}}
.sb-logo-title{{color:white;font-weight:800;font-size:1rem;margin:0}}
.sb-logo-sub{{color:rgba(255,255,255,.7);font-size:.68rem;margin:0}}
.sb-sec{{font-size:.65rem;font-weight:700;color:{C_GRAY};
  text-transform:uppercase;letter-spacing:.1em;margin:.7rem 0 .2rem}}

/* Header */
.enex-header{{
  background:linear-gradient(135deg,{C_NAVY} 0%,#0055A8 100%);
  padding:.9rem 1.4rem;border-radius:12px;margin-bottom:.7rem;
  display:flex;align-items:center;justify-content:space-between
}}
.enex-title{{color:white;font-size:1.3rem;font-weight:800;margin:0;letter-spacing:-.3px}}
.enex-sub{{color:rgba(255,255,255,.75);font-size:.78rem;margin:2px 0 0}}
.enex-badge{{
  background:rgba(255,255,255,.15);color:white;padding:3px 10px;
  border-radius:20px;font-size:.72rem;font-weight:600;
  border:1px solid rgba(255,255,255,.3);white-space:nowrap
}}

/* KPI Cards */
.kpi-card{{
  background:white;border-radius:10px;padding:.8rem .5rem;
  box-shadow:0 1px 4px rgba(0,0,0,.08),0 0 0 1px rgba(0,0,0,.04);
  text-align:center;height:80px;
  display:flex;flex-direction:column;justify-content:center
}}
.kv  {{font-size:1.2rem;font-weight:800;color:{C_NAVY};line-height:1.1}}
.kv-g{{font-size:1.2rem;font-weight:800;color:{C_GREEN};line-height:1.1}}
.kv-r{{font-size:1.2rem;font-weight:800;color:{C_ALERT};line-height:1.1}}
.kv-a{{font-size:1.2rem;font-weight:800;color:{C_AMBER};line-height:1.1}}
.kl  {{font-size:.6rem;color:#94A3B8;margin-top:3px;
  text-transform:uppercase;letter-spacing:.07em;font-weight:500}}

/* Section headers */
.sec{{
  font-size:.78rem;font-weight:700;color:{C_NAVY};
  border-left:3px solid {C_RED};padding:2px 0 2px 8px;
  margin:1rem 0 .4rem;text-transform:uppercase;letter-spacing:.06em
}}

/* Alert chips */
.chip-r{{display:inline-block;background:#FFF5F5;border:1px solid #FEB2B2;
  border-radius:6px;padding:.5rem .8rem;margin:.2rem;font-size:.78rem}}
.chip-a{{display:inline-block;background:#FFFBEB;border:1px solid #FBD38D;
  border-radius:6px;padding:.5rem .8rem;margin:.2rem;font-size:.78rem}}
.chip-g{{display:inline-block;background:#F0FFF4;border:1px solid #9AE6B4;
  border-radius:6px;padding:.5rem .8rem;margin:.2rem;font-size:.78rem}}

/* Tabs */
div.stTabs [data-baseweb="tab-list"]{{
  gap:2px;background:{C_LIGHT};padding:4px;border-radius:8px
}}
div.stTabs [data-baseweb="tab"]{{
  background:transparent;border-radius:6px;padding:.25rem .65rem;
  font-weight:600;font-size:.78rem;color:{C_GRAY}
}}
div.stTabs [aria-selected="true"]{{
  background:white!important;color:{C_NAVY}!important;
  box-shadow:0 1px 3px rgba(0,0,0,.1)
}}
</style>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
#  CACHE BUILD
# ══════════════════════════════════════════════════════════════════
COL_NAMES = [
    "Tiendas","Transaccion","Fecha_raw","skip1",
    "Cod_prod","Nom_prod","skip2","Categoria",
    "Cantidad","Precio","Importe","skip3",
    "Cod_promo","Nom_promo","skip4","Costo_unit",
    "Tipo","Sistema"
]
USECOLS = ["Tiendas","Transaccion","Fecha_raw","Cod_prod","Nom_prod","Categoria",
           "Cantidad","Precio","Importe","Cod_promo","Nom_promo","Costo_unit","Tipo","Sistema"]
KEY_MAIN      = ["Anio_mes","Tiendas","Categoria","Cod_promo","Nom_promo","Tipo","Sistema"]
KEY_PROD      = ["Anio_mes","Cod_prod","Nom_prod","Categoria"]
KEY_PROMO_PROD= ["Anio_mes","Cod_prod","Nom_prod","Categoria","has_promo"]

def cache_valido(files):
    if not files:
        return False
    caches = [CACHE_MAIN, CACHE_PROD, CACHE_TKT, CACHE_PROMO_PROD]
    if not all(os.path.exists(p) for p in caches):
        return False
    cm = min(os.path.getmtime(p) for p in caches)
    return cm > max(os.path.getmtime(f) for f in files)

def procesar_archivo(fp):
    facts, prods_f, tkt_pairs, promo_prods_f = [], [], [], []
    for chunk in pd.read_csv(fp, sep=";", header=None, names=COL_NAMES, usecols=USECOLS,
                             dtype=str, chunksize=150_000, encoding="utf-8",
                             on_bad_lines="skip", low_memory=False):
        f = chunk["Fecha_raw"].str.strip()
        mask = f.str.len() == 8
        chunk = chunk[mask].copy()
        if chunk.empty: continue
        chunk["Anio_mes"] = f[mask].str[:4] + "-" + f[mask].str[4:6]
        for col in ["Cantidad","Precio","Importe","Costo_unit"]:
            chunk[col] = pd.to_numeric(chunk[col], errors="coerce").fillna(0).astype("float32")
        chunk["Costo_total"]   = (chunk["Costo_unit"] * chunk["Cantidad"]).astype("float32")
        chunk["Precio_x_Cant"] = (chunk["Precio"]     * chunk["Cantidad"]).astype("float32")
        chunk["Cod_promo"]     = chunk["Cod_promo"].str.strip().replace("", np.nan)
        chunk["Nom_promo"]     = chunk["Nom_promo"].str.strip().replace("", np.nan)
        chunk["has_promo"]     = chunk["Cod_promo"].notna().astype("int8")
        tkt_pairs.append(
            chunk[["Anio_mes","Tiendas","Transaccion","Cod_promo"]].drop_duplicates(["Anio_mes","Tiendas","Transaccion"])
        )
        facts.append(
            chunk.groupby(KEY_MAIN, dropna=False)
            .agg(Importe=("Importe","sum"), Cantidad=("Cantidad","sum"),
                 Precio_x_Cant=("Precio_x_Cant","sum"), Costo_total=("Costo_total","sum"))
            .reset_index()
        )
        prods_f.append(
            chunk.groupby(KEY_PROD, dropna=False)
            .agg(Importe=("Importe","sum"), Cantidad=("Cantidad","sum"),
                 Precio_x_Cant=("Precio_x_Cant","sum"), Costo_total=("Costo_total","sum"))
            .reset_index()
        )
        promo_prods_f.append(
            chunk.groupby(KEY_PROMO_PROD, dropna=False)
            .agg(Importe=("Importe","sum"), Cantidad=("Cantidad","sum"),
                 Precio_x_Cant=("Precio_x_Cant","sum"), Costo_total=("Costo_total","sum"))
            .reset_index()
        )
    if not facts:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    dm_ = pd.concat(facts).groupby(KEY_MAIN, dropna=False)[
        ["Importe","Cantidad","Precio_x_Cant","Costo_total"]].sum().reset_index()
    dp_ = pd.concat(prods_f).groupby(KEY_PROD, dropna=False)[
        ["Importe","Cantidad","Precio_x_Cant","Costo_total"]].sum().reset_index()
    tkt_raw = pd.concat(tkt_pairs).drop_duplicates(["Anio_mes","Tiendas","Transaccion"])
    tkt_raw["has_promo"] = tkt_raw["Cod_promo"].notna().astype("int8")
    dt_ = tkt_raw.groupby(["Anio_mes","Tiendas"]).agg(
        n_tickets=("Transaccion","count"), n_tkt_promo=("has_promo","sum")).reset_index()
    dpp_ = pd.concat(promo_prods_f).groupby(KEY_PROMO_PROD, dropna=False)[
        ["Importe","Cantidad","Precio_x_Cant","Costo_total"]].sum().reset_index()
    return dm_, dp_, dt_, dpp_

def build_cache(files):
    ph = st.empty(); bar = st.progress(0)
    mains, prods, tkts, promo_prods = [], [], [], []
    for i, f in enumerate(files):
        ph.info(f"Procesando **{os.path.basename(f)}** ({i+1}/{len(files)})…")
        bar.progress(int(i/len(files)*100))
        dm_, dp_, dt_, dpp_ = procesar_archivo(f)
        if not dm_.empty:  mains.append(dm_)
        if not dp_.empty:  prods.append(dp_)
        if not dt_.empty:  tkts.append(dt_)
        if not dpp_.empty: promo_prods.append(dpp_)
        gc.collect()
    bar.progress(100); ph.info("Guardando cache…")
    pd.concat(mains).groupby(KEY_MAIN, dropna=False)[
        ["Importe","Cantidad","Precio_x_Cant","Costo_total"]].sum().reset_index().to_parquet(CACHE_MAIN, index=False)
    pd.concat(prods).groupby(KEY_PROD, dropna=False)[
        ["Importe","Cantidad","Precio_x_Cant","Costo_total"]].sum().reset_index().to_parquet(CACHE_PROD, index=False)
    pd.concat(tkts).groupby(["Anio_mes","Tiendas"])[
        ["n_tickets","n_tkt_promo"]].sum().reset_index().to_parquet(CACHE_TKT, index=False)
    pd.concat(promo_prods).groupby(KEY_PROMO_PROD, dropna=False)[
        ["Importe","Cantidad","Precio_x_Cant","Costo_total"]].sum().reset_index().to_parquet(CACHE_PROMO_PROD, index=False)
    ph.empty(); bar.empty()

@st.cache_resource(show_spinner=False)
def load_data():
    files = sorted(glob.glob(os.path.join(FOLDER, "*.csv"))) if os.path.isdir(FOLDER) else []
    # If cache already exists, use it directly without needing CSVs
    cache_exists = os.path.exists(CACHE_MAIN) and os.path.exists(CACHE_PROD) and os.path.exists(CACHE_TKT)
    if not cache_exists:
        if not files: return None, None, None, None, 0
        build_cache(files)
    elif files and not cache_valido(files):
        build_cache(files)
    dm_  = pd.read_parquet(CACHE_MAIN)
    dp_  = pd.read_parquet(CACHE_PROD)
    dt_  = pd.read_parquet(CACHE_TKT)
    dpp_ = pd.read_parquet(CACHE_PROMO_PROD) if os.path.exists(CACHE_PROMO_PROD) else pd.DataFrame()
    if "Precio_x_Cant" not in dp_.columns:
        dp_["Precio_x_Cant"] = dp_["Importe"]
    # ── Excluir categorías fuera del modelo ──
    if "Categoria" in dm_.columns:
        dm_ = dm_[~dm_["Categoria"].isin(CATS_EXCLUIDAS)].copy()
    if "Categoria" in dp_.columns:
        dp_ = dp_[~dp_["Categoria"].isin(CATS_EXCLUIDAS)].copy()
    if "Categoria" in dpp_.columns and not dpp_.empty:
        dpp_ = dpp_[~dpp_["Categoria"].isin(CATS_EXCLUIDAS)].copy()
    # ── Columnas derivadas del código de promo ──
    if "Cod_promo" in dm_.columns:
        parsed = dm_["Cod_promo"].apply(lambda x: parse_cod_promo(x))
        dm_["CicloNum_p"] = parsed.apply(lambda t: t[0])
        dm_["CicloAño_p"] = parsed.apply(lambda t: t[1])
        dm_["PromoAño"]   = parsed.apply(lambda t: t[2])
        dm_["PromoNum"]   = parsed.apply(lambda t: t[3])
        dm_["TipoPromo"]  = parsed.apply(lambda t: t[4])
    # ── CicloAño desde Anio_mes (columna de ciclo de la fila) ──
    if "Anio_mes" in dm_.columns:
        dm_["CicloAño_mes"] = dm_["Anio_mes"].map(MES_CICLO).fillna("")
    # Tiendas como string para join
    if "Tiendas" in dm_.columns:
        dm_["Tiendas"] = dm_["Tiendas"].astype(str).str.strip()
    return dm_, dp_, dt_, dpp_, len(files)


def merge_csv_al_cache(uploaded_file):
    """Procesa un CSV subido y lo fusiona con el caché existente sin tocar los otros archivos."""
    import tempfile
    ph = st.empty(); bar = st.progress(0)
    ph.info("📖 Leyendo archivo…")
    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as tmp:
        tmp.write(uploaded_file.getvalue())
        tmp_path = tmp.name
    try:
        ph.info("⚙️ Procesando datos…"); bar.progress(15)
        dm_new, dp_new, dt_new, dpp_new = procesar_archivo(tmp_path)
    finally:
        os.unlink(tmp_path)
    if dm_new.empty:
        ph.error("❌ El archivo no tiene datos válidos. Verifica el formato.")
        bar.empty(); return []
    meses_nuevos = dm_new["Anio_mes"].unique().tolist()
    ph.info(f"🔀 Fusionando {len(meses_nuevos)} mes(es)…"); bar.progress(35)

    # ── ventas ──
    if os.path.exists(CACHE_MAIN):
        _ex = pd.read_parquet(CACHE_MAIN)
        _ex = _ex[~_ex["Anio_mes"].isin(meses_nuevos)]
        dm_fin = pd.concat([_ex, dm_new]).groupby(KEY_MAIN, dropna=False)[
            ["Importe","Cantidad","Precio_x_Cant","Costo_total"]].sum().reset_index()
    else:
        dm_fin = dm_new
    bar.progress(50)

    # ── productos ──
    if os.path.exists(CACHE_PROD):
        _ex = pd.read_parquet(CACHE_PROD)
        _ex = _ex[~_ex["Anio_mes"].isin(meses_nuevos)]
        dp_fin = pd.concat([_ex, dp_new]).groupby(KEY_PROD, dropna=False)[
            ["Importe","Cantidad","Precio_x_Cant","Costo_total"]].sum().reset_index()
    else:
        dp_fin = dp_new
    bar.progress(65)

    # ── tickets ──
    if os.path.exists(CACHE_TKT):
        _ex = pd.read_parquet(CACHE_TKT)
        _ex = _ex[~_ex["Anio_mes"].isin(meses_nuevos)]
        dt_fin = pd.concat([_ex, dt_new]).groupby(["Anio_mes","Tiendas"])[
            ["n_tickets","n_tkt_promo"]].sum().reset_index()
    else:
        dt_fin = dt_new
    bar.progress(78)

    # ── promo_prod ──
    if not dpp_new.empty:
        if os.path.exists(CACHE_PROMO_PROD):
            _ex = pd.read_parquet(CACHE_PROMO_PROD)
            _ex = _ex[~_ex["Anio_mes"].isin(meses_nuevos)]
            dpp_fin = pd.concat([_ex, dpp_new]).groupby(KEY_PROMO_PROD, dropna=False)[
                ["Importe","Cantidad","Precio_x_Cant","Costo_total"]].sum().reset_index()
        else:
            dpp_fin = dpp_new
        dpp_fin.to_parquet(CACHE_PROMO_PROD, index=False)
    bar.progress(90)

    ph.info("💾 Guardando…")
    dm_fin.to_parquet(CACHE_MAIN,  index=False)
    dp_fin.to_parquet(CACHE_PROD,  index=False)
    dt_fin.to_parquet(CACHE_TKT,   index=False)
    bar.progress(100); ph.empty(); bar.empty()
    return sorted(meses_nuevos)


@st.cache_resource(show_spinner=False)
def load_tiendas():
    """Carga info de tiendas desde cache parquet o desde Excel."""
    if os.path.exists(CACHE_TIENDAS):
        try:
            df_t = pd.read_parquet(CACHE_TIENDAS)
            if not df_t.empty:
                return df_t
        except Exception:
            pass
    if not os.path.exists(EXCEL_MODEL):
        return pd.DataFrame()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_MODEL, read_only=True, data_only=True)
        ws = wb["info de tiendas"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        if len(rows) < 2:
            return pd.DataFrame()
        headers = [str(c).strip() if c is not None else f"col{i}" for i, c in enumerate(rows[0])]
        data = [list(r) for r in rows[1:] if any(v is not None for v in r)]
        df_t = pd.DataFrame(data, columns=headers)
        if "FileCORR" in df_t.columns:
            def _to_str(x):
                try:
                    v = str(int(float(x)))
                    return v
                except Exception:
                    return str(x).strip() if x is not None else ""
            df_t["FileCORR"] = df_t["FileCORR"].apply(_to_str)
        df_t.to_parquet(CACHE_TIENDAS, index=False)
        return df_t
    except Exception:
        return pd.DataFrame()

@st.cache_resource(show_spinner=False)
def load_catalogo():
    """Carga el catálogo de promos (Cod_promo, Nom_promo, Formato_promo, Unidades_por_promo).
    Si existe un archivo subido por el usuario (catalogo_promos_custom.csv) lo usa en lugar del default."""
    custom = os.path.join(DIR, "catalogo_promos_custom.csv")
    path = custom if os.path.exists(custom) else CATALOGO_PROMOS
    if not os.path.exists(path):
        return pd.DataFrame()
    try:
        df_cat = pd.read_csv(path, dtype=str)
        df_cat.columns = [c.strip() for c in df_cat.columns]
        # Normalizar nombre de columnas esperadas
        rename_map = {}
        for col in df_cat.columns:
            cl = col.lower().replace(" ", "_")
            if "cod" in cl and "promo" in cl:
                rename_map[col] = "Cod_promo"
            elif "nom" in cl and "promo" in cl:
                rename_map[col] = "Nom_promo"
            elif "formato" in cl:
                rename_map[col] = "Formato_promo"
            elif "unidades" in cl or "unid" in cl:
                rename_map[col] = "Unidades_por_promo"
            elif "código" in cl or "codigo" in cl:
                rename_map[col] = "Cod_promo"
        df_cat = df_cat.rename(columns=rename_map)
        for col in ["Cod_promo","Nom_promo","Formato_promo","Unidades_por_promo"]:
            if col not in df_cat.columns:
                df_cat[col] = ""
        df_cat["Cod_promo"] = df_cat["Cod_promo"].str.strip().str.upper()
        df_cat["Unidades_por_promo"] = pd.to_numeric(df_cat["Unidades_por_promo"], errors="coerce").fillna(1).astype(int)
        return df_cat[["Cod_promo","Nom_promo","Formato_promo","Unidades_por_promo"]].drop_duplicates("Cod_promo")
    except Exception:
        return pd.DataFrame()


dm, dp, dt, dpp, n_files = load_data()
if dm is None:
    st.error(f"No se encontraron CSV en:\n`{FOLDER}`"); st.stop()
if dm.empty:
    st.error("Los CSV están vacíos."); st.stop()

df_tiendas  = load_tiendas()   # puede ser empty DataFrame si falla
df_catalogo = load_catalogo()  # catálogo de promos con Unidades_por_promo


# ══════════════════════════════════════════════════════════════════
#  HELPERS UI
# ══════════════════════════════════════════════════════════════════
def clp(v):
    try:
        x = float(v)
        if abs(x)>=1e9:  return f"${x/1e9:.1f} mil mill."
        if abs(x)>=1e6:  return f"${x/1e6:.1f} mill."
        if abs(x)>=1e3:  return f"${x/1e3:.0f} mil"
        return f"${x:,.0f}"
    except: return "–"

def num(v):
    try:
        x=float(v)
        if abs(x)>=1e9: return f"{x/1e9:.1f} mil mill."
        if abs(x)>=1e6: return f"{x/1e6:.1f} mill."
        if abs(x)>=1e3: return f"{x/1e3:.0f} mil"
        return f"{x:,.0f}"
    except: return "–"

def pct(v,d=1):
    try: return f"{float(v)*100:.{d}f}%"
    except: return "–"

def s(v):
    try:
        r=float(v); return 0.0 if r!=r else r
    except: return 0.0

def sz(series): return series.clip(lower=0.01)

def kpi(col, val, lbl, c=""):
    cls = f"kv{'-'+c if c else ''}"
    col.markdown(f'<div class="kpi-card"><div class="{cls}">{val}</div>'
                 f'<div class="kl">{lbl}</div></div>', unsafe_allow_html=True)

def sec(txt):
    st.markdown(f'<div class="sec">{txt}</div>', unsafe_allow_html=True)

def pcfg(fig, h=300, legend=True, title=None):
    """Estilo limpio ENEX para todos los gráficos."""
    layout = dict(
        plot_bgcolor="white", paper_bgcolor="white",
        height=h,
        margin=dict(t=12, b=8, l=8, r=8),
        font=dict(family="Segoe UI,Arial", size=11, color="#374151"),
        showlegend=legend,
        hoverlabel=dict(
            bgcolor="white", bordercolor="#E5E7EB",
            font_size=12, font_family="Segoe UI,Arial"
        ),
    )
    if legend:
        layout["legend"] = dict(
            orientation="h", y=1.08, x=0,
            font=dict(size=10), bgcolor="rgba(0,0,0,0)",
            bordercolor="rgba(0,0,0,0)"
        )
    fig.update_layout(**layout)
    fig.update_xaxes(
        showgrid=False, zeroline=False,
        linecolor="#E5E7EB", linewidth=1,
        tickfont=dict(size=10, color="#6B7280"),
        title_font=dict(size=11, color="#374151")
    )
    fig.update_yaxes(
        gridcolor="#F3F4F6", gridwidth=1, zeroline=False,
        linecolor="#E5E7EB", linewidth=1,
        tickfont=dict(size=10, color="#6B7280"),
        title_font=dict(size=11, color="#374151")
    )
    return fig

def _fmt_clp(x):
    """Formatea número como CLP para ejes (sin símbolo $)."""
    if abs(x) >= 1e9:  return f"{x/1e9:.1f} B"
    if abs(x) >= 1e6:  return f"{x/1e6:.0f} M"
    if abs(x) >= 1e3:  return f"{x/1e3:.0f} K"
    return f"{x:.0f}"

def bar_chart(df, x, y, color=None, color_col=None, orient="v",
              text_col=None, title=None, h=300, colorscale=None):
    """Gráfico de barras limpio con etiquetas."""
    if orient == "h":
        fig = px.bar(df, x=x, y=y, orientation="h",
                     color=color_col,
                     color_discrete_sequence=[color] if color and not color_col else None,
                     color_continuous_scale=colorscale,
                     text=text_col)
        fig.update_traces(
            textposition="outside",
            textfont_size=10,
            marker_line_width=0,
            insidetextanchor="middle"
        )
        fig.update_layout(yaxis=dict(autorange="reversed"))
    else:
        fig = px.bar(df, x=x, y=y,
                     color=color_col,
                     color_discrete_sequence=[color] if color and not color_col else None,
                     color_continuous_scale=colorscale,
                     text=text_col)
        fig.update_traces(
            textposition="outside",
            textfont_size=10,
            marker_line_width=0
        )
    return pcfg(fig, h=h, legend=bool(color_col), title=title)


# ══════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════
def generar_excel(D, filtro_desc=""):
    """Genera un Excel con múltiples hojas de datos y gráficos nativos de Excel."""
    import openpyxl
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY  = "003B7A"; RED = "C8102E"; AMBER = "F5A623"; GREEN = "1A9E5C"
    HDR_FILL  = PatternFill("solid", fgColor=NAVY)
    HDR_FONT  = Font(color="FFFFFF", bold=True, size=10)
    TITLE_FONT= Font(color=NAVY, bold=True, size=13)
    SUB_FONT  = Font(color="64748B", italic=True, size=9)
    thin = Side(style="thin", color="E2E8F0")
    BORDER= Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write_header(ws, title, subtitle=""):
        ws["A1"] = title;     ws["A1"].font = TITLE_FONT
        ws["A2"] = subtitle;  ws["A2"].font = SUB_FONT
        ws["A3"] = ""

    def _write_df(ws, df, start_row=4, pct_cols=None, money_cols=None):
        pct_cols   = pct_cols   or []
        money_cols = money_cols or []
        # Header
        for c, col in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=c, value=col)
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
        # Data
        for r, row in enumerate(df.itertuples(index=False), start_row+1):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = BORDER
                col_name = df.columns[c-1]
                if col_name in pct_cols and isinstance(val, (int,float)):
                    cell.number_format = "0.0%"
                elif col_name in money_cols and isinstance(val, (int,float)):
                    cell.number_format = '#,##0'
        # Column widths
        for c, col in enumerate(df.columns, 1):
            max_w = max(len(str(col)), *(len(str(v)) for v in df.iloc[:,c-1].astype(str)))
            ws.column_dimensions[get_column_letter(c)].width = min(max_w+3, 30)
        return start_row, start_row + len(df)

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ──
    ws = wb.active; ws.title = "Resumen"
    _write_header(ws, "ENEX Pricing Intelligence — Resumen", filtro_desc)
    ev = D.get("evol_v", pd.DataFrame())
    if not ev.empty:
        _ev_cols = [c for c in ["Anio_mes","Ingresos","Margen","Mg_pct","Descuento"] if c in ev.columns]
        df_ev = ev[_ev_cols].copy()
        sr, er = _write_df(ws, df_ev, start_row=4, pct_cols=["Mg_pct","Mix_promo"], money_cols=["Ingresos","Margen"])
        # Gráfico de barras: Ingresos por mes
        chart = BarChart()
        chart.type = "col"; chart.grouping = "clustered"
        chart.title = "Ingresos por Mes"; chart.style = 10
        chart.y_axis.title = "Ingresos ($)"; chart.x_axis.title = "Mes"
        data = Reference(ws, min_col=2, max_col=2, min_row=sr+4, max_row=er+4)
        cats = Reference(ws, min_col=1, min_row=sr+5, max_row=er+4)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4; chart.width = 20; chart.height = 12
        ws.add_chart(chart, f"G4")
        # Gráfico de línea: Margen %
        chart2 = LineChart()
        chart2.title = "Margen % por Mes"; chart2.style = 10
        chart2.y_axis.title = "Margen %"; chart2.y_axis.numFmt = "0%"
        data2 = Reference(ws, min_col=4, max_col=4, min_row=sr+4, max_row=er+4)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.width = 20; chart2.height = 12
        ws.add_chart(chart2, "G24")

    # ── Hoja 2: Categorías ──
    ws2 = wb.create_sheet("Categorías")
    _write_header(ws2, "Ventas por Categoría", filtro_desc)
    tm = D.get("treemap", pd.DataFrame())
    if not tm.empty:
        _tm_cols = [c for c in ["Categoria","Ingresos","Unidades","Margen","Mg_pct","Mix_promo","Desc_pct"] if c in tm.columns]
        df_cat = tm[_tm_cols].copy()
        sr2, er2 = _write_df(ws2, df_cat, start_row=4,
                              pct_cols=["Mg_pct","Mix_promo","Desc_pct"],
                              money_cols=["Ingresos","Margen"])
        chart3 = BarChart()
        chart3.type = "bar"; chart3.grouping = "clustered"
        chart3.title = "Ingresos por Categoría"; chart3.style = 10
        data3 = Reference(ws2, min_col=2, max_col=2, min_row=sr2+4, max_row=er2+4)
        cats3 = Reference(ws2, min_col=1, min_row=sr2+5, max_row=er2+4)
        chart3.add_data(data3, titles_from_data=True)
        chart3.set_categories(cats3)
        chart3.width = 20; chart3.height = 14
        ws2.add_chart(chart3, "J4")

    # ── Hoja 3: SKU Ranking ──
    ws3 = wb.create_sheet("SKU Ranking")
    _write_header(ws3, "Ranking de Productos (SKUs)", filtro_desc)
    rk = D.get("sku_ranking", pd.DataFrame())
    if not rk.empty:
        _rk_cols = [c for c in ["Producto","Categoria","Ingresos","Unidades","Margen",
                     "Mg_pct","Precio_ef","Desc_pct","Primer_mes","Ultimo_mes"] if c in rk.columns]
        df_rk = rk[_rk_cols].copy()
        _write_df(ws3, df_rk, start_row=4,
                  pct_cols=["Mg_pct","Desc_pct"], money_cols=["Ingresos","Margen","Precio_ef"])

    # ── Hoja 4: Tiendas ──
    ws4 = wb.create_sheet("Tiendas")
    _write_header(ws4, "Ranking de Tiendas", filtro_desc)
    tn = D.get("tiendas_v", pd.DataFrame())
    if not tn.empty:
        _tn_cols = [c for c in ["Tiendas","Ingresos","Unidades","Margen","Mg_pct","Mix_promo","Desc_pct","Precio_ef"] if c in tn.columns]
        df_tn = tn[_tn_cols].copy()
        sr4, er4 = _write_df(ws4, df_tn, start_row=4,
                              pct_cols=["Mg_pct","Mix_promo","Desc_pct"],
                              money_cols=["Ingresos","Margen","Precio_ef"])
        chart4 = BarChart()
        chart4.type = "bar"; chart4.grouping = "clustered"
        chart4.title = "Top Tiendas por Ingresos"; chart4.style = 10
        data4 = Reference(ws4, min_col=2, max_col=2, min_row=sr4+4, max_row=min(sr4+19, er4+4))
        cats4 = Reference(ws4, min_col=1, min_row=sr4+5, max_row=min(sr4+19, er4+4))
        chart4.add_data(data4, titles_from_data=True)
        chart4.set_categories(cats4)
        chart4.width = 22; chart4.height = 14
        ws4.add_chart(chart4, "J4")

    # ── Hoja 5: Promociones ──
    ws5 = wb.create_sheet("Promociones")
    _write_header(ws5, "Ranking de Promociones", filtro_desc)
    pr = D.get("promos", pd.DataFrame())
    if not pr.empty:
        _pr_cols = [c for c in ["Promo","Primer_mes","Ultimo_mes","Meses","Tiendas",
                     "Ingresos","Unidades","Margen","Mg_pct","Desc_pct","ROI"] if c in pr.columns]
        df_pr = pr[_pr_cols].copy()
        _write_df(ws5, df_pr, start_row=4,
                  pct_cols=["Mg_pct","Desc_pct"], money_cols=["Ingresos","Margen"])

    # ── Hoja 6: YoY ──
    ws6 = wb.create_sheet("YoY")
    _write_header(ws6, "Comparativa Año vs Año", filtro_desc)
    yoy = D.get("yoy_mes", pd.DataFrame())
    if not yoy.empty:
        _yoy_cols = [c for c in ["Anio_mes","Año","Mes","Ingresos","Unidades","Margen","Mg_pct","Mix_promo"] if c in yoy.columns]
        df_yoy = yoy[_yoy_cols].copy()
        sr6, er6 = _write_df(ws6, df_yoy, start_row=4,
                              pct_cols=["Mg_pct","Mix_promo"], money_cols=["Ingresos","Margen"])
        chart6 = LineChart()
        chart6.title = "Ingresos por Mes — YoY"; chart6.style = 10
        chart6.y_axis.title = "Ingresos ($)"
        data6 = Reference(ws6, min_col=4, max_col=4, min_row=sr6+4, max_row=er6+4)
        cats6 = Reference(ws6, min_col=1, min_row=sr6+5, max_row=er6+4)
        chart6.add_data(data6, titles_from_data=True)
        chart6.set_categories(cats6)
        chart6.width = 22; chart6.height = 12
        ws6.add_chart(chart6, "J4")

    # ── Hoja 7: Datos Raw ──
    ws7 = wb.create_sheet("Datos")
    _write_header(ws7, "Datos Agregados (filtro aplicado)", filtro_desc)
    res = D.get("resumen_export", pd.DataFrame())
    if not res.empty:
        _write_df(ws7, res, start_row=4,
                  money_cols=["Ingresos","Costo","Margen","Descuento"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(f"""
    <div class="sb-logo">
      <div class="sb-logo-title">⛽ ENEX S.A.</div>
      <div class="sb-logo-sub">Pricing Intelligence</div>
    </div>""", unsafe_allow_html=True)
    st.caption(f"{n_files} archivos · caché activo")

    meses_all = sorted(dm["Anio_mes"].dropna().unique().tolist())
    cats_all  = sorted(dm["Categoria"].dropna().unique().tolist())
    tipos_all = sorted(dm["Tipo"].dropna().unique().tolist())
    sists_all = sorted(dm["Sistema"].dropna().unique().tolist())

    # ── Período (mes) ──
    st.markdown('<div class="sb-sec">Período</div>', unsafe_allow_html=True)
    meses_sel = st.multiselect("Meses", meses_all, default=meses_all, key="ms",
                               label_visibility="collapsed")

    # ── Ciclo ──
    ciclo_años_all = sorted(CICLO_MESES.keys())   # CX-YYYY
    ciclo_nums_all = ["C1","C2","C3","C4","C5","C6"]
    años_all       = sorted({m[:4] for m in meses_all})

    st.markdown('<div class="sb-sec">Ciclo</div>', unsafe_allow_html=True)
    ciclo_años_sel = st.multiselect("Ciclo-Año (ej: C2-2026)", ciclo_años_all,
                                    default=[], key="cay", label_visibility="visible")
    ciclo_nums_sel = st.multiselect("Número de Ciclo", ciclo_nums_all,
                                    default=[], key="cnu", label_visibility="visible")
    años_sel       = st.multiselect("Año", años_all, default=[], key="ano",
                                    label_visibility="visible")

    # ── Categoría ──
    st.markdown('<div class="sb-sec">Categoría</div>', unsafe_allow_html=True)
    cats_sel = st.multiselect("Categorías", cats_all, default=cats_all, key="cs",
                              label_visibility="collapsed")

    # ── Tipo · Sistema ──
    st.markdown('<div class="sb-sec">Tipo · Sistema</div>', unsafe_allow_html=True)
    tipos_sel = st.multiselect("Tipos", tipos_all, default=tipos_all, key="ts",
                               label_visibility="collapsed")
    sists_sel = st.multiselect("Sistemas", sists_all, default=sists_all, key="ss",
                               label_visibility="collapsed")

    # ── Filtros de Promo por código ──
    tipos_promo_all  = sorted(t for t in dm["TipoPromo"].dropna().unique() if t) \
                       if "TipoPromo" in dm.columns else []
    ciclos_promo_all = sorted(t for t in dm["CicloAño_p"].dropna().unique() if t) \
                       if "CicloAño_p" in dm.columns else []
    nums_promo_all   = sorted(t for t in dm["PromoNum"].dropna().unique() if t) \
                       if "PromoNum" in dm.columns else []

    st.markdown('<div class="sb-sec">Tipo de Promo</div>', unsafe_allow_html=True)
    tipos_promo_sel  = st.multiselect("Tipo (CIRM, CRM, CA…)", tipos_promo_all,
                                      default=[], key="tpr", label_visibility="visible")
    st.markdown('<div class="sb-sec">Ciclo de Promo</div>', unsafe_allow_html=True)
    ciclos_promo_sel = st.multiselect("Ciclo-Año de Promo", ciclos_promo_all,
                                      default=[], key="cpr", label_visibility="visible")
    nums_promo_sel   = st.multiselect("Número de Promo", nums_promo_all,
                                      default=[], key="npr", label_visibility="visible")

    # ── Tiendas: Región, Comuna, Formato, Segmento ──
    if not df_tiendas.empty:
        regiones_all  = sorted(df_tiendas["Región"].dropna().unique().tolist())  \
                        if "Región"   in df_tiendas.columns else []
        comunas_all   = sorted(df_tiendas["Comuna"].dropna().unique().tolist())  \
                        if "Comuna"   in df_tiendas.columns else []
        formatos_all  = sorted(df_tiendas["Formato"].dropna().unique().tolist()) \
                        if "Formato"  in df_tiendas.columns else []
        segmentos_all = sorted(df_tiendas["Segmento"].dropna().unique().tolist())\
                        if "Segmento" in df_tiendas.columns else []

        st.markdown('<div class="sb-sec">Tiendas</div>', unsafe_allow_html=True)
        regiones_sel  = st.multiselect("Región",   regiones_all,  default=[], key="reg")
        comunas_sel   = st.multiselect("Comuna",   comunas_all,   default=[], key="com")
        formatos_sel  = st.multiselect("Formato",  formatos_all,  default=[], key="fmt")
        segmentos_sel = st.multiselect("Segmento", segmentos_all, default=[], key="seg")
    else:
        regiones_all  = comunas_all  = formatos_all  = segmentos_all = []
        regiones_sel  = comunas_sel  = formatos_sel  = segmentos_sel = []

    # ── Filtro directo por Tienda ──
    tiendas_all = sorted(dm["Tiendas"].dropna().unique().tolist()) \
                  if "Tiendas" in dm.columns else []
    st.markdown('<div class="sb-sec">Tienda (código)</div>', unsafe_allow_html=True)
    tiendas_dir_sel = st.multiselect("Tienda", tiendas_all, default=[], key="tnd",
                                     label_visibility="collapsed")

    # ── Filtro directo por Producto ──
    prods_all = sorted(dp["Nom_prod"].dropna().unique().tolist()) \
                if "Nom_prod" in dp.columns else []
    st.markdown('<div class="sb-sec">Producto</div>', unsafe_allow_html=True)
    prods_dir_sel = st.multiselect("Producto", prods_all, default=[], key="prd",
                                   label_visibility="collapsed")

    # ── Búsqueda texto ──
    st.markdown('<div class="sb-sec">Búsqueda texto</div>', unsafe_allow_html=True)
    promo_search = st.text_input("Promo", placeholder="ej: Monster, 2x1…", key="ps")
    prod_search  = st.text_input("Producto", placeholder="ej: Coca-Cola…", key="prs")
    cod_search   = st.text_input("Código", placeholder="ej: 2702324", key="cds")

    st.markdown('<div class="sb-sec">Ranking</div>', unsafe_allow_html=True)
    top_n = st.slider("Top N", 5, 30, 15)

    # ── Subir CSV nuevo ──
    st.divider()
    st.markdown('<div class="sb-sec">📁 Agregar datos (CSV)</div>', unsafe_allow_html=True)

    _meses_actuales = sorted(dm["Anio_mes"].dropna().unique().tolist())
    if _meses_actuales:
        st.caption(f"En caché: **{_meses_actuales[0]}** → **{_meses_actuales[-1]}** ({len(_meses_actuales)} meses)")

    _csv_up = st.file_uploader(
        "Arrastra o selecciona un CSV", type=["csv"],
        key="csv_upload", label_visibility="collapsed",
        help="El archivo debe tener el mismo formato que los CSV de ciclo ENEX."
    )
    if _csv_up is not None:
        if st.button("✅ Incorporar al dashboard", use_container_width=True, key="btn_upload"):
            _meses_ok = merge_csv_al_cache(_csv_up)
            if _meses_ok:
                st.success(f"Incorporado: {', '.join(_meses_ok)}")
                st.cache_resource.clear(); st.cache_data.clear(); st.rerun()

    # ── Subir catálogo de promos ──
    st.divider()
    st.markdown('<div class="sb-sec">📋 Catálogo de Promos</div>', unsafe_allow_html=True)
    if not df_catalogo.empty:
        st.caption(f"Catálogo cargado: **{len(df_catalogo)}** promos")
    _cat_up = st.file_uploader(
        "Subir catálogo CSV (Cod_promo, Nom_promo, Formato_promo, Unidades_por_promo)",
        type=["csv"], key="cat_upload", label_visibility="collapsed",
        help="CSV con columnas: Código promoción, Promoción, Formato promo, Unidades por promoción"
    )
    if _cat_up is not None:
        if st.button("✅ Cargar catálogo", use_container_width=True, key="btn_cat_upload"):
            _cat_path = os.path.join(DIR, "catalogo_promos_custom.csv")
            with open(_cat_path, "wb") as f_:
                f_.write(_cat_up.getvalue())
            st.cache_resource.clear(); st.rerun()

    st.divider()
    if st.button("🔄 Recalcular cache completo", use_container_width=True):
        for p_ in [CACHE_MAIN, CACHE_PROD, CACHE_TKT, CACHE_TIENDAS, CACHE_PROMO_PROD]:
            if os.path.exists(p_): os.remove(p_)
        st.cache_resource.clear(); st.cache_data.clear(); st.rerun()


# ══════════════════════════════════════════════════════════════════
#  FILTRO RÁPIDO EN PANDAS + TODAS LAS QUERIES EN UNA FUNCIÓN
#  → cached por combo de filtros, no por WHERE string
# ══════════════════════════════════════════════════════════════════
_ms            = tuple(sorted(meses_sel       or meses_all))
_cats          = tuple(sorted(cats_sel        or cats_all))
_tipos         = tuple(sorted(tipos_sel       or tipos_all))
_sists         = tuple(sorted(sists_sel       or sists_all))
_ps            = promo_search.strip()
_prs           = prod_search.strip()
_cs            = cod_search.strip()
_ciclo_años    = tuple(sorted(ciclo_años_sel))
_ciclo_nums    = tuple(sorted(ciclo_nums_sel))
_años          = tuple(sorted(años_sel))
_tipos_promo   = tuple(sorted(tipos_promo_sel))
_ciclos_promo  = tuple(sorted(ciclos_promo_sel))
_nums_promo    = tuple(sorted(nums_promo_sel))
_regiones      = tuple(sorted(regiones_sel))
_comunas       = tuple(sorted(comunas_sel))
_formatos_t    = tuple(sorted(formatos_sel))
_segmentos     = tuple(sorted(segmentos_sel))
_tiendas_dir   = tuple(sorted(tiendas_dir_sel))
_prods_dir     = tuple(sorted(prods_dir_sel))

filter_key = (
    _ms, _cats, _tipos, _sists, _ps, _prs, _cs,
    _ciclo_años, _ciclo_nums, _años,
    _tipos_promo, _ciclos_promo, _nums_promo,
    _regiones, _comunas, _formatos_t, _segmentos,
    _tiendas_dir, _prods_dir,
)

@st.cache_data(ttl=900, show_spinner="Calculando…", max_entries=20)
def all_data(fk):
    """Una sola función cacheada que pre-filtra en pandas y corre todas las queries."""
    (ms, cats, tipos, sists, ps, prs, cs,
     ciclo_años, ciclo_nums, años_flt,
     tipos_promo, ciclos_promo, nums_promo,
     regiones, comunas, formatos_t, segmentos,
     tiendas_dir, prods_dir) = fk

    ms_s    = set(ms);    cats_s  = set(cats)
    tipos_s = set(tipos); sists_s = set(sists)

    # ── Conjunto efectivo de meses considerando filtros de ciclo / año ──
    eff_ms = set(ms) if ms else set(meses_all)
    if ciclo_años:
        cy = set()
        for ca in ciclo_años:
            cy.update(CICLO_MESES.get(ca, set()))
        eff_ms = eff_ms & cy if eff_ms else cy
    if ciclo_nums:
        cn = set()
        for ca, mset in CICLO_MESES.items():
            if ca.split("-")[0] in set(ciclo_nums):
                cn.update(mset)
        eff_ms = eff_ms & cn if eff_ms else cn
    if años_flt:
        eff_ms = {m for m in eff_ms if m[:4] in set(años_flt)}

    # ── Pre-filtro pandas (muy rápido en datos pre-agregados) ──
    v = dm.copy()
    if eff_ms != set(meses_all):   v = v[v["Anio_mes"].isin(eff_ms)]
    if cats_s  != set(cats_all):   v = v[v["Categoria"].isin(cats_s)]
    if tipos_s != set(tipos_all):  v = v[v["Tipo"].isin(tipos_s)]
    if sists_s != set(sists_all):  v = v[v["Sistema"].isin(sists_s)]
    if ps: v = v[v["Nom_promo"].str.contains(ps, case=False, na=False)]

    # ── Filtros de código de promo ──
    if tipos_promo and "TipoPromo" in v.columns:
        v = v[v["TipoPromo"].isin(set(tipos_promo))]
    if ciclos_promo and "CicloAño_p" in v.columns:
        v = v[v["CicloAño_p"].isin(set(ciclos_promo))]
    if nums_promo and "PromoNum" in v.columns:
        v = v[v["PromoNum"].isin(set(nums_promo))]

    # ── Filtro de tiendas por Región / Comuna / Formato / Segmento ──
    if (regiones or comunas or formatos_t or segmentos) and not df_tiendas.empty:
        tf = df_tiendas.copy()
        if regiones   and "Región"   in tf.columns: tf = tf[tf["Región"].isin(set(regiones))]
        if comunas    and "Comuna"   in tf.columns: tf = tf[tf["Comuna"].isin(set(comunas))]
        if formatos_t and "Formato"  in tf.columns: tf = tf[tf["Formato"].isin(set(formatos_t))]
        if segmentos  and "Segmento" in tf.columns: tf = tf[tf["Segmento"].isin(set(segmentos))]
        valid_t = set(tf["FileCORR"].astype(str)) if "FileCORR" in tf.columns else set()
        if valid_t:
            v = v[v["Tiendas"].isin(valid_t)]

    # ── Filtro directo por código de tienda ──
    if tiendas_dir and "Tiendas" in v.columns:
        v = v[v["Tiendas"].isin(set(tiendas_dir))]

    p = dp.copy()
    if eff_ms  != set(meses_all): p = p[p["Anio_mes"].isin(eff_ms)]
    if cats_s  != set(cats_all):  p = p[p["Categoria"].isin(cats_s)]
    if prs: p = p[p["Nom_prod"].str.contains(prs, case=False, na=False)]
    if cs:  p = p[p["Cod_prod"].str.contains(cs, case=False, na=False)]
    # ── Filtro directo por producto ──
    if prods_dir and "Nom_prod" in p.columns:
        p = p[p["Nom_prod"].isin(set(prods_dir))]

    t = dt.copy()
    if eff_ms != set(meses_all): t = t[t["Anio_mes"].isin(eff_ms)]
    # Filtro de tiendas en tickets también
    if (regiones or comunas or formatos_t or segmentos) and not df_tiendas.empty:
        tf2 = df_tiendas.copy()
        if regiones   and "Región"   in tf2.columns: tf2 = tf2[tf2["Región"].isin(set(regiones))]
        if comunas    and "Comuna"   in tf2.columns: tf2 = tf2[tf2["Comuna"].isin(set(comunas))]
        if formatos_t and "Formato"  in tf2.columns: tf2 = tf2[tf2["Formato"].isin(set(formatos_t))]
        if segmentos  and "Segmento" in tf2.columns: tf2 = tf2[tf2["Segmento"].isin(set(segmentos))]
        valid_t2 = set(tf2["FileCORR"].astype(str)) if "FileCORR" in tf2.columns else set()
        if valid_t2:
            t = t[t["Tiendas"].isin(valid_t2)]
    if tiendas_dir and "Tiendas" in t.columns:
        t = t[t["Tiendas"].isin(set(tiendas_dir))]

    # ══════════════════════════════════════════════════════════
    #  CROSS-FILTERS BIDIRECCIONALES
    #  Garantizan que TODOS los filtros afecten TODOS los gráficos
    #
    #  Problema: ventas no tiene Nom_prod; productos no tiene Tiendas
    #  Solución: propagar via (Anio_mes, Categoria) como llave común
    #
    #  1) Tiendas → Productos: v filtrado → propagar a p (via Anio_mes+Cat)
    #  2) Producto → Ventas:   p filtrado → propagar a v (via Anio_mes+Cat)
    #  3) Ambos   → promo_prod: usar v final como referencia
    # ══════════════════════════════════════════════════════════

    # ── 1) Tiendas → Productos ──
    if regiones or comunas or formatos_t or segmentos or tiendas_dir:
        _xf_tienda = v[["Anio_mes","Categoria"]].drop_duplicates()
        if not _xf_tienda.empty and "Anio_mes" in p.columns and "Categoria" in p.columns:
            p = p.merge(_xf_tienda, on=["Anio_mes","Categoria"], how="inner")

    # ── 2) Producto → Ventas y Tickets ──
    if (prods_dir or prs or cs) and not p.empty:
        _xf_prod = p[["Anio_mes","Categoria"]].drop_duplicates()
        if not _xf_prod.empty:
            if "Anio_mes" in v.columns and "Categoria" in v.columns:
                v = v.merge(_xf_prod, on=["Anio_mes","Categoria"], how="inner")
            # Tickets no tiene Categoria → filtrar solo por meses
            _xf_prod_ms = _xf_prod[["Anio_mes"]].drop_duplicates()
            if "Anio_mes" in t.columns:
                t = t.merge(_xf_prod_ms, on="Anio_mes", how="inner")

    # ── 3) Filtrar promo_prod (pp) aplicando todos los cross-filters ──
    pp = dpp.copy() if not dpp.empty else pd.DataFrame()
    if not pp.empty:
        if eff_ms != set(meses_all) and "Anio_mes" in pp.columns:
            pp = pp[pp["Anio_mes"].isin(eff_ms)]
        if cats_s != set(cats_all) and "Categoria" in pp.columns:
            pp = pp[pp["Categoria"].isin(cats_s)]
        if prods_dir and "Nom_prod" in pp.columns:
            pp = pp[pp["Nom_prod"].isin(set(prods_dir))]
        # Propagar ambos cross-filters usando v ya totalmente filtrado
        if "Anio_mes" in pp.columns and "Categoria" in pp.columns and not v.empty:
            _xf_v = v[["Anio_mes","Categoria"]].drop_duplicates()
            if not _xf_v.empty:
                pp = pp.merge(_xf_v, on=["Anio_mes","Categoria"], how="inner")

    # ── DuckDB sobre datos ya filtrados (mucho más rápido) ──
    c = duckdb.connect()
    c.register("ventas",      v)
    c.register("productos",   p)
    c.register("tickets",     t)
    if not pp.empty:
        c.register("promo_prod", pp)
    def q(sql): return c.execute(sql).df()
    wp = "WHERE ventas.Cod_promo IS NOT NULL"

    D = {"_vf": v, "_pf": p, "_tf": t}

    # KPIs globales
    D["kpis_v"] = q("""
        SELECT SUM(Importe) AS ing, SUM(Cantidad) AS und,
               SUM(Costo_total) AS costo,
               SUM(Importe-Costo_total) AS margen,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS mg_pct,
               SUM(Precio_x_Cant) AS ing_lista,
               SUM(Precio_x_Cant-Importe) AS desc_total,
               (SUM(Precio_x_Cant)-SUM(Importe))/NULLIF(SUM(Precio_x_Cant),0) AS desc_pct,
               SUM(CASE WHEN ventas.Cod_promo IS NOT NULL THEN Importe ELSE 0 END) AS ing_promo,
               COUNT(DISTINCT ventas.Cod_promo) AS n_promos
        FROM ventas""").iloc[0]
    D["kpis_t"] = q("""
        SELECT SUM(n_tickets) AS tkt, SUM(n_tkt_promo) AS tkt_p,
               COUNT(DISTINCT tickets.Tiendas) AS tiendas
        FROM tickets""").iloc[0]

    # ── TAB 0: Resumen ──
    D["treemap"] = q("""
        SELECT COALESCE(Categoria,'Sin cat.') AS Categoria,
               SUM(Importe) AS Ingresos,
               SUM(Importe-Costo_total) AS Margen,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct
        FROM ventas GROUP BY 1 ORDER BY Ingresos DESC""")

    D["pareto"] = q("""
        SELECT COALESCE(NULLIF(TRIM(Nom_prod),''),Cod_prod,'?') AS Prod,
               SUM(Importe-Costo_total) AS Margen
        FROM productos
        GROUP BY 1 HAVING SUM(Importe-Costo_total) > 0
        ORDER BY Margen DESC LIMIT 150""")

    D["evol_v"] = q("""
        SELECT Anio_mes,
               SUM(Importe) AS Ingresos,
               SUM(Importe-Costo_total) AS Margen,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct,
               SUM(Precio_x_Cant-Importe) AS Descuento
        FROM ventas GROUP BY 1 ORDER BY 1""")
    D["evol_t"] = q("SELECT Anio_mes, SUM(n_tickets) AS Tickets FROM tickets GROUP BY 1")

    # ── TAB 1: Margen ──
    D["prod_mg"] = q("""
        SELECT COALESCE(NULLIF(TRIM(Nom_prod),''),Cod_prod,'?') AS Producto,
               Cod_prod AS Codigo,
               COALESCE(Categoria,'–') AS Categoria,
               SUM(Importe) AS Ingresos, SUM(Cantidad) AS Unidades,
               SUM(Costo_total) AS Costo,
               SUM(Importe-Costo_total) AS Margen,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Costo_total),0) AS Markup,
               SUM(Importe)/NULLIF(SUM(Cantidad),0) AS Precio_ef,
               SUM(Costo_total)/NULLIF(SUM(Cantidad),0) AS Costo_unit
        FROM productos
        GROUP BY 1,2,3 ORDER BY Ingresos DESC LIMIT 500""")

    # ── TAB 2: Precio & Costo ──
    D["precio_evol"] = q("""
        SELECT Anio_mes, COALESCE(Categoria,'Sin cat.') AS Categoria,
               SUM(Importe)/NULLIF(SUM(Cantidad),0) AS Precio_ef,
               SUM(Precio_x_Cant)/NULLIF(SUM(Cantidad),0) AS Precio_lista,
               SUM(Costo_total)/NULLIF(SUM(Cantidad),0) AS Costo_unit,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct,
               SUM(Importe) AS Ingresos, SUM(Cantidad) AS Unidades
        FROM ventas GROUP BY 1,2 ORDER BY 1,2""")

    D["pvol"] = q("""
        SELECT COALESCE(NULLIF(TRIM(Nom_prod),''),Cod_prod,'?') AS Producto,
               COALESCE(Categoria,'–') AS Categoria,
               SUM(Importe)/NULLIF(SUM(Cantidad),0) AS Precio_ef,
               SUM(Cantidad) AS Unidades, SUM(Importe) AS Ingresos,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct
        FROM productos GROUP BY 1,2 ORDER BY Ingresos DESC LIMIT 300""")

    try:
        D["compresion"] = q("""
        WITH base AS (
            SELECT COALESCE(NULLIF(TRIM(Nom_prod),''),Cod_prod,'?') AS Producto,
                   COALESCE(Categoria,'–') AS Categoria, Anio_mes,
                   SUM(Importe)/NULLIF(SUM(Cantidad),0) AS Precio_ef,
                   SUM(Costo_total)/NULLIF(SUM(Cantidad),0) AS Costo_unit,
                   (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct
            FROM productos GROUP BY 1,2,3
        ),
        fl AS (
            SELECT Producto, Categoria,
                   FIRST(Precio_ef  ORDER BY Anio_mes) AS Precio_inicio,
                   LAST(Precio_ef   ORDER BY Anio_mes) AS Precio_fin,
                   FIRST(Costo_unit ORDER BY Anio_mes) AS Costo_inicio,
                   LAST(Costo_unit  ORDER BY Anio_mes) AS Costo_fin,
                   FIRST(Mg_pct     ORDER BY Anio_mes) AS Mg_inicio,
                   LAST(Mg_pct      ORDER BY Anio_mes) AS Mg_fin,
                   COUNT(DISTINCT Anio_mes) AS n_meses
            FROM base GROUP BY 1,2
        )
        SELECT Producto, Categoria, n_meses,
               ROUND(Precio_inicio,0) AS Precio_inicio, ROUND(Precio_fin,0) AS Precio_fin,
               ROUND(Costo_inicio,0)  AS Costo_inicio,  ROUND(Costo_fin,0)  AS Costo_fin,
               ROUND((Precio_fin-Precio_inicio)/NULLIF(Precio_inicio,0)*100,1) AS Cambio_precio_pct,
               ROUND((Costo_fin-Costo_inicio)/NULLIF(Costo_inicio,0)*100,1)   AS Cambio_costo_pct,
               ROUND(Mg_inicio*100,1) AS Mg_inicio_pct, ROUND(Mg_fin*100,1)  AS Mg_fin_pct,
               ROUND(Mg_fin-Mg_inicio,3) AS Delta_mg
        FROM fl WHERE n_meses>=3 AND Costo_inicio>0 AND Precio_inicio>0
        ORDER BY Delta_mg ASC LIMIT 200""")
    except Exception:
        D["compresion"] = pd.DataFrame()

    # ── TAB 3: Promociones ──
    D["promos"] = q(f"""
        SELECT COALESCE(NULLIF(TRIM(Nom_promo),''),Cod_promo,'?') AS Promo,
               Cod_promo,
               MIN(Anio_mes) AS Primer_mes, MAX(Anio_mes) AS Ultimo_mes,
               COUNT(DISTINCT Anio_mes) AS Meses,
               COUNT(DISTINCT ventas.Tiendas) AS Tiendas,
               SUM(Importe) AS Ingresos, SUM(Cantidad) AS Unidades,
               SUM(Importe-Costo_total) AS Margen,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct,
               SUM(Precio_x_Cant-Importe) AS Descuento,
               (SUM(Precio_x_Cant)-SUM(Importe))/NULLIF(SUM(Precio_x_Cant),0) AS Desc_pct,
               SUM(Importe)/NULLIF(SUM(Precio_x_Cant-Importe),0) AS ROI
        FROM ventas {wp} GROUP BY 1,2 ORDER BY Ingresos DESC""")

    top20p = D["promos"]["Promo"].head(20).tolist()
    if top20p:
        lst20 = ",".join(f"'{x.replace(chr(39),chr(39)*2)}'" for x in top20p)
        D["heatmap_promo"] = q(f"""
            SELECT COALESCE(NULLIF(TRIM(Nom_promo),''),Cod_promo,'?') AS Promo,
                   Anio_mes, SUM(Importe) AS Ingresos
            FROM ventas WHERE Cod_promo IS NOT NULL
              AND COALESCE(NULLIF(TRIM(Nom_promo),''),Cod_promo,'?') IN ({lst20})
            GROUP BY 1,2 ORDER BY 1,2""")
    else:
        D["heatmap_promo"] = pd.DataFrame()

    D["mecanica"] = q(f"""
        SELECT
            CASE
              WHEN LOWER(Nom_promo) LIKE '%2x%' OR LOWER(Nom_promo) LIKE '%2 x%' THEN '2x1 / 2x precio'
              WHEN LOWER(Nom_promo) LIKE '%combo%' OR LOWER(Nom_promo) LIKE '%pack%' THEN 'Combo / Pack'
              WHEN LOWER(Nom_promo) LIKE '%gratis%' OR LOWER(Nom_promo) LIKE '%lleva%' THEN 'Lleva X Gratis'
              WHEN LOWER(Nom_promo) LIKE '%descuento%' OR LOWER(Nom_promo) LIKE '% % off%' THEN 'Descuento Directo'
              WHEN LOWER(Nom_promo) LIKE '%crm%' OR LOWER(Nom_promo) LIKE '%um%' THEN 'Promo CRM/Lealtad'
              ELSE 'Otras'
            END AS Mecanica,
            COUNT(DISTINCT Cod_promo) AS Promos, SUM(Importe) AS Ingresos,
            SUM(Importe-Costo_total) AS Margen,
            (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct,
            SUM(Precio_x_Cant-Importe) AS Inversion,
            SUM(Importe)/NULLIF(SUM(Precio_x_Cant-Importe),0) AS ROI
        FROM ventas {wp} GROUP BY 1 ORDER BY Ingresos DESC""")

    # ── TAB 4: Tiendas ──
    D["tiendas_v"] = q("""
        SELECT ventas.Tiendas,
               SUM(Importe) AS Ingresos, SUM(Cantidad) AS Unidades,
               SUM(Costo_total) AS Costo,
               SUM(Importe-Costo_total) AS Margen,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct,
               (SUM(Precio_x_Cant)-SUM(Importe))/NULLIF(SUM(Precio_x_Cant),0) AS Desc_pct,
               SUM(Precio_x_Cant)/NULLIF(SUM(Cantidad),0) AS Precio_lista,
               SUM(Importe)/NULLIF(SUM(Cantidad),0) AS Precio_ef,
               SUM(CASE WHEN Cod_promo IS NOT NULL THEN Importe ELSE 0 END)
                   /NULLIF(SUM(Importe),0) AS Mix_promo
        FROM ventas WHERE ventas.Tiendas IS NOT NULL
        GROUP BY 1 ORDER BY Ingresos DESC""")
    D["tiendas_t"] = q("""
        SELECT tickets.Tiendas, SUM(n_tickets) AS Tickets, SUM(n_tkt_promo) AS Tkt_promo
        FROM tickets GROUP BY 1""")

    top_tn = D["tiendas_v"]["Tiendas"].head(20).tolist()
    if top_tn:
        lst_tn = ",".join(repr(x) for x in top_tn)
        D["heat_tiendas"] = q(f"""
            SELECT ventas.Tiendas, Anio_mes, SUM(Importe) AS Ingresos
            FROM ventas WHERE ventas.Tiendas IN ({lst_tn})
            GROUP BY 1,2 ORDER BY 1,2""")
    else:
        D["heat_tiendas"] = pd.DataFrame()

    # ── TAB 5: Alertas ──
    D["alertas"] = q("""
        SELECT COALESCE(NULLIF(TRIM(Nom_prod),''),Cod_prod,'?') AS Producto,
               COALESCE(Categoria,'–') AS Categoria,
               SUM(Importe) AS Ingresos, SUM(Cantidad) AS Unidades,
               SUM(Importe-Costo_total) AS Margen,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct,
               SUM(Precio_x_Cant)/NULLIF(SUM(Cantidad),0) AS Precio_lista,
               SUM(Importe)/NULLIF(SUM(Cantidad),0) AS Precio_ef,
               SUM(Costo_total)/NULLIF(SUM(Cantidad),0) AS Costo_unit,
               (SUM(Precio_x_Cant)-SUM(Importe))/NULLIF(SUM(Precio_x_Cant),0) AS Desc_pct
        FROM productos GROUP BY 1,2 HAVING SUM(Importe)>100""")

    D["alertas_promo"] = q(f"""
        SELECT COALESCE(NULLIF(TRIM(Nom_promo),''),Cod_promo,'?') AS Promo,
               SUM(Importe) AS Ingresos,
               SUM(Precio_x_Cant-Importe) AS Inversion,
               SUM(Importe)/NULLIF(SUM(Precio_x_Cant-Importe),0) AS ROI,
               (SUM(Precio_x_Cant)-SUM(Importe))/NULLIF(SUM(Precio_x_Cant),0) AS Desc_pct,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct
        FROM ventas {wp} GROUP BY 1 HAVING SUM(Importe)>100""")

    # ── TICKET PROMO vs SIN PROMO (#8) ──
    D["ticket_promo_mes"] = q("""
        WITH ing AS (
            SELECT Anio_mes,
                   SUM(CASE WHEN Cod_promo IS NOT NULL THEN Importe ELSE 0 END) AS Ing_promo,
                   SUM(CASE WHEN Cod_promo IS NULL     THEN Importe ELSE 0 END) AS Ing_sin_promo,
                   SUM(Importe) AS Ing_total
            FROM ventas GROUP BY 1
        ),
        tkt AS (
            SELECT Anio_mes,
                   SUM(n_tickets)   AS Tkts_total,
                   SUM(n_tkt_promo) AS Tkts_promo
            FROM tickets GROUP BY 1
        )
        SELECT ing.Anio_mes,
               ing.Ing_promo, ing.Ing_sin_promo, ing.Ing_total,
               tkt.Tkts_total, tkt.Tkts_promo,
               (tkt.Tkts_total - tkt.Tkts_promo) AS Tkts_sin_promo,
               ing.Ing_promo    / NULLIF(tkt.Tkts_promo, 0)                         AS Tkt_con_promo,
               ing.Ing_sin_promo/ NULLIF(tkt.Tkts_total - tkt.Tkts_promo, 0)        AS Tkt_sin_promo,
               ing.Ing_total    / NULLIF(tkt.Tkts_total, 0)                         AS Tkt_promedio
        FROM ing JOIN tkt ON ing.Anio_mes = tkt.Anio_mes
        ORDER BY ing.Anio_mes""")

    D["ticket_promo_cat"] = q("""
        WITH ing AS (
            SELECT COALESCE(Categoria,'–') AS Categoria,
                   SUM(CASE WHEN Cod_promo IS NOT NULL THEN Importe ELSE 0 END) AS Ing_promo,
                   SUM(CASE WHEN Cod_promo IS NULL     THEN Importe ELSE 0 END) AS Ing_sin_promo,
                   SUM(Importe) AS Ing_total
            FROM ventas GROUP BY 1
        ),
        tkt AS (
            SELECT SUM(n_tickets)   AS Tkts_total,
                   SUM(n_tkt_promo) AS Tkts_promo
            FROM tickets
        )
        SELECT ing.Categoria,
               ing.Ing_promo, ing.Ing_sin_promo, ing.Ing_total,
               ing.Ing_promo    / NULLIF((SELECT SUM(n_tkt_promo) FROM tickets), 0) AS Tkt_con_promo_ref,
               (SUM(ing.Ing_promo) OVER ()) AS Ing_promo_total,
               ing.Ing_promo / NULLIF(SUM(ing.Ing_promo) OVER (), 0)               AS Mix_cat_promo
        FROM ing ORDER BY Ing_promo DESC""")

    # ── SKUs POR PROMO (#3) ──
    D["sku_por_promo"] = q(f"""
        SELECT COALESCE(NULLIF(TRIM(Nom_promo),''),Cod_promo,'?') AS Promo,
               Cod_promo,
               COALESCE(Categoria,'–') AS Categoria,
               COUNT(DISTINCT ventas.Tiendas) AS Tiendas,
               SUM(Importe) AS Ingresos,
               SUM(Cantidad) AS Unidades,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct,
               (SUM(Precio_x_Cant)-SUM(Importe))/NULLIF(SUM(Precio_x_Cant),0) AS Desc_pct,
               SUM(Importe)/NULLIF(SUM(Precio_x_Cant-Importe),0) AS ROI,
               MIN(Anio_mes) AS Primer_mes, MAX(Anio_mes) AS Ultimo_mes
        FROM ventas {wp}
        GROUP BY 1,2,3 ORDER BY Ingresos DESC""")

    D["promo_por_cat_ciclo"] = q(f"""
        SELECT COALESCE(Categoria,'–') AS Categoria,
               COALESCE(NULLIF(TRIM(Nom_promo),''),Cod_promo,'?') AS Promo,
               Anio_mes,
               SUM(Importe)   AS Ingresos,
               SUM(Cantidad)  AS Unidades,
               COUNT(DISTINCT ventas.Tiendas) AS Tiendas,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct
        FROM ventas {wp}
        GROUP BY 1,2,3 ORDER BY 1,2,3""")

    # ── PENETRACIÓN PROMO ──
    if not pp.empty:
        D["penetracion_prod"] = q("""
            SELECT COALESCE(NULLIF(TRIM(Nom_prod),''),Cod_prod,'?') AS Producto,
                   COALESCE(Categoria,'–') AS Categoria,
                   SUM(CASE WHEN has_promo=1 THEN Cantidad ELSE 0 END) AS Cant_promo,
                   SUM(CASE WHEN has_promo=0 THEN Cantidad ELSE 0 END) AS Cant_sin_promo,
                   SUM(Cantidad)                                         AS Cant_total,
                   SUM(CASE WHEN has_promo=1 THEN Cantidad ELSE 0 END)
                       / NULLIF(SUM(Cantidad),0)                         AS Penetracion,
                   SUM(CASE WHEN has_promo=1 THEN Importe ELSE 0 END)   AS Ing_promo,
                   SUM(CASE WHEN has_promo=0 THEN Importe ELSE 0 END)   AS Ing_sin_promo,
                   SUM(Importe)                                           AS Ingresos,
                   SUM(CASE WHEN has_promo=1 THEN Costo_total ELSE 0 END) AS Costo_promo,
                   SUM(CASE WHEN has_promo=0 THEN Costo_total ELSE 0 END) AS Costo_sin_promo
            FROM promo_prod
            GROUP BY 1,2
            HAVING SUM(Cantidad) > 0
            ORDER BY Ingresos DESC""")

        D["penetracion_mes"] = q("""
            SELECT Anio_mes,
                   SUM(CASE WHEN has_promo=1 THEN Cantidad ELSE 0 END) AS Cant_promo,
                   SUM(CASE WHEN has_promo=0 THEN Cantidad ELSE 0 END) AS Cant_sin_promo,
                   SUM(Cantidad)                                         AS Cant_total,
                   SUM(CASE WHEN has_promo=1 THEN Cantidad ELSE 0 END)
                       / NULLIF(SUM(Cantidad),0)                         AS Penetracion,
                   SUM(CASE WHEN has_promo=1 THEN Importe ELSE 0 END)   AS Ing_promo,
                   SUM(CASE WHEN has_promo=0 THEN Importe ELSE 0 END)   AS Ing_sin_promo,
                   SUM(Importe)                                           AS Ingresos
            FROM promo_prod
            GROUP BY 1 ORDER BY 1""")

        D["penetracion_cat"] = q("""
            SELECT COALESCE(Categoria,'–') AS Categoria,
                   SUM(CASE WHEN has_promo=1 THEN Cantidad ELSE 0 END) AS Cant_promo,
                   SUM(CASE WHEN has_promo=0 THEN Cantidad ELSE 0 END) AS Cant_sin_promo,
                   SUM(Cantidad)                                         AS Cant_total,
                   SUM(CASE WHEN has_promo=1 THEN Cantidad ELSE 0 END)
                       / NULLIF(SUM(Cantidad),0)                         AS Penetracion,
                   SUM(CASE WHEN has_promo=1 THEN Importe ELSE 0 END)
                       / NULLIF(SUM(Importe),0)                          AS Penetracion_ing,
                   SUM(Importe) AS Ingresos
            FROM promo_prod
            GROUP BY 1 ORDER BY Penetracion DESC""")

        D["penetracion_cat_mes"] = q("""
            SELECT COALESCE(Categoria,'–') AS Categoria, Anio_mes,
                   SUM(CASE WHEN has_promo=1 THEN Cantidad ELSE 0 END)
                       / NULLIF(SUM(Cantidad),0) AS Penetracion
            FROM promo_prod
            GROUP BY 1,2 ORDER BY 1,2""")
    else:
        D["penetracion_prod"]    = pd.DataFrame()
        D["penetracion_mes"]     = pd.DataFrame()
        D["penetracion_cat"]     = pd.DataFrame()
        D["penetracion_cat_mes"] = pd.DataFrame()

    # ── TAB 6: CRUCES ──
    D["mg_cat_mes"] = q("""
        SELECT COALESCE(Categoria,'Sin cat.') AS Categoria, Anio_mes,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct,
               SUM(Importe) AS Ingresos
        FROM ventas GROUP BY 1,2 ORDER BY 1,2""")

    D["desc_mg_prod"] = q("""
        SELECT COALESCE(NULLIF(TRIM(Nom_prod),''),Cod_prod,'?') AS Producto,
               COALESCE(Categoria,'–') AS Categoria,
               (SUM(Precio_x_Cant)-SUM(Importe))/NULLIF(SUM(Precio_x_Cant),0) AS Desc_pct,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0)          AS Mg_pct,
               SUM(Importe) AS Ingresos, SUM(Cantidad) AS Unidades
        FROM productos GROUP BY 1,2 HAVING SUM(Importe)>10000 ORDER BY Ingresos DESC""")

    D["cat_promo_cross"] = q("""
        SELECT COALESCE(Categoria,'Sin cat.') AS Categoria,
               SUM(Importe) AS Ingresos,
               SUM(CASE WHEN Cod_promo IS NOT NULL THEN Importe ELSE 0 END)
                   /NULLIF(SUM(Importe),0) AS Mix_promo,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct,
               (SUM(Precio_x_Cant)-SUM(Importe))/NULLIF(SUM(Precio_x_Cant),0) AS Desc_pct
        FROM ventas GROUP BY 1 ORDER BY Ingresos DESC""")

    D["precio_gap"] = q("""
        SELECT Anio_mes, COALESCE(Categoria,'Sin cat.') AS Categoria,
               SUM(Precio_x_Cant)/NULLIF(SUM(Cantidad),0) AS Precio_lista,
               SUM(Importe)/NULLIF(SUM(Cantidad),0)        AS Precio_ef,
               (SUM(Precio_x_Cant)-SUM(Importe))/NULLIF(SUM(Precio_x_Cant),0) AS Gap_pct,
               SUM(Importe) AS Ingresos
        FROM ventas GROUP BY 1,2 ORDER BY 1,2""")

    # ── YoY (Año vs Año) ──
    D["yoy_mes"] = q("""
        SELECT Anio_mes,
               SUBSTR(Anio_mes,1,4) AS Año,
               SUBSTR(Anio_mes,6,2) AS Mes,
               SUM(Importe)                AS Ingresos,
               SUM(Cantidad)               AS Unidades,
               SUM(Importe-Costo_total)    AS Margen,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct,
               SUM(CASE WHEN Cod_promo IS NOT NULL THEN Importe ELSE 0 END)
                   /NULLIF(SUM(Importe),0) AS Mix_promo
        FROM ventas GROUP BY 1,2,3 ORDER BY 1""")

    D["yoy_cat"] = q("""
        SELECT COALESCE(Categoria,'–') AS Categoria,
               SUBSTR(Anio_mes,1,4) AS Año,
               SUM(Importe)                AS Ingresos,
               SUM(Cantidad)               AS Unidades,
               SUM(Importe-Costo_total)    AS Margen,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct
        FROM ventas GROUP BY 1,2 ORDER BY 1,2""")

    # ── SKU Ranking completo ──
    D["sku_ranking"] = q("""
        SELECT COALESCE(NULLIF(TRIM(Nom_prod),''),Cod_prod,'?') AS Producto,
               COALESCE(Categoria,'–') AS Categoria,
               SUM(Importe)                AS Ingresos,
               SUM(Cantidad)               AS Unidades,
               SUM(Importe-Costo_total)    AS Margen,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct,
               SUM(Precio_x_Cant)/NULLIF(SUM(Cantidad),0)             AS Precio_lista,
               SUM(Importe)/NULLIF(SUM(Cantidad),0)                   AS Precio_ef,
               (SUM(Precio_x_Cant)-SUM(Importe))/NULLIF(SUM(Precio_x_Cant),0) AS Desc_pct,
               MIN(Anio_mes) AS Primer_mes, MAX(Anio_mes) AS Ultimo_mes
        FROM productos GROUP BY 1,2 HAVING SUM(Importe)>0 ORDER BY Ingresos DESC LIMIT 500""")

    # ── Price Ladder por categoría ──
    D["price_ladder"] = q("""
        SELECT COALESCE(NULLIF(TRIM(Nom_prod),''),Cod_prod,'?') AS Producto,
               COALESCE(Categoria,'–') AS Categoria,
               SUM(Importe)/NULLIF(SUM(Cantidad),0)       AS Precio_ef,
               SUM(Precio_x_Cant)/NULLIF(SUM(Cantidad),0) AS Precio_lista,
               SUM(Importe)                               AS Ingresos,
               SUM(Cantidad)                              AS Unidades,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct
        FROM productos GROUP BY 1,2 HAVING SUM(Importe)>10000 ORDER BY 2,3""")

    # ── Cobertura de tiendas por categoría (ventas tiene Tiendas, productos no) ──
    D["cobertura_prod"] = q("""
        SELECT COALESCE(Categoria,'–') AS Categoria,
               COUNT(DISTINCT ventas.Tiendas) AS N_tiendas,
               SUM(Importe)                   AS Ingresos,
               SUM(Cantidad)                  AS Unidades,
               (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct,
               (SUM(Precio_x_Cant)-SUM(Importe))/NULLIF(SUM(Precio_x_Cant),0) AS Desc_pct
        FROM ventas GROUP BY 1 HAVING SUM(Importe)>0 ORDER BY N_tiendas DESC""")

    # ── TAB 8 (legacy Datos) ──
    D["resumen_export"] = q("""
        SELECT Anio_mes, ventas.Tiendas,
               COALESCE(Categoria,'–') AS Categoria,
               COALESCE(Tipo,'–') AS Tipo,
               COALESCE(Sistema,'–') AS Sistema,
               SUM(Importe) AS Ingresos, SUM(Cantidad) AS Unidades,
               SUM(Costo_total) AS Costo,
               SUM(Importe-Costo_total) AS Margen,
               ROUND((SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0)*100,2) AS Margen_pct,
               SUM(Precio_x_Cant-Importe) AS Descuento,
               COUNT(DISTINCT Cod_promo) AS Promos_distintas
        FROM ventas GROUP BY 1,2,3,4,5 ORDER BY 1,2,3""")

    c.close()
    return D


# ── Obtener todos los datos (cached por combo de filtros) ──
D = all_data(filter_key)

# Desempaquetar KPIs globales
row     = D["kpis_v"]
tkt_row = D["kpis_t"]
ing       = s(row["ing"])
costo     = s(row["costo"])
margen    = s(row["margen"])
mg_pct    = s(row["mg_pct"])
desc_t    = s(row["desc_total"])
desc_pct  = s(row["desc_pct"])
ing_promo = s(row["ing_promo"])
n_promo   = s(row["n_promos"])
und       = s(row["und"])
total_tkt = s(tkt_row["tkt"])
n_tiendas = s(tkt_row["tiendas"])
tkt_prom  = ing/total_tkt if total_tkt else 0

# DataFrames desempaquetados
df_ev_v = D["evol_v"]
df_ev_t = D["evol_t"]
df_evol = df_ev_v.merge(df_ev_t, on="Anio_mes", how="left")
df_tm   = D["treemap"]
df_par  = D["pareto"]
df_prod = D["prod_mg"]
df_pev  = D["precio_evol"]
df_pvol = D["pvol"]
df_cmp  = D["compresion"]
df_ap   = D["promos"]
df_hm   = D["heatmap_promo"]
df_mec  = D["mecanica"]
df_tn_v = D["tiendas_v"]
df_tn_t = D["tiendas_t"]
df_ht   = D["heat_tiendas"]
df_al   = D["alertas"]
df_alp  = D["alertas_promo"]
df_res  = D["resumen_export"]
df_mg_cat_mes   = D["mg_cat_mes"]
df_desc_mg      = D["desc_mg_prod"]
df_cat_promo    = D["cat_promo_cross"]
df_precio_gap   = D["precio_gap"]
df_pen_prod     = D["penetracion_prod"]
df_pen_mes      = D["penetracion_mes"]
df_pen_cat      = D["penetracion_cat"]
df_pen_cat_mes  = D["penetracion_cat_mes"]
df_tkt_mes      = D["ticket_promo_mes"]
df_tkt_cat      = D["ticket_promo_cat"]
df_sku_promo    = D["sku_por_promo"]
df_promo_ciclo  = D["promo_por_cat_ciclo"]
df_yoy_mes      = D["yoy_mes"]
df_yoy_cat      = D["yoy_cat"]
df_sku_ranking  = D["sku_ranking"]
df_price_ladder = D["price_ladder"]
df_cobertura    = D["cobertura_prod"]
vf      = D["_vf"]
pf      = D["_pf"]


# ══════════════════════════════════════════════════════════════════
#  HEADER
# ══════════════════════════════════════════════════════════════════
# Calcular meses efectivos para el header
_eff_ms = set(_ms) if _ms else set(meses_all)
if _ciclo_años:
    _cy = set()
    for _ca in _ciclo_años: _cy.update(CICLO_MESES.get(_ca, set()))
    _eff_ms &= _cy
if _ciclo_nums:
    _cn = set()
    for _ca, _mset in CICLO_MESES.items():
        if _ca.split("-")[0] in set(_ciclo_nums): _cn.update(_mset)
    _eff_ms &= _cn
if _años:
    _eff_ms = {m for m in _eff_ms if m[:4] in set(_años)}
_eff_ms_sorted = sorted(_eff_ms)
rng = (f"{_eff_ms_sorted[0]} → {_eff_ms_sorted[-1]}" if _eff_ms_sorted else "–")
_ciclo_badge = ""
if _ciclo_años:  _ciclo_badge += " · " + ", ".join(_ciclo_años)
elif _ciclo_nums: _ciclo_badge += " · " + ", ".join(_ciclo_nums)
st.markdown(f"""
<div class="enex-header">
  <div>
    <div class="enex-title">⛽ Pricing Intelligence · ENEX S.A.</div>
    <div class="enex-sub">Gerencia Comercial · {len(_eff_ms_sorted)} meses · {rng}{_ciclo_badge}</div>
  </div>
  <div class="enex-badge">{len(_cats)} categorías · {int(n_tiendas)} tiendas</div>
</div>""", unsafe_allow_html=True)

k = st.columns(8)
kpi(k[0], clp(ing),       "Ingresos Totales")
kpi(k[1], clp(margen),    "Margen Bruto $", "g" if margen>0 else "r")
kpi(k[2], pct(mg_pct),    "Margen %",       "g" if mg_pct>=0.15 else ("a" if mg_pct>0 else "r"))
kpi(k[3], pct(1-s(row["costo"])/ing if ing else 0), "Markup Bruto", "g")
kpi(k[4], num(total_tkt), "Tickets")
kpi(k[5], clp(tkt_prom),  "Ticket Promedio")
kpi(k[6], pct(ing_promo/ing if ing else 0), "Mix Promo", "a")
kpi(k[7], pct(desc_pct),  "Descuento %",    "a")


# ══════════════════════════════════════════════════════════════════
#  TABS
# ══════════════════════════════════════════════════════════════════
T = st.tabs([
    "📊 Resumen","💰 Margen","📈 Precio & Costo",
    "🎯 Promociones","🏪 Tiendas","🔔 Alertas","🔀 Cruces","📣 Análisis Promo",
    "📅 YoY","🏆 Ranking SKU","📋 Datos"
])


# ══════════════════════════════════════════════════ TAB 0: RESUMEN
with T[0]:
    c1,c2 = st.columns([3,2])
    with c1:
        if not df_evol.empty:
            _ev = df_evol.copy()
            _ev["Ing_lbl"]  = _ev["Ingresos"].apply(lambda x: _fmt_clp(x))
            _ev["Mg_lbl"]   = _ev["Mg_pct"].apply(lambda x: f"{x*100:.1f}%")
            fig = go.Figure()
            fig.add_trace(go.Bar(
                name="Ingresos", x=_ev["Anio_mes"], y=_ev["Ingresos"],
                marker_color=C_NAVY, opacity=0.9,
                hovertemplate="<b>%{x}</b><br>Ingresos: $%{y:,.0f}<extra></extra>"
            ))
            fig.add_trace(go.Bar(
                name="Margen $", x=_ev["Anio_mes"], y=_ev["Margen"],
                marker_color=C_GREEN, opacity=0.85,
                hovertemplate="<b>%{x}</b><br>Margen: $%{y:,.0f}<extra></extra>"
            ))
            fig.add_trace(go.Scatter(
                name="Margen %", x=_ev["Anio_mes"], y=_ev["Mg_pct"],
                mode="lines+markers+text",
                text=_ev["Mg_lbl"], textposition="top center", textfont=dict(size=9, color=C_AMBER),
                line=dict(color=C_AMBER, width=2.5), marker=dict(size=6),
                yaxis="y2",
                hovertemplate="<b>%{x}</b><br>Margen %%: %{y:.1%}<extra></extra>"
            ))
            fig.update_layout(
                barmode="overlay",
                yaxis=dict(title="Ingresos / Margen ($)", tickformat=",.0f",
                           gridcolor="#F3F4F6", tickprefix="$"),
                yaxis2=dict(title="Margen %", overlaying="y", side="right",
                            tickformat=".0%", showgrid=False),
                legend=dict(orientation="h", y=1.12, x=0, font_size=11)
            )
            sec("Evolución mensual de Ingresos y Margen")
            st.plotly_chart(pcfg(fig, 310, legend=False), use_container_width=True)

    with c2:
        if not df_tm.empty:
            _tm = df_tm.copy()
            _tm["texto"] = _tm.apply(
                lambda r: f"{r['Categoria']}<br>{r['Mg_pct']*100:.1f}% margen", axis=1)
            fig2 = px.treemap(
                _tm, path=["Categoria"], values="Ingresos",
                color="Mg_pct",
                color_continuous_scale=["#C8102E","#F5A623","#1A9E5C"],
                custom_data=["Margen","Mg_pct"]
            )
            fig2.update_traces(
                texttemplate="<b>%{label}</b><br>%{customdata[1]:.1%} margen",
                textfont_size=11,
                hovertemplate="<b>%{label}</b><br>"
                              "Ingresos: $%{value:,.0f}<br>"
                              "Margen: $%{customdata[0]:,.0f}<br>"
                              "Margen %%: %{customdata[1]:.1%}<extra></extra>"
            )
            fig2.update_layout(
                coloraxis_colorbar=dict(title="Margen %", tickformat=".0%", len=0.6)
            )
            sec("Mix de Ingresos por Categoría")
            st.plotly_chart(pcfg(fig2, 310, legend=False), use_container_width=True)

    if not df_par.empty:
        df_par = df_par.copy()
        df_par["Pct"]     = df_par["Margen"] / df_par["Margen"].sum()
        df_par["Cum_pct"] = df_par["Pct"].cumsum()
        df_par["Rank"]    = range(1, len(df_par)+1)
        cut80 = int(df_par[df_par["Cum_pct"]<=0.80]["Rank"].max()) \
                if not df_par[df_par["Cum_pct"]<=0.80].empty else len(df_par)
        pct_skus = cut80 / len(df_par) * 100
        fig3 = go.Figure()
        fig3.add_trace(go.Bar(
            x=df_par["Rank"], y=df_par["Pct"],
            marker_color=C_NAVY, marker_line_width=0, name="% Margen individual",
            opacity=0.8,
            hovertemplate="Producto #%{x}<br>Aporte al margen: %{y:.2%}<extra></extra>"
        ))
        fig3.add_trace(go.Scatter(
            x=df_par["Rank"], y=df_par["Cum_pct"],
            line=dict(color=C_AMBER, width=2.5), name="% Acumulado",
            yaxis="y2",
            hovertemplate="Producto #%{x}<br>Acumulado: %{y:.1%}<extra></extra>"
        ))
        fig3.add_vline(x=cut80, line_dash="dash", line_color=C_RED, line_width=1.5,
                       annotation_text=f"  {cut80} SKUs = 80% del margen ({pct_skus:.0f}% del catálogo)",
                       annotation_font_size=10, annotation_font_color=C_RED)
        fig3.add_hline(y=0.80, line_dash="dot", line_color=C_RED, line_width=1,
                       yref="y2")
        fig3.update_layout(
            yaxis=dict(tickformat=".1%", gridcolor="#F3F4F6", title="Aporte al margen"),
            yaxis2=dict(overlaying="y", side="right", tickformat=".0%",
                        title="% Acumulado", range=[0,1.05]),
            xaxis_title="Ranking de productos por margen (mayor a menor)",
            legend=dict(orientation="h", y=1.12)
        )
        sec(f"Análisis Pareto — {cut80} SKUs generan el 80% del margen ({pct_skus:.0f}% del catálogo)")
        st.plotly_chart(pcfg(fig3, 260), use_container_width=True)

    sec("Resumen por Categoría")
    if not df_tm.empty:
        d = df_tm.copy().sort_values("Ingresos", ascending=False)
        d["% del total"] = (d["Ingresos"] / d["Ingresos"].sum()).apply(lambda x: f"{x*100:.1f}%")
        d["Ingresos"]    = d["Ingresos"].apply(clp)
        d["Margen"]      = d["Margen"].apply(clp)
        d["Margen %"]    = d["Mg_pct"].apply(lambda x: f"{x*100:.1f}%")
        d["Estado"]      = d["Mg_pct"].apply(lambda x: "🟢 Saludable" if x>=0.20 else ("🟡 Ajustado" if x>=0.10 else "🔴 Crítico"))
        st.dataframe(d[["Estado","Categoria","Ingresos","% del total","Margen","Margen %"]],
                     use_container_width=True, hide_index=True, height=240)


# ══════════════════════════════════════════════════ TAB 1: MARGEN
with T[1]:
    if df_prod.empty:
        st.info("Sin datos con los filtros actuales.")
    else:
        top15 = df_prod.nlargest(top_n, "Mg_pct").copy()
        bot15 = df_prod[df_prod["Ingresos"]>1000].nsmallest(top_n, "Mg_pct").copy()

        c1,c2 = st.columns(2)
        with c1:
            top15["lbl"] = top15.apply(lambda r: f"{r['Mg_pct']*100:.1f}%  |  {clp(r['Ingresos'])}", axis=1)
            fig = px.bar(top15, x="Mg_pct", y="Producto", orientation="h",
                         color="Mg_pct", color_continuous_scale=["#F5A623","#1A9E5C"],
                         text="lbl",
                         custom_data=["Ingresos","Margen","Categoria"])
            fig.update_traces(
                textposition="outside", textfont_size=9.5,
                marker_line_width=0,
                hovertemplate=(
                    "<b>%{y}</b><br>"
                    "Categoría: %{customdata[2]}<br>"
                    "Margen %%: %{x:.1%}<br>"
                    "Ingresos: $%{customdata[0]:,.0f}<br>"
                    "Margen $: $%{customdata[1]:,.0f}<extra></extra>"
                )
            )
            fig.update_layout(yaxis=dict(autorange="reversed"),
                              xaxis=dict(tickformat=".0%", title="Margen %"),
                              coloraxis_showscale=False)
            sec(f"Top {top_n} productos con mayor margen %")
            st.plotly_chart(pcfg(fig, max(300,top_n*28), legend=False),
                            use_container_width=True)

        with c2:
            if not bot15.empty:
                bot15["lbl"] = bot15.apply(lambda r: f"{r['Mg_pct']*100:.1f}%  |  {clp(r['Ingresos'])}", axis=1)
                fig2 = px.bar(bot15, x="Mg_pct", y="Producto", orientation="h",
                              color="Mg_pct", color_continuous_scale=["#C8102E","#F5A623"],
                              text="lbl",
                              custom_data=["Ingresos","Margen","Categoria"])
                fig2.update_traces(
                    textposition="outside", textfont_size=9.5,
                    marker_line_width=0,
                    hovertemplate=(
                        "<b>%{y}</b><br>"
                        "Categoría: %{customdata[2]}<br>"
                        "Margen %%: %{x:.1%}<br>"
                        "Ingresos: $%{customdata[0]:,.0f}<br>"
                        "Margen $: $%{customdata[1]:,.0f}<extra></extra>"
                    )
                )
                fig2.update_layout(yaxis=dict(autorange="reversed"),
                                   xaxis=dict(tickformat=".0%", title="Margen %"),
                                   coloraxis_showscale=False)
                sec(f"Bottom {top_n} productos con menor margen % (ingresos >$1.000)")
                st.plotly_chart(pcfg(fig2, max(300,top_n*28), legend=False),
                                use_container_width=True)

        df_sc = df_prod[df_prod["Ingresos"]>0].copy()
        med_und = df_sc["Unidades"].median()
        med_mg  = df_sc["Mg_pct"].median()
        df_sc["Cuadrante"] = df_sc.apply(lambda r:
            "Alto margen · alta rotación 🎯" if r["Mg_pct"]>=med_mg and r["Unidades"]>=med_und
            else ("Alto margen · baja rotación 💡" if r["Mg_pct"]>=med_mg and r["Unidades"]<med_und
            else ("Bajo margen · alta rotación ⚠️" if r["Mg_pct"]<med_mg and r["Unidades"]>=med_und
            else "Bajo margen · baja rotación 🔴")), axis=1)
        cuad_colors = {
            "Alto margen · alta rotación 🎯": C_GREEN,
            "Alto margen · baja rotación 💡": C_BLUE2,
            "Bajo margen · alta rotación ⚠️": C_AMBER,
            "Bajo margen · baja rotación 🔴": C_ALERT
        }
        fig3 = px.scatter(
            df_sc, x="Unidades", y="Mg_pct", size=sz(df_sc["Ingresos"]),
            hover_name="Producto", color="Cuadrante",
            color_discrete_map=cuad_colors, size_max=50, opacity=0.8,
            custom_data=["Ingresos","Margen","Categoria"]
        )
        fig3.update_traces(
            hovertemplate=(
                "<b>%{hovertext}</b><br>"
                "Categoría: %{customdata[2]}<br>"
                "Margen %%: %{y:.1%}<br>"
                "Unidades: %{x:,.0f}<br>"
                "Ingresos: $%{customdata[0]:,.0f}<extra></extra>"
            )
        )
        fig3.add_hline(y=med_mg, line_dash="dot", line_color="#9CA3AF", line_width=1,
                       annotation_text=f"  Mediana margen: {med_mg*100:.1f}%",
                       annotation_font_size=10, annotation_font_color=C_GRAY)
        fig3.add_vline(x=med_und, line_dash="dot", line_color="#9CA3AF", line_width=1,
                       annotation_text=f"  Mediana vol.: {med_und:,.0f} un.",
                       annotation_font_size=10, annotation_font_color=C_GRAY)
        # Etiquetas de cuadrantes
        fig3.add_annotation(x=df_sc["Unidades"].max()*0.95, y=df_sc["Mg_pct"].max()*0.95,
            text="🎯 Estrella", showarrow=False, font=dict(color=C_GREEN, size=10))
        fig3.add_annotation(x=df_sc["Unidades"].max()*0.95, y=med_mg*0.3,
            text="⚠️ Vacas lecheras a revisar", showarrow=False, font=dict(color=C_AMBER, size=10))
        fig3.update_yaxes(tickformat=".0%", title="Margen %")
        fig3.update_xaxes(title="Unidades vendidas")
        sec("Diagnóstico de portafolio — Margen % vs Volumen (burbuja = ingresos)")
        st.plotly_chart(pcfg(fig3, 400),
                        use_container_width=True)

        sec("Ranking completo de productos")
        df_show = df_prod.head(top_n*3).copy()
        df_show["#"]         = range(1, len(df_show)+1)
        df_show["Margen %"]  = df_show["Mg_pct"].apply(lambda x: f"{x*100:.1f}%")
        df_show["Markup"]    = df_show["Markup"].apply(lambda x: f"{x*100:.1f}%" if pd.notna(x) else "–")
        df_show["Ingresos"]  = df_show["Ingresos"].apply(clp)
        df_show["Margen $"]  = df_show["Margen"].apply(clp)
        df_show["Precio ef."]= df_show["Precio_ef"].apply(lambda x: f"${x:,.0f}".replace(",",".") if pd.notna(x) else "–")
        df_show["Costo u."]  = df_show["Costo_unit"].apply(lambda x: f"${x:,.0f}".replace(",",".") if pd.notna(x) else "–")
        df_show["Estado"]    = df_show["Mg_pct"].apply(lambda x: "🔴" if x<0 else ("🟡" if x<0.10 else "🟢"))
        st.dataframe(df_show[["Estado","#","Producto","Categoria","Ingresos","Margen $","Margen %","Precio ef.","Costo u.","Markup"]],
                     use_container_width=True, hide_index=True, height=340)


# ══════════════════════════════════════════════ TAB 2: PRECIO & COSTO
with T[2]:
    if not df_pev.empty:
        cats_ev = sorted(df_pev["Categoria"].unique().tolist())
        cat_sel_ev = st.selectbox("Categoría para análisis de precio:", ["(Todas)"] + cats_ev, key="cat_ev")
        df_pev_f = df_pev if cat_sel_ev=="(Todas)" else df_pev[df_pev["Categoria"]==cat_sel_ev]
        df_agg_ev = df_pev_f.groupby("Anio_mes").agg(
            Precio_lista=("Precio_lista","mean"),
            Precio_ef=("Precio_ef","mean"),
            Costo_unit=("Costo_unit","mean"),
            Mg_pct=("Mg_pct","mean"),
            Unidades=("Unidades","sum")
        ).reset_index()

        c1,c2 = st.columns(2)
        with c1:
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                name="Precio Lista", x=df_agg_ev["Anio_mes"], y=df_agg_ev["Precio_lista"],
                mode="lines+markers", line=dict(color=C_NAVY, width=2.5), marker=dict(size=6),
                hovertemplate="<b>%{x}</b><br>Precio Lista: $%{y:,.0f}<extra></extra>"
            ))
            fig.add_trace(go.Scatter(
                name="Precio Efectivo (con descuentos)", x=df_agg_ev["Anio_mes"], y=df_agg_ev["Precio_ef"],
                mode="lines+markers", line=dict(color=C_BLUE2, width=2, dash="dot"), marker=dict(size=6),
                hovertemplate="<b>%{x}</b><br>Precio Efectivo: $%{y:,.0f}<extra></extra>"
            ))
            fig.add_trace(go.Scatter(
                name="Costo Unitario", x=df_agg_ev["Anio_mes"], y=df_agg_ev["Costo_unit"],
                mode="lines+markers", line=dict(color=C_RED, width=2), marker=dict(size=6),
                fill="tonexty" if len(df_agg_ev)>1 else None,
                fillcolor="rgba(200,16,46,0.06)",
                hovertemplate="<b>%{x}</b><br>Costo Unitario: $%{y:,.0f}<extra></extra>"
            ))
            fig.update_yaxes(title="Precio / Costo promedio ponderado ($)", tickprefix="$", tickformat=",.0f")
            fig.update_xaxes(title="Mes")
            sec("Evolución: Precio Lista · Precio Efectivo · Costo Unitario")
            st.plotly_chart(pcfg(fig, 310),
                            use_container_width=True)

        with c2:
            fig2 = go.Figure()
            fig2.add_trace(go.Scatter(
                name="Margen %", x=df_agg_ev["Anio_mes"], y=df_agg_ev["Mg_pct"],
                mode="lines+markers+text",
                text=df_agg_ev["Mg_pct"].apply(lambda x: f"{x*100:.1f}%"),
                textposition="top center", textfont=dict(size=9, color=C_GREEN),
                line=dict(color=C_GREEN, width=2.5), marker=dict(size=7),
                yaxis="y",
                hovertemplate="<b>%{x}</b><br>Margen %%: %{y:.1%}<extra></extra>"
            ))
            fig2.add_trace(go.Bar(
                name="Unidades vendidas", x=df_agg_ev["Anio_mes"], y=df_agg_ev["Unidades"],
                marker_color=C_NAVY, opacity=0.2, yaxis="y2",
                hovertemplate="<b>%{x}</b><br>Unidades: %{y:,.0f}<extra></extra>"
            ))
            fig2.update_layout(
                yaxis=dict(title="Margen %", tickformat=".1%", gridcolor="#F3F4F6"),
                yaxis2=dict(title="Unidades vendidas", overlaying="y", side="right",
                            showgrid=False, tickformat=",.0f")
            )
            sec("Evolución de Margen % y Volumen mensual")
            st.plotly_chart(pcfg(fig2, 310),
                            use_container_width=True)

    if not df_pvol.empty:
        fig3 = px.scatter(
            df_pvol, x="Precio_ef", y="Unidades",
            size=sz(df_pvol["Ingresos"]), hover_name="Producto",
            color="Categoria", color_discrete_sequence=CAT_COLORS,
            size_max=45, opacity=0.8,
            custom_data=["Ingresos","Mg_pct","Categoria"]
        )
        fig3.update_traces(
            hovertemplate=(
                "<b>%{hovertext}</b><br>"
                "Categoría: %{customdata[2]}<br>"
                "Precio efectivo: $%{x:,.0f}<br>"
                "Unidades vendidas: %{y:,.0f}<br>"
                "Ingresos: $%{customdata[0]:,.0f}<br>"
                "Margen %%: %{customdata[1]:.1%}<extra></extra>"
            )
        )
        fig3.update_xaxes(title="Precio efectivo ($) — precio real pagado por el cliente", tickprefix="$", tickformat=",.0f")
        fig3.update_yaxes(title="Unidades vendidas — mayor volumen = mayor círculo")
        sec("Precio Efectivo vs Volumen — burbuja proporcional a ingresos")
        st.plotly_chart(pcfg(fig3, 340),
                        use_container_width=True)

    if not df_cmp.empty:
        df_cmp_r = df_cmp[df_cmp["Delta_mg"] < -0.01].head(top_n)
        df_cmp_g = df_cmp[df_cmp["Delta_mg"] > 0.01].head(top_n)

        c3,c4 = st.columns(2)
        with c3:
            if not df_cmp_r.empty:
                fig4 = go.Figure()
                fig4.add_trace(go.Bar(
                    name="Variación del costo %", x=df_cmp_r["Producto"],
                    y=df_cmp_r["Cambio_costo_pct"], marker_color=C_RED, opacity=0.85,
                    marker_line_width=0,
                    hovertemplate="<b>%{x}</b><br>Costo varió: %{y:.1f}%<extra></extra>"
                ))
                fig4.add_trace(go.Bar(
                    name="Variación del precio %", x=df_cmp_r["Producto"],
                    y=df_cmp_r["Cambio_precio_pct"], marker_color=C_NAVY, opacity=0.85,
                    marker_line_width=0,
                    hovertemplate="<b>%{x}</b><br>Precio varió: %{y:.1f}%<extra></extra>"
                ))
                fig4.update_layout(barmode="group", xaxis_tickangle=-30,
                                   yaxis_title="Variación % (período completo)")
                sec("⚠️ Compresión de margen — el costo subió más que el precio")
                st.plotly_chart(pcfg(fig4, 290),
                                use_container_width=True)
            else:
                st.success("No se detectan productos con compresión de margen significativa.")

        with c4:
            if not df_cmp_g.empty:
                fig5 = go.Figure()
                fig5.add_trace(go.Bar(
                    name="Variación del precio %", x=df_cmp_g["Producto"],
                    y=df_cmp_g["Cambio_precio_pct"], marker_color=C_GREEN, opacity=0.85,
                    marker_line_width=0,
                    hovertemplate="<b>%{x}</b><br>Precio varió: %{y:.1f}%<extra></extra>"
                ))
                fig5.add_trace(go.Bar(
                    name="Variación del costo %", x=df_cmp_g["Producto"],
                    y=df_cmp_g["Cambio_costo_pct"], marker_color=C_AMBER, opacity=0.85,
                    marker_line_width=0,
                    hovertemplate="<b>%{x}</b><br>Costo varió: %{y:.1f}%<extra></extra>"
                ))
                fig5.update_layout(barmode="group", xaxis_tickangle=-30,
                                   yaxis_title="Variación % (período completo)")
                sec("✅ Expansión de margen — el precio subió más que el costo")
                st.plotly_chart(pcfg(fig5, 290),
                                use_container_width=True)
            else:
                st.info("Sin productos con expansión de margen en el período.")

        sec("Oportunidades de Repricing")
        c5,c6 = st.columns(2)
        with c5:
            st.markdown("**Subir precio** — Costo comprimió el margen")
            sub_precio = df_cmp[df_cmp["Delta_mg"] < -0.02].head(15)
            if not sub_precio.empty:
                sp = sub_precio[["Producto","Categoria","Precio_inicio","Precio_fin",
                                  "Costo_fin","Mg_inicio_pct","Mg_fin_pct"]].copy()
                sp["Δ Margen"] = sp["Mg_fin_pct"].apply(lambda x: f"{x:.1f}%") + " ← " + \
                                 sp["Mg_inicio_pct"].apply(lambda x: f"{x:.1f}%")
                st.dataframe(sp[["Producto","Categoria","Precio_fin","Costo_fin","Δ Margen"]],
                             use_container_width=True, hide_index=True, height=220)
        with c6:
            st.markdown("**Bajar precio** — Alto margen, bajo volumen")
            baj_precio = df_cmp[(df_cmp["Mg_fin_pct"]>35) & (df_cmp["Delta_mg"]>0)].head(15)
            if not baj_precio.empty:
                bp = baj_precio[["Producto","Categoria","Precio_fin","Mg_fin_pct"]].copy()
                bp["Margen actual"] = bp["Mg_fin_pct"].apply(lambda x: f"{x:.1f}%")
                st.dataframe(bp[["Producto","Categoria","Precio_fin","Margen actual"]],
                             use_container_width=True, hide_index=True, height=220)
            else:
                st.info("No se identifican oportunidades claras de baja de precio.")


# ══════════════════════════════════════════════ TAB 3: PROMOCIONES
with T[3]:
    if df_ap.empty:
        st.info("Sin datos de promociones.")
    else:
        df_ap = df_ap.copy()
        df_ap["Cobertura"] = df_ap["Tiendas"]/n_tiendas if n_tiendas>0 else 0
        df_ap["Semaforo"]  = df_ap.apply(lambda r:
            "🟢 Buena"   if r["Mg_pct"]>=0.20 and r["ROI"]>=2 and r["Cobertura"]>=0.3
            else ("🟡 Regular" if (r["Mg_pct"]>=0.10 or s(r["ROI"])>=1)
            else "🔴 Crítica"), axis=1)

        promo_ing = s(df_ap["Ingresos"].sum())
        promo_mg  = s(df_ap["Margen"].sum())
        promo_inv = s(df_ap["Descuento"].sum())
        roi_glob  = promo_ing/promo_inv if promo_inv>0 else 0
        buenas    = len(df_ap[df_ap["Semaforo"]=="🟢 Buena"])
        criticas  = len(df_ap[df_ap["Semaforo"]=="🔴 Crítica"])

        k = st.columns(6)
        kpi(k[0], clp(promo_ing),  "Ingresos Promo",   "g")
        kpi(k[1], clp(promo_mg),   "Margen Promo",     "g" if promo_mg>0 else "r")
        kpi(k[2], clp(promo_inv),  "Inversión Markdown")
        kpi(k[3], f"{roi_glob:.2f}×", "ROI Global",    "g" if roi_glob>=2 else ("a" if roi_glob>=1 else "r"))
        kpi(k[4], str(buenas),     "Promos Buenas 🟢",  "g")
        kpi(k[5], str(criticas),   "Promos Críticas 🔴","r")

        c1,c2 = st.columns([2,1])
        with c1:
            df_top_pr = df_ap.head(top_n).copy()
            df_top_pr["lbl"] = df_top_pr.apply(
                lambda r: f"Mg {r['Mg_pct']*100:.1f}%  ·  ROI {s(r['ROI']):.1f}×", axis=1)
            fig = px.bar(
                df_top_pr, x="Ingresos", y="Promo", orientation="h",
                color="Semaforo",
                color_discrete_map={"🟢 Buena":C_GREEN,"🟡 Regular":C_AMBER,"🔴 Crítica":C_RED},
                text="lbl",
                custom_data=["Margen","Mg_pct","Descuento","ROI","Tiendas","Cobertura"]
            )
            fig.update_traces(
                textposition="inside", textfont=dict(size=9.5, color="white"),
                insidetextanchor="middle", marker_line_width=0,
                hovertemplate=(
                    "<b>%{y}</b><br>"
                    "Ingresos: $%{x:,.0f}<br>"
                    "Margen: $%{customdata[0]:,.0f} (%{customdata[1]:.1%})<br>"
                    "Inversión markdown: $%{customdata[2]:,.0f}<br>"
                    "ROI: %{customdata[3]:.2f}×<br>"
                    "Tiendas: %{customdata[4]} (%{customdata[5]:.0%} cobertura)<extra></extra>"
                )
            )
            fig.update_layout(yaxis=dict(autorange="reversed"),
                              xaxis=dict(title="Ingresos ($)", tickprefix="$", tickformat=",.0f"))
            sec(f"Top {top_n} promociones por ingresos — color = semáforo de rentabilidad")
            st.plotly_chart(pcfg(fig, max(320,top_n*30)),
                            use_container_width=True)

        with c2:
            if not df_mec.empty:
                df_mec_s = df_mec.copy()
                df_mec_s["lbl"] = df_mec_s["ROI"].apply(lambda x: f"{x:.1f}×")
                fig2 = px.bar(
                    df_mec_s, x="Mecanica", y="ROI",
                    color="ROI", color_continuous_scale=["#C8102E","#F5A623","#1A9E5C"],
                    text="lbl",
                    custom_data=["Ingresos","Mg_pct","Promos","Inversion"]
                )
                fig2.update_traces(
                    textposition="outside", textfont_size=10,
                    marker_line_width=0,
                    hovertemplate=(
                        "<b>%{x}</b><br>"
                        "ROI: %{y:.2f}×<br>"
                        "Ingresos: $%{customdata[0]:,.0f}<br>"
                        "Margen %%: %{customdata[1]:.1%}<br>"
                        "# Promos: %{customdata[2]}<br>"
                        "Inversión: $%{customdata[3]:,.0f}<extra></extra>"
                    )
                )
                fig2.add_hline(y=1, line_dash="dash", line_color=C_RED, line_width=1.5,
                               annotation_text="  ROI = 1× (break-even)",
                               annotation_font_size=9, annotation_font_color=C_RED)
                fig2.update_layout(xaxis_tickangle=-15, coloraxis_showscale=False)
                fig2.update_yaxes(title="ROI")
                sec("ROI por mecánica de promoción")
                st.plotly_chart(pcfg(fig2, max(320,top_n*30), legend=False),
                                use_container_width=True)

        if not df_hm.empty:
            piv = df_hm.pivot(index="Promo", columns="Anio_mes", values="Ingresos").fillna(0)
            fig3 = px.imshow(
                piv, color_continuous_scale=["#EEF2F7","#93C5FD",C_NAVY],
                aspect="auto",
                text_auto=False,
                zmin=0
            )
            fig3.update_traces(
                hovertemplate="<b>%{y}</b><br>Mes: %{x}<br>Ingresos: $%{z:,.0f}<extra></extra>"
            )
            fig3.update_layout(
                coloraxis_colorbar=dict(title="Ingresos $", tickformat=",.0f"),
                xaxis_title="Mes", yaxis_title=""
            )
            sec("Mapa de calor — Ingresos por promo y mes (Top 20)")
            st.plotly_chart(pcfg(fig3, max(380,len(piv)*28), legend=False),
                            use_container_width=True)

        # Filtrar outliers para escala legible (percentil 95 en ROI, margen entre -1 y 2)
        _sc_full = df_ap[df_ap["ROI"].notna()].copy()
        _roi_max = float(_sc_full["ROI"].quantile(0.93)) if len(_sc_full) > 5 else 20
        _mg_min  = max(-1.0, float(_sc_full["Mg_pct"].quantile(0.02)))
        _mg_max  = min(2.0,  float(_sc_full["Mg_pct"].quantile(0.98)))
        df_sc = _sc_full[
            (_sc_full["ROI"] >= 0) & (_sc_full["ROI"] <= _roi_max) &
            (_sc_full["Mg_pct"] >= _mg_min) & (_sc_full["Mg_pct"] <= _mg_max)
        ].copy()
        _excl = len(_sc_full) - len(df_sc)
        df_sc["Cob_sz"] = (df_sc["Cobertura"]*100).clip(lower=1)
        if not df_sc.empty:
            fig4 = px.scatter(
                df_sc, x="ROI", y="Mg_pct", size="Cob_sz",
                hover_name="Promo", color="Semaforo",
                color_discrete_map={"🟢 Buena":C_GREEN,"🟡 Regular":C_AMBER,"🔴 Crítica":C_RED},
                size_max=40, opacity=0.85,
                custom_data=["Ingresos","Descuento","Tiendas","Cobertura"]
            )
            fig4.update_traces(
                hovertemplate=(
                    "<b>%{hovertext}</b><br>"
                    "ROI: %{x:.1f}×  ·  Margen: %{y:.1%}<br>"
                    "Ingresos: $%{customdata[0]:,.0f}<br>"
                    "Inversión: $%{customdata[1]:,.0f}<br>"
                    "Tiendas: %{customdata[2]} (%{customdata[3]:.0%} cobertura)<extra></extra>"
                )
            )
            fig4.add_vline(x=1, line_dash="dash", line_color="#6B7280", line_width=1.5,
                           annotation_text="ROI = 1×", annotation_font_size=9,
                           annotation_font_color="#6B7280")
            fig4.add_hline(y=0.15, line_dash="dash", line_color="#6B7280", line_width=1.5,
                           annotation_text="Margen 15%", annotation_font_size=9,
                           annotation_font_color="#6B7280")
            # Zona ideal (arriba derecha)
            fig4.add_shape(type="rect",
                x0=1, x1=_roi_max, y0=0.15, y1=_mg_max,
                fillcolor="rgba(26,158,92,0.06)", line_width=0)
            fig4.add_annotation(x=_roi_max*0.7, y=_mg_max*0.85,
                text="⭐ Zona ideal", showarrow=False,
                font=dict(color=C_GREEN, size=11, family="Segoe UI"))
            fig4.update_yaxes(tickformat=".0%", title="Margen % sobre ingresos",
                              range=[_mg_min - 0.05, _mg_max + 0.05])
            fig4.update_xaxes(title="ROI (× ingresos por cada $ invertido en descuento)",
                              range=[-0.3, _roi_max + 0.5])
            _note = f" — {_excl} promos con valores extremos excluidas de la vista" if _excl > 0 else ""
            sec(f"Diagnóstico ROI vs Margen % · burbuja = cobertura de tiendas{_note}")
            st.plotly_chart(pcfg(fig4, 380), use_container_width=True)

        # Drill-down individual
        st.divider()
        sec("Drill-down por Promo Individual")
        promo_lista = ["(Selecciona)"] + df_ap["Promo"].tolist()
        promo_sel   = st.selectbox("Promo:", promo_lista, index=0)

        if promo_sel != "(Selecciona)":
            safe_p = promo_sel.replace("'","''")

            @st.cache_data(ttl=900, show_spinner=False)
            def drill_data(fk_, promo_name_):
                (ms, cats, tipos, sists, ps, prs, cs,
                 ciclo_años_, ciclo_nums_, años_flt_,
                 tipos_promo_, ciclos_promo_, nums_promo_,
                 regiones_, comunas_, formatos_t_, segmentos_,
                 tiendas_dir_, prods_dir_) = fk_
                eff_ = set(ms) if ms else set(meses_all)
                if ciclo_años_:
                    cy_ = set()
                    for ca_ in ciclo_años_: cy_.update(CICLO_MESES.get(ca_, set()))
                    eff_ &= cy_
                if ciclo_nums_:
                    cn_ = set()
                    for ca_, mset_ in CICLO_MESES.items():
                        if ca_.split("-")[0] in set(ciclo_nums_): cn_.update(mset_)
                    eff_ &= cn_
                if años_flt_:
                    eff_ = {m_ for m_ in eff_ if m_[:4] in set(años_flt_)}
                v_ = dm.copy()
                if eff_         != set(meses_all): v_ = v_[v_["Anio_mes"].isin(eff_)]
                if set(cats)    != set(cats_all):  v_ = v_[v_["Categoria"].isin(set(cats))]
                if set(tipos)   != set(tipos_all): v_ = v_[v_["Tipo"].isin(set(tipos))]
                if set(sists)   != set(sists_all): v_ = v_[v_["Sistema"].isin(set(sists))]
                if tiendas_dir_ and "Tiendas" in v_.columns:
                    v_ = v_[v_["Tiendas"].isin(set(tiendas_dir_))]
                c_ = duckdb.connect(); c_.register("ventas", v_)
                safe_ = promo_name_.replace("'","''")
                ev_ = c_.execute(f"""
                    SELECT Anio_mes, SUM(Importe) AS Ingresos, SUM(Cantidad) AS Unidades,
                           COUNT(DISTINCT ventas.Tiendas) AS Tiendas,
                           (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct,
                           SUM(Importe)/NULLIF(SUM(Precio_x_Cant-Importe),0) AS ROI
                    FROM ventas WHERE COALESCE(NULLIF(TRIM(Nom_promo),''),Cod_promo,'?')='{safe_}'
                    GROUP BY 1 ORDER BY 1""").df()
                tn_ = c_.execute(f"""
                    SELECT ventas.Tiendas, SUM(Importe) AS Ingresos,
                           (SUM(Importe)-SUM(Costo_total))/NULLIF(SUM(Importe),0) AS Mg_pct
                    FROM ventas WHERE COALESCE(NULLIF(TRIM(Nom_promo),''),Cod_promo,'?')='{safe_}'
                    GROUP BY 1 ORDER BY Ingresos DESC LIMIT 20""").df()
                c_.close()
                return ev_, tn_

            df_dr, df_dt = drill_data(filter_key, promo_sel)
            row_p = df_ap[df_ap["Promo"]==promo_sel].iloc[0]

            st.markdown(f"### {row_p['Semaforo']} {promo_sel}")
            st.caption(f"{row_p['Primer_mes']} → {row_p['Ultimo_mes']} · "
                       f"{int(row_p['Meses'])} meses · {int(row_p['Tiendas'])} tiendas · "
                       f"Cobertura {row_p['Cobertura']*100:.0f}%")

            dk = st.columns(5)
            kpi(dk[0], clp(row_p["Ingresos"]),  "Ingresos")
            kpi(dk[1], clp(row_p["Margen"]),     "Margen", "g" if row_p["Margen"]>0 else "r")
            kpi(dk[2], pct(row_p["Mg_pct"]),     "Margen %", "g" if s(row_p["Mg_pct"])>=0.15 else "r")
            kpi(dk[3], clp(row_p["Descuento"]),  "Inversión")
            kpi(dk[4], f"{s(row_p['ROI']):.2f}×","ROI","g" if s(row_p["ROI"])>=2 else "r")

            if not df_dr.empty:
                dc1,dc2 = st.columns(2)
                with dc1:
                    fig5 = go.Figure()
                    fig5.add_trace(go.Bar(name="Ingresos",x=df_dr["Anio_mes"],y=df_dr["Ingresos"],
                                          marker_color=C_NAVY,yaxis="y"))
                    fig5.add_trace(go.Scatter(name="Tiendas",x=df_dr["Anio_mes"],y=df_dr["Tiendas"],
                                              mode="lines+markers",line=dict(color=C_AMBER,width=2),yaxis="y2"))
                    fig5.update_layout(
                        yaxis=dict(title="Ingresos ($)", gridcolor="#F1F5F9"),
                        yaxis2=dict(title="Tiendas", overlaying="y", side="right"))
                    sec("Evolución Ingresos & Cobertura")
                    st.plotly_chart(pcfg(fig5,250), use_container_width=True)

                with dc2:
                    fig6 = go.Figure()
                    fig6.add_trace(go.Scatter(name="ROI",x=df_dr["Anio_mes"],y=df_dr["ROI"],
                                              mode="lines+markers",line=dict(color=C_GREEN,width=2.5)))
                    fig6.add_trace(go.Scatter(name="Margen %",x=df_dr["Anio_mes"],y=df_dr["Mg_pct"],
                                              mode="lines+markers",line=dict(color=C_AMBER,width=2,dash="dot"),
                                              yaxis="y2"))
                    fig6.add_hline(y=1,line_dash="dash",line_color=C_RED,annotation_text="ROI=1×")
                    fig6.update_layout(
                        yaxis=dict(title="ROI",gridcolor="#F1F5F9"),
                        yaxis2=dict(title="Margen %",overlaying="y",side="right",tickformat=".0%"))
                    sec("ROI & Margen por Mes")
                    st.plotly_chart(pcfg(fig6,250), use_container_width=True)

                if len(df_dr)>=3:
                    last3 = df_dr.tail(3)["Ingresos"].tolist()
                    if last3[-1] < last3[0]*0.75:
                        st.warning(f"⚠️ **Promo Fatigue**: los ingresos cayeron {(1-last3[-1]/last3[0])*100:.0f}% en los últimos 3 meses.")
                    elif last3[-1] > last3[0]*1.20:
                        st.success(f"📈 **Crecimiento sostenido**: +{(last3[-1]/last3[0]-1)*100:.0f}% en los últimos 3 meses.")

            if not df_dt.empty:
                sec(f"Top {top_n} Tiendas para esta Promo")
                fig7 = px.bar(df_dt, x="Ingresos", y="Tiendas", orientation="h",
                              color="Mg_pct", color_continuous_scale=SEQ_RG, text_auto=".2s")
                fig7.update_layout(yaxis=dict(autorange="reversed"),
                                   coloraxis_colorbar=dict(title="Margen %", tickformat=".0%"))
                st.plotly_chart(pcfg(fig7, max(250,len(df_dt)*26),False), use_container_width=True)

        sec("Ranking Completo de Promociones")
        df_t = df_ap.copy()
        df_t["Ingresos"]  = df_t["Ingresos"].apply(clp)
        df_t["Margen"]    = df_t["Margen"].apply(clp)
        df_t["Descuento"] = df_t["Descuento"].apply(clp)
        df_t["Mg %"]      = df_t["Mg_pct"].apply(lambda x: f"{x*100:.1f}%")
        df_t["Desc %"]    = df_t["Desc_pct"].apply(lambda x: f"{x*100:.1f}%")
        df_t["ROI"]       = df_t["ROI"].apply(lambda x: f"{x:.2f}×" if pd.notna(x) else "–")
        df_t["Cobertura"] = df_t["Cobertura"].apply(lambda x: f"{x*100:.0f}%")
        st.dataframe(df_t[["Semaforo","Promo","Primer_mes","Ultimo_mes","Meses",
                            "Tiendas","Cobertura","Ingresos","Mg %","Desc %","ROI","Descuento"]],
                     use_container_width=True, hide_index=True, height=300)


# ══════════════════════════════════════════════ TAB 4: TIENDAS
with T[4]:
    df_tn = df_tn_v.merge(df_tn_t, on="Tiendas", how="left")
    df_tn["Ticket_prom"] = df_tn["Ingresos"]/df_tn["Tickets"].replace(0,np.nan)
    df_tn = df_tn.head(top_n)

    if df_tn.empty:
        st.info("Sin datos.")
    else:
        k = st.columns(4)
        kpi(k[0], num(len(df_tn)),              "Tiendas en ranking")
        kpi(k[1], clp(df_tn["Ingresos"].mean()), "Ingreso Prom/Tienda")
        kpi(k[2], pct(df_tn["Mg_pct"].mean()),   "Margen % Promedio", "g")
        kpi(k[3], pct(df_tn["Mix_promo"].mean()), "Mix Promo Promedio","a")

        _tn2 = df_tn.dropna(subset=["Precio_ef","Ticket_prom"]).copy()
        fig2 = px.scatter(
            _tn2, x="Precio_ef", y="Ticket_prom", size=sz(_tn2["Ingresos"]),
            hover_name="Tiendas", color="Mg_pct", text="Tiendas",
            color_continuous_scale=SEQ_RG, size_max=40, opacity=0.85,
            custom_data=["Ingresos","Margen","Mg_pct","Mix_promo"]
        )
        fig2.update_traces(
            textposition="top center", textfont_size=8,
            hovertemplate=(
                "<b>Tienda %{hovertext}</b><br>"
                "Precio efectivo: $%{x:,.0f}<br>"
                "Ticket promedio: $%{y:,.0f}<br>"
                "Ingresos: $%{customdata[0]:,.0f}<br>"
                "Margen %%: %{customdata[2]:.1%}<br>"
                "Mix promo: %{customdata[3]:.1%}<extra></extra>"
            )
        )
        fig2.update_xaxes(title="Precio efectivo ($)", tickprefix="$", tickformat=",.0f",
                          tickfont=dict(size=10))
        fig2.update_yaxes(title="Ticket promedio ($)", tickprefix="$", tickformat=",.0f",
                          tickfont=dict(size=10))
        fig2.update_layout(coloraxis_colorbar=dict(title="Margen %", tickformat=".0%"))
        sec("Precio Efectivo vs Ticket Promedio por Tienda")
        st.plotly_chart(pcfg(fig2, max(420, top_n*20), legend=False),
                        use_container_width=True)

        if not df_ht.empty:
            piv = df_ht.pivot(index="Tiendas", columns="Anio_mes", values="Ingresos").fillna(0)
            # Filtrar filas con datos reales (al menos 30% de meses con ingresos)
            min_meses = max(1, int(len(piv.columns) * 0.3))
            piv = piv[(piv > 0).sum(axis=1) >= min_meses].head(min(top_n, 20))
            if not piv.empty:
                # Texto dentro de celdas en formato legible
                def _heat_lbl(v):
                    if v == 0: return ""
                    if v >= 1e9: return f"${v/1e9:.1f}B"
                    if v >= 1e6: return f"${v/1e6:.0f}M"
                    return f"${v/1e3:.0f}K"
                lbl_matrix = piv.map(_heat_lbl).values
                fig3 = px.imshow(
                    piv, color_continuous_scale=["#F8FAFC","#93C5FD","#1D4ED8"],
                    aspect="auto", zmin=0, text_auto=False
                )
                fig3.update_traces(
                    text=lbl_matrix, texttemplate="%{text}",
                    textfont=dict(size=9, color="white"),
                    hovertemplate="<b>Tienda %{y}</b><br>%{x}<br>Ingresos: $%{z:,.0f}<extra></extra>"
                )
                fig3.update_layout(
                    coloraxis_colorbar=dict(title="Ingresos", tickformat="$~s", len=0.6),
                    xaxis_title="", yaxis_title=""
                )
                sec("Mapa de calor — Ingresos por tienda y mes")
                st.plotly_chart(pcfg(fig3, max(300, len(piv)*36), legend=False),
                                use_container_width=True)

        sec("Detalle por Tienda")
        ds = df_tn.copy()
        ds["Ingresos"]    = ds["Ingresos"].apply(clp)
        ds["Margen"]      = ds["Margen"].apply(clp)
        ds["Ticket_prom"] = ds["Ticket_prom"].apply(lambda x: f"${x:,.0f}".replace(",",".") if pd.notna(x) else "–")
        ds["Mg_pct"]      = ds["Mg_pct"].apply(lambda x: f"{x*100:.1f}%")
        ds["Desc_pct"]    = ds["Desc_pct"].apply(lambda x: f"{x*100:.1f}%")
        ds["Mix_promo"]   = ds["Mix_promo"].apply(lambda x: f"{x*100:.1f}%")
        ds["Tickets"]     = ds["Tickets"].apply(lambda x: f"{int(x):,}".replace(",",".") if pd.notna(x) else "–")
        st.dataframe(ds[["Tiendas","Ingresos","Margen","Mg_pct","Ticket_prom","Tickets","Desc_pct","Mix_promo"]],
                     use_container_width=True, hide_index=True)

        # ── Cobertura de Categorías por Tienda ──
        if not df_cobertura.empty:
            st.divider()
            sec("🗺️ Cobertura por Categoría — N° tiendas activas")
            st.caption("Cuántas tiendas tienen venta real en cada categoría. Detecta categorías con baja distribución.")

            _max_t = df_cobertura["N_tiendas"].max() if not df_cobertura.empty else 1
            _cob_df = df_cobertura.copy()
            _cob_df["Cobertura_pct"] = _cob_df["N_tiendas"] / _max_t
            _cob_df["lbl"] = _cob_df.apply(
                lambda r: f"{int(r['N_tiendas'])} tiendas ({r['Cobertura_pct']*100:.0f}%)", axis=1)
            fig_cob = px.bar(
                _cob_df.sort_values("N_tiendas"),
                x="N_tiendas", y="Categoria", orientation="h",
                color="Mg_pct", color_continuous_scale=SEQ_RG,
                text="lbl",
                custom_data=["Ingresos","Mg_pct","Desc_pct","Unidades"]
            )
            fig_cob.update_traces(
                textposition="outside", textfont_size=9, marker_line_width=0,
                hovertemplate=(
                    "<b>%{y}</b><br>Tiendas activas: %{x}<br>"
                    "Ingresos: $%{customdata[0]:,.0f}<br>"
                    "Margen %%: %{customdata[1]:.1%}<br>"
                    "Descuento %%: %{customdata[2]:.1%}<extra></extra>"
                )
            )
            fig_cob.update_layout(
                xaxis=dict(title="N° tiendas con venta en la categoría"),
                coloraxis_colorbar=dict(title="Margen %", tickformat=".0%")
            )
            st.plotly_chart(pcfg(fig_cob, max(300, len(_cob_df)*30), legend=False),
                            use_container_width=True)


# ══════════════════════════════════════════════ TAB 5: ALERTAS
with T[5]:
    st.markdown("### 🔔 Panel de Alertas de Pricing")
    st.caption("Revisa estos puntos prioritarios antes de tomar decisiones de precio.")

    if df_al.empty:
        st.info("Sin datos suficientes para generar alertas.")
    else:
        neg_mg    = df_al[df_al["Mg_pct"] < 0].sort_values("Margen")
        low_mg    = df_al[(df_al["Mg_pct"] >= 0) & (df_al["Mg_pct"] < 0.05) & (df_al["Ingresos"]>50000)].sort_values("Ingresos",ascending=False)
        high_desc = df_al[df_al["Desc_pct"] > 0.30].sort_values("Desc_pct",ascending=False)
        precio_bajo = df_al[df_al["Precio_ef"] < df_al["Costo_unit"]].sort_values("Ingresos",ascending=False)
        roi_bajo    = df_alp[df_alp["ROI"] < 1].sort_values("Inversion",ascending=False) if not df_alp.empty else pd.DataFrame()
        desc_exc    = df_alp[df_alp["Desc_pct"] > 0.40].sort_values("Desc_pct",ascending=False) if not df_alp.empty else pd.DataFrame()

        st.markdown(f"""
        <div style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:1rem">
          <div class="chip-r">🔴 <b>{len(neg_mg)}</b> productos con margen negativo</div>
          <div class="chip-a">🟡 <b>{len(low_mg)}</b> productos con margen &lt;5%</div>
          <div class="chip-a">🟡 <b>{len(high_desc)}</b> productos con descuento &gt;30%</div>
          <div class="chip-r">🔴 <b>{len(precio_bajo)}</b> productos precio &lt; costo</div>
          <div class="chip-r">🔴 <b>{len(roi_bajo)}</b> promos con ROI &lt; 1×</div>
          <div class="chip-a">🟡 <b>{len(desc_exc)}</b> promos con descuento &gt;40%</div>
        </div>""", unsafe_allow_html=True)

        a1,a2 = st.columns(2)
        with a1:
            if not neg_mg.empty:
                sec(f"🔴 Productos con Margen NEGATIVO ({len(neg_mg)})")
                d = neg_mg[["Producto","Categoria","Ingresos","Margen","Mg_pct","Costo_unit","Precio_ef"]].copy()
                d["Ingresos"] = d["Ingresos"].apply(clp)
                d["Margen"]   = d["Margen"].apply(clp)
                d["Mg_pct"]   = d["Mg_pct"].apply(lambda x: f"{x*100:.1f}%")
                d["Costo"]    = d["Costo_unit"].apply(lambda x: f"${x:,.0f}".replace(",","."))
                d["Precio"]   = d["Precio_ef"].apply(lambda x: f"${x:,.0f}".replace(",","."))
                st.dataframe(d[["Producto","Categoria","Ingresos","Margen","Mg_pct","Costo","Precio"]],
                             use_container_width=True, hide_index=True, height=220)
            else:
                st.success("✅ No hay productos con margen negativo.")

            if not precio_bajo.empty:
                sec(f"🔴 Precio de venta MENOR al costo ({len(precio_bajo)})")
                d2 = precio_bajo[["Producto","Categoria","Precio_ef","Costo_unit","Ingresos"]].copy()
                d2["Gap"] = d2["Precio_ef"]-d2["Costo_unit"]
                d2 = d2.sort_values("Gap")
                d2["Precio_ef"]  = d2["Precio_ef"].apply(lambda x: f"${x:,.0f}".replace(",","."))
                d2["Costo_unit"] = d2["Costo_unit"].apply(lambda x: f"${x:,.0f}".replace(",","."))
                d2["Gap $"]      = d2["Gap"].apply(lambda x: f"${x:,.0f}".replace(",","."))
                d2["Ingresos"]   = d2["Ingresos"].apply(clp)
                st.dataframe(d2[["Producto","Categoria","Precio_ef","Costo_unit","Gap $","Ingresos"]],
                             use_container_width=True, hide_index=True, height=200)

        with a2:
            if not roi_bajo.empty:
                sec(f"🔴 Promos con ROI &lt; 1× — pierden dinero ({len(roi_bajo)})")
                d3 = roi_bajo[["Promo","Ingresos","Inversion","ROI","Desc_pct","Mg_pct"]].copy()
                d3["Ingresos"]  = d3["Ingresos"].apply(clp)
                d3["Inversion"] = d3["Inversion"].apply(clp)
                d3["ROI"]       = d3["ROI"].apply(lambda x: f"{x:.2f}×")
                d3["Desc_pct"]  = d3["Desc_pct"].apply(lambda x: f"{x*100:.1f}%")
                d3["Mg_pct"]    = d3["Mg_pct"].apply(lambda x: f"{x*100:.1f}%")
                st.dataframe(d3, use_container_width=True, hide_index=True, height=200)
            else:
                st.success("✅ Todas las promos tienen ROI ≥ 1×.")

            if not low_mg.empty:
                sec(f"🟡 Productos con Margen &lt; 5% — {len(low_mg)}")
                d4 = low_mg[["Producto","Categoria","Ingresos","Mg_pct"]].copy()
                d4["Ingresos"] = d4["Ingresos"].apply(clp)
                d4["Mg_pct"]   = d4["Mg_pct"].apply(lambda x: f"{x*100:.2f}%")
                st.dataframe(d4, use_container_width=True, hide_index=True, height=200)

        if len(df_al) > 0:
            # ── Gráfico 1: Margen % de los top productos por ingresos ──
            df_al2 = df_al.copy()
            df_al2["Estado"] = df_al2["Mg_pct"].apply(
                lambda x: "Negativo" if x<0 else ("Bajo <5%" if x<0.05 else ("Normal" if x<0.35 else "Alto")))
            _color_map = {"Negativo":C_RED,"Bajo <5%":C_AMBER,"Normal":C_GREEN,"Alto":C_BLUE2}
            # Top 20 por ingresos, margen entre -200% y 200%
            _top_al = df_al2[df_al2["Mg_pct"].between(-2,2)].nlargest(20,"Ingresos").copy()
            _top_al["lbl"] = _top_al["Mg_pct"].apply(lambda x: f"{x*100:.1f}%")
            _top_al["Ing_lbl"] = _top_al["Ingresos"].apply(clp)

            c_al1, c_al2 = st.columns(2)
            with c_al1:
                fig_al1 = px.bar(
                    _top_al, x="Mg_pct", y="Producto", orientation="h",
                    color="Estado", color_discrete_map=_color_map,
                    text="lbl",
                    custom_data=["Ingresos","Unidades","Precio_ef","Costo_unit"]
                )
                fig_al1.update_traces(
                    textposition="outside", textfont_size=9, marker_line_width=0,
                    hovertemplate=(
                        "<b>%{y}</b><br>"
                        "Margen %%: %{x:.1%}<br>"
                        "Ingresos: $%{customdata[0]:,.0f}<br>"
                        "Unidades: %{customdata[1]:,.0f}<br>"
                        "Precio ef.: $%{customdata[2]:,.0f}<br>"
                        "Costo u.: $%{customdata[3]:,.0f}<extra></extra>"
                    )
                )
                fig_al1.add_vline(x=0, line_color=C_RED, line_width=1.5)
                fig_al1.add_vline(x=0.05, line_dash="dash", line_color=C_AMBER, line_width=1)
                fig_al1.update_layout(
                    yaxis=dict(autorange="reversed"),
                    xaxis=dict(tickformat=".0%"),
                    legend=dict(orientation="h", y=1.1, title_text="")
                )
                st.plotly_chart(pcfg(fig_al1, 460),
                                use_container_width=True)

            with c_al2:
                # Resumen por categoría: margen % promedio ponderado
                _cat_al = df_al.groupby("Categoria").apply(
                    lambda g: pd.Series({
                        "Mg_pct_w": (g["Margen"].sum() / g["Ingresos"].sum()) if g["Ingresos"].sum() > 0 else 0,
                        "Ingresos": g["Ingresos"].sum(),
                        "Productos": len(g),
                        "Neg": (g["Mg_pct"] < 0).sum(),
                        "Bajo": ((g["Mg_pct"] >= 0) & (g["Mg_pct"] < 0.05)).sum()
                    }), include_groups=False
                ).reset_index().sort_values("Mg_pct_w")
                _cat_al["lbl"] = _cat_al["Mg_pct_w"].apply(lambda x: f"{x*100:.1f}%")
                _cat_al["color"] = _cat_al["Mg_pct_w"].apply(
                    lambda x: "Negativo" if x<0 else ("Bajo <5%" if x<0.05 else ("Normal" if x<0.35 else "Alto")))
                fig_al2 = px.bar(
                    _cat_al, x="Mg_pct_w", y="Categoria", orientation="h",
                    color="color", color_discrete_map=_color_map,
                    text="lbl",
                    custom_data=["Ingresos","Productos","Neg","Bajo"]
                )
                fig_al2.update_traces(
                    textposition="outside", textfont_size=10, marker_line_width=0,
                    hovertemplate=(
                        "<b>%{y}</b><br>"
                        "Margen %% pond.: %{x:.1%}<br>"
                        "Ingresos: $%{customdata[0]:,.0f}<br>"
                        "Productos: %{customdata[1]}<br>"
                        "Con margen negativo: %{customdata[2]}<br>"
                        "Con margen &lt;5%%: %{customdata[3]}<extra></extra>"
                    )
                )
                fig_al2.add_vline(x=0, line_color=C_RED, line_width=1.5)
                fig_al2.add_vline(x=0.05, line_dash="dash", line_color=C_AMBER, line_width=1)
                fig_al2.update_layout(
                    yaxis=dict(autorange="reversed"),
                    xaxis=dict(tickformat=".0%", title="Margen % ponderado por ingresos"),
                    showlegend=False
                )
                sec("Margen % por categoría — ponderado por ingresos")
                st.plotly_chart(pcfg(fig_al2, 460),
                                use_container_width=True)


# ══════════════════════════════════════════════ TAB 6: CRUCES
with T[6]:
    st.markdown("### 🔀 Análisis Cruzado de Pricing")
    st.caption("Combina dimensiones para detectar problemas y oportunidades de precio.")

    # ══════════════════════════════════════════════════════════
    #  SECCIÓN 1 — MARGEN POR CATEGORÍA × MES
    # ══════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### 📅 Sección 1 — Evolución del Margen por Categoría y Mes")
    st.caption("Detecta qué categorías están mejorando o deteriorando su rentabilidad en el tiempo.")

    if not df_mg_cat_mes.empty:
        # Heatmap
        piv_mg = df_mg_cat_mes.pivot(index="Categoria", columns="Anio_mes", values="Mg_pct").fillna(np.nan)
        piv_mg = piv_mg.dropna(how="all")
        # Ordenar por margen promedio descendente
        piv_mg = piv_mg.loc[piv_mg.mean(axis=1).sort_values(ascending=False).index]
        sec("Heatmap — Margen % por Categoría y Mes (rojo = bajo · verde = alto)")
        fig_hm_mg = px.imshow(
            (piv_mg * 100).round(1),
            color_continuous_scale=["#C8102E","#F5A623","#1A9E5C"],
            zmin=0, zmax=50,
            aspect="auto",
            text_auto=".1f"
        )
        fig_hm_mg.update_traces(
            textfont=dict(size=9),
            hovertemplate="<b>%{y}</b><br>Mes: %{x}<br>Margen: %{z:.1f}%%<extra></extra>"
        )
        fig_hm_mg.update_layout(
            coloraxis_colorbar=dict(title="Margen %", ticksuffix="%", len=0.6),
            xaxis_title="", yaxis_title=""
        )
        st.plotly_chart(pcfg(fig_hm_mg, max(320, len(piv_mg)*34), legend=False),
                        use_container_width=True)

        # Tendencia por categoría: primera vs última mitad del período
        c1_s1, c2_s1 = st.columns(2)
        with c1_s1:
            meses_ord = sorted(piv_mg.columns.tolist())
            if len(meses_ord) >= 2:
                mitad = len(meses_ord) // 2
                prim_meses = meses_ord[:mitad]
                ult_meses  = meses_ord[mitad:]
                mg_prim = piv_mg[prim_meses].mean(axis=1)
                mg_ult  = piv_mg[ult_meses].mean(axis=1)
                df_tend = pd.DataFrame({
                    "Categoria": piv_mg.index,
                    "Mg_inicio": mg_prim.values,
                    "Mg_fin":    mg_ult.values
                }).dropna()
                df_tend["Delta"] = df_tend["Mg_fin"] - df_tend["Mg_inicio"]
                df_tend["Tendencia"] = df_tend["Delta"].apply(
                    lambda x: "🟢 Mejorando" if x > 0.01 else ("🔴 Deteriorando" if x < -0.01 else "🟡 Estable"))
                df_tend = df_tend.sort_values("Delta")
                df_tend["lbl"] = df_tend["Delta"].apply(lambda x: f"{x*100:+.1f}pp")
                fig_tend = px.bar(
                    df_tend, x="Delta", y="Categoria", orientation="h",
                    color="Tendencia",
                    color_discrete_map={"🟢 Mejorando":C_GREEN,"🔴 Deteriorando":C_RED,"🟡 Estable":C_AMBER},
                    text="lbl",
                    custom_data=["Mg_inicio","Mg_fin"]
                )
                fig_tend.update_traces(
                    textposition="outside", textfont_size=9.5, marker_line_width=0,
                    hovertemplate=(
                        "<b>%{y}</b><br>"
                        "Mg inicio: %{customdata[0]:.1%}<br>"
                        "Mg fin: %{customdata[1]:.1%}<br>"
                        "Variación: %{x:.1%}<extra></extra>"
                    )
                )
                fig_tend.update_layout(
                    yaxis=dict(autorange="reversed"),
                    xaxis=dict(tickformat="+.0%", title="Variación de margen (pp)"),
                    showlegend=True,
                    legend=dict(orientation="h", y=1.08, x=0, font_size=10)
                )
                sec("Tendencia — Variación de margen: primera vs segunda mitad del período")
                st.plotly_chart(pcfg(fig_tend, max(300, len(df_tend)*32)), use_container_width=True)

        with c2_s1:
            # Volatilidad: desviación estándar del margen por categoría
            df_vol = df_mg_cat_mes.groupby("Categoria")["Mg_pct"].agg(
                Media="mean", StdDev="std", Meses="count"
            ).reset_index().dropna().sort_values("StdDev", ascending=False)
            df_vol["lbl"] = df_vol["StdDev"].apply(lambda x: f"±{x*100:.1f}pp")
            fig_vol = px.bar(
                df_vol.head(15), x="StdDev", y="Categoria", orientation="h",
                color="Media", color_continuous_scale=["#C8102E","#F5A623","#1A9E5C"],
                text="lbl",
                custom_data=["Media","Meses"]
            )
            fig_vol.update_traces(
                textposition="outside", textfont_size=9.5, marker_line_width=0,
                hovertemplate=(
                    "<b>%{y}</b><br>"
                    "Volatilidad (std): %{x:.1%}<br>"
                    "Margen promedio: %{customdata[0]:.1%}<br>"
                    "Meses con datos: %{customdata[1]}<extra></extra>"
                )
            )
            fig_vol.update_layout(
                yaxis=dict(autorange="reversed"),
                xaxis=dict(tickformat=".0%", title="Desviación estándar del margen %"),
                coloraxis_showscale=False
            )
            sec("Volatilidad de Margen % por Categoría — mayor barra = más inestable")
            st.plotly_chart(pcfg(fig_vol, max(300, len(df_vol.head(15))*32), legend=False),
                            use_container_width=True)

    # ══════════════════════════════════════════════════════════
    #  SECCIÓN 2 — DESCUENTO × MARGEN (ZONA DE PELIGRO)
    # ══════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### ⚠️ Sección 2 — Cruce Descuento % × Margen % por Producto")
    st.caption("Identifica productos en la zona de peligro: alto descuento con margen bajo o negativo.")

    if not df_desc_mg.empty:
        _dm = df_desc_mg.copy()
        _dm = _dm[_dm["Desc_pct"].between(0, 1) & _dm["Mg_pct"].between(-0.5, 1.0)]
        _dm["Zona"] = _dm.apply(lambda r:
            "🔴 Peligro: desc. alto + margen bajo" if r["Desc_pct"] > 0.20 and r["Mg_pct"] < 0.15
            else ("🟢 Ideal: desc. bajo + margen alto"  if r["Desc_pct"] <= 0.20 and r["Mg_pct"] >= 0.15
            else ("🟡 Alto margen con descuento alto"   if r["Desc_pct"] > 0.20 and r["Mg_pct"] >= 0.15
            else "⚪ Sin descuento significativo")), axis=1)
        zona_col = {
            "🔴 Peligro: desc. alto + margen bajo": C_RED,
            "🟢 Ideal: desc. bajo + margen alto":   C_GREEN,
            "🟡 Alto margen con descuento alto":     C_AMBER,
            "⚪ Sin descuento significativo":         C_GRAY,
        }
        n_peligro = len(_dm[_dm["Zona"].str.startswith("🔴")])
        n_ideal   = len(_dm[_dm["Zona"].str.startswith("🟢")])

        st.markdown(f"""
        <div style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:.8rem">
          <div class="chip-r">🔴 <b>{n_peligro}</b> productos en zona de peligro</div>
          <div class="chip-g">🟢 <b>{n_ideal}</b> productos en zona ideal</div>
        </div>""", unsafe_allow_html=True)

        c1_s2, c2_s2 = st.columns([3, 2])
        with c1_s2:
            fig_sc2 = px.scatter(
                _dm, x="Desc_pct", y="Mg_pct",
                size=sz(_dm["Ingresos"]),
                hover_name="Producto", color="Zona",
                color_discrete_map=zona_col,
                size_max=50, opacity=0.78,
                custom_data=["Ingresos","Unidades","Categoria"]
            )
            fig_sc2.update_traces(
                hovertemplate=(
                    "<b>%{hovertext}</b><br>"
                    "Categoría: %{customdata[2]}<br>"
                    "Descuento: %{x:.1%} · Margen: %{y:.1%}<br>"
                    "Ingresos: $%{customdata[0]:,.0f}<br>"
                    "Unidades: %{customdata[1]:,.0f}<extra></extra>"
                )
            )
            fig_sc2.add_vline(x=0.20, line_dash="dash", line_color="#9CA3AF", line_width=1.5,
                              annotation_text="Desc. 20%", annotation_font_size=9,
                              annotation_font_color=C_GRAY)
            fig_sc2.add_hline(y=0.15, line_dash="dash", line_color="#9CA3AF", line_width=1.5,
                              annotation_text="Margen 15%", annotation_font_size=9,
                              annotation_font_color=C_GRAY)
            fig_sc2.add_shape(type="rect", x0=0.20, x1=1.0, y0=-0.5, y1=0.15,
                              fillcolor="rgba(200,16,46,0.06)", line_width=0)
            fig_sc2.add_shape(type="rect", x0=0, x1=0.20, y0=0.15, y1=1.0,
                              fillcolor="rgba(26,158,92,0.06)", line_width=0)
            fig_sc2.add_annotation(x=0.60, y=-0.30, text="⚠️ Zona de peligro",
                                   showarrow=False, font=dict(color=C_RED, size=11))
            fig_sc2.add_annotation(x=0.05, y=0.80, text="🎯 Zona ideal",
                                   showarrow=False, font=dict(color=C_GREEN, size=11))
            fig_sc2.update_xaxes(tickformat=".0%", title="Descuento % (precio lista → efectivo)")
            fig_sc2.update_yaxes(tickformat=".0%", title="Margen % sobre ingresos")
            sec("Descuento % vs Margen % — burbuja proporcional a ingresos")
            st.plotly_chart(pcfg(fig_sc2, 420), use_container_width=True)

        with c2_s2:
            # Top productos en zona peligro
            df_peligro = _dm[_dm["Zona"].str.startswith("🔴")].nlargest(top_n, "Ingresos").copy()
            if not df_peligro.empty:
                df_peligro["lbl"] = df_peligro.apply(
                    lambda r: f"Desc {r['Desc_pct']*100:.0f}% · Mg {r['Mg_pct']*100:.1f}%", axis=1)
                fig_pel = px.bar(
                    df_peligro, x="Ingresos", y="Producto", orientation="h",
                    color="Desc_pct", color_continuous_scale=["#F5A623","#C8102E"],
                    text="lbl",
                    custom_data=["Mg_pct","Categoria"]
                )
                fig_pel.update_traces(
                    textposition="inside", textfont=dict(size=9, color="white"),
                    insidetextanchor="middle", marker_line_width=0,
                    hovertemplate=(
                        "<b>%{y}</b><br>Categoría: %{customdata[1]}<br>"
                        "Ingresos: $%{x:,.0f}<br>"
                        "Margen: %{customdata[0]:.1%}<extra></extra>"
                    )
                )
                fig_pel.update_layout(
                    yaxis=dict(autorange="reversed"),
                    xaxis=dict(tickprefix="$", tickformat="~s"),
                    coloraxis_colorbar=dict(title="Desc %", tickformat=".0%")
                )
                sec(f"⚠️ Productos zona peligro con mayor exposición ({len(df_peligro)})")
                st.plotly_chart(pcfg(fig_pel, max(300, len(df_peligro)*30), legend=False),
                                use_container_width=True)
            else:
                st.success("✅ No hay productos en zona de peligro.")

    # ══════════════════════════════════════════════════════════
    #  SECCIÓN 3 — CATEGORÍA × PROMOCIÓN × RENTABILIDAD
    # ══════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### 🎯 Sección 3 — Intensidad Promocional vs Rentabilidad por Categoría")
    st.caption("¿Las categorías con más promoción tienen mejor o peor margen?")

    if not df_cat_promo.empty:
        c1_s3, c2_s3 = st.columns(2)
        with c1_s3:
            _cp = df_cat_promo.copy().sort_values("Mg_pct", ascending=False)
            fig_cp1 = go.Figure()
            fig_cp1.add_trace(go.Bar(
                name="Margen %", x=_cp["Categoria"], y=_cp["Mg_pct"]*100,
                marker_color=C_NAVY, opacity=0.85, yaxis="y",
                text=_cp["Mg_pct"].apply(lambda x: f"{x*100:.1f}%"),
                textposition="outside", textfont_size=9, marker_line_width=0,
                hovertemplate="<b>%{x}</b><br>Margen %%: %{y:.1f}%%<extra></extra>"
            ))
            fig_cp1.add_trace(go.Scatter(
                name="Mix Promo %", x=_cp["Categoria"], y=_cp["Mix_promo"]*100,
                mode="lines+markers", yaxis="y2",
                line=dict(color=C_RED, width=2.5), marker=dict(size=8),
                hovertemplate="<b>%{x}</b><br>Mix promo: %{y:.1f}%%<extra></extra>"
            ))
            fig_cp1.update_layout(
                xaxis_tickangle=-30,
                yaxis=dict(title="Margen %", ticksuffix="%", gridcolor="#F3F4F6"),
                yaxis2=dict(title="Mix Promo %", overlaying="y", side="right",
                            ticksuffix="%", showgrid=False),
                legend=dict(orientation="h", y=1.1, x=0, font_size=10)
            )
            sec("Margen % (barras) vs Mix Promo % (línea) por Categoría")
            st.plotly_chart(pcfg(fig_cp1, 340, legend=False), use_container_width=True)

        with c2_s3:
            # Scatter: Mix promo vs Mg_pct por categoría
            fig_cp2 = px.scatter(
                _cp, x="Mix_promo", y="Mg_pct",
                size=sz(_cp["Ingresos"]),
                hover_name="Categoria",
                color="Desc_pct",
                color_continuous_scale=["#1A9E5C","#F5A623","#C8102E"],
                size_max=55, opacity=0.85,
                text="Categoria",
                custom_data=["Ingresos","Desc_pct"]
            )
            fig_cp2.update_traces(
                textposition="top center", textfont_size=9,
                hovertemplate=(
                    "<b>%{hovertext}</b><br>"
                    "Mix promo: %{x:.1%}<br>"
                    "Margen %%: %{y:.1%}<br>"
                    "Ingresos: $%{customdata[0]:,.0f}<br>"
                    "Desc. aplicado: %{customdata[1]:.1%}<extra></extra>"
                )
            )
            med_mix = _cp["Mix_promo"].median()
            med_mgc = _cp["Mg_pct"].median()
            fig_cp2.add_vline(x=med_mix, line_dash="dot", line_color="#9CA3AF", line_width=1.5,
                              annotation_text=f"Mediana {med_mix*100:.0f}%", annotation_font_size=9)
            fig_cp2.add_hline(y=med_mgc, line_dash="dot", line_color="#9CA3AF", line_width=1.5,
                              annotation_text=f"Mediana {med_mgc*100:.1f}%", annotation_font_size=9)
            fig_cp2.update_xaxes(tickformat=".0%", title="Mix Promo % (ingresos con promo)")
            fig_cp2.update_yaxes(tickformat=".0%", title="Margen % sobre ingresos")
            fig_cp2.update_layout(coloraxis_colorbar=dict(title="Desc %", tickformat=".0%"))
            sec("Mix Promo vs Margen % — burbuja = ingresos · color = descuento aplicado")
            st.plotly_chart(pcfg(fig_cp2, 340, legend=False), use_container_width=True)

        # Ranking tabla resumen
        sec("Ranking de Categorías: Rentabilidad vs Exposición Promocional")
        _cp_t = df_cat_promo.copy().sort_values("Mg_pct", ascending=False)
        _cp_t["Ingresos"]   = _cp_t["Ingresos"].apply(clp)
        _cp_t["Margen %"]   = _cp_t["Mg_pct"].apply(lambda x: f"{x*100:.1f}%")
        _cp_t["Mix Promo"]  = _cp_t["Mix_promo"].apply(lambda x: f"{x*100:.1f}%")
        _cp_t["Descuento %"]= _cp_t["Desc_pct"].apply(lambda x: f"{x*100:.1f}%")
        _cp_t["Estado"]     = _cp_t["Mg_pct"].apply(lambda x:
            "🟢" if x >= 0.20 else ("🟡" if x >= 0.10 else "🔴"))
        st.dataframe(_cp_t[["Estado","Categoria","Ingresos","Margen %","Mix Promo","Descuento %"]],
                     use_container_width=True, hide_index=True, height=250)

    # ══════════════════════════════════════════════════════════
    #  SECCIÓN 4 — BRECHA PRECIO LISTA vs EFECTIVO
    # ══════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### 💲 Sección 4 — Brecha Precio Lista vs Precio Efectivo")
    st.caption("Cuánto descuento real se está aplicando. La brecha es el dinero que se deja sobre la mesa.")

    if not df_precio_gap.empty:
        c1_s4, c2_s4 = st.columns([1, 3])
        with c1_s4:
            cat_gap_sel = st.selectbox("Filtrar categoría:", ["(Todas)"] +
                                        sorted(df_precio_gap["Categoria"].unique().tolist()),
                                        key="cat_gap")
        with c2_s4:
            st.write("")

        df_gap_f = df_precio_gap if cat_gap_sel == "(Todas)" \
                   else df_precio_gap[df_precio_gap["Categoria"] == cat_gap_sel]
        df_gap_agg = df_gap_f.groupby("Anio_mes").agg(
            Precio_lista=("Precio_lista","mean"),
            Precio_ef=("Precio_ef","mean"),
            Gap_pct=("Gap_pct","mean")
        ).reset_index()

        fig_gap = go.Figure()
        fig_gap.add_trace(go.Scatter(
            name="Precio Lista", x=df_gap_agg["Anio_mes"], y=df_gap_agg["Precio_lista"],
            mode="lines+markers", line=dict(color=C_NAVY, width=2.5), marker=dict(size=6),
            hovertemplate="<b>%{x}</b><br>Precio Lista: $%{y:,.0f}<extra></extra>"
        ))
        fig_gap.add_trace(go.Scatter(
            name="Precio Efectivo (con descuento)", x=df_gap_agg["Anio_mes"], y=df_gap_agg["Precio_ef"],
            mode="lines+markers", line=dict(color=C_BLUE2, width=2, dash="dot"), marker=dict(size=6),
            fill="tonexty", fillcolor="rgba(200,16,46,0.10)",
            hovertemplate="<b>%{x}</b><br>Precio Efectivo: $%{y:,.0f}<extra></extra>"
        ))
        fig_gap.add_trace(go.Scatter(
            name="Descuento aplicado %", x=df_gap_agg["Anio_mes"], y=df_gap_agg["Gap_pct"],
            mode="lines+markers+text",
            text=df_gap_agg["Gap_pct"].apply(lambda x: f"{x*100:.1f}%"),
            textposition="top center", textfont=dict(size=9, color=C_RED),
            line=dict(color=C_RED, width=1.5, dash="dash"), marker=dict(size=5),
            yaxis="y2",
            hovertemplate="<b>%{x}</b><br>Descuento: %{y:.1%}<extra></extra>"
        ))
        fig_gap.update_layout(
            yaxis=dict(title="Precio ($)", tickprefix="$", tickformat=",.0f",
                       gridcolor="#F3F4F6"),
            yaxis2=dict(title="Descuento %", overlaying="y", side="right",
                        tickformat=".1%", showgrid=False),
            legend=dict(orientation="h", y=1.1, x=0, font_size=10)
        )
        sec("Precio Lista vs Efectivo — área roja = dinero dejado sobre la mesa")
        st.plotly_chart(pcfg(fig_gap, 340, legend=False), use_container_width=True)

        # Por categoría: brecha promedio en barras
        df_gap_cat = df_precio_gap.groupby("Categoria").agg(
            Gap_pct=("Gap_pct","mean"),
            Precio_lista=("Precio_lista","mean"),
            Precio_ef=("Precio_ef","mean"),
            Ingresos=("Ingresos","sum")
        ).reset_index().sort_values("Gap_pct", ascending=False)
        df_gap_cat["lbl"] = df_gap_cat["Gap_pct"].apply(lambda x: f"{x*100:.1f}%")
        fig_gap2 = px.bar(
            df_gap_cat, x="Gap_pct", y="Categoria", orientation="h",
            color="Gap_pct",
            color_continuous_scale=["#1A9E5C","#F5A623","#C8102E"],
            text="lbl",
            custom_data=["Precio_lista","Precio_ef","Ingresos"]
        )
        fig_gap2.update_traces(
            textposition="outside", textfont_size=10, marker_line_width=0,
            hovertemplate=(
                "<b>%{y}</b><br>"
                "Descuento promedio: %{x:.1%}<br>"
                "Precio lista: $%{customdata[0]:,.0f}<br>"
                "Precio efectivo: $%{customdata[1]:,.0f}<br>"
                "Ingresos: $%{customdata[2]:,.0f}<extra></extra>"
            )
        )
        fig_gap2.update_layout(
            yaxis=dict(autorange="reversed"),
            xaxis=dict(tickformat=".0%", title="Descuento promedio aplicado"),
            coloraxis_showscale=False
        )
        sec("Descuento Promedio Aplicado por Categoría — mayor % = más diferencia entre lista y efectivo")
        st.plotly_chart(pcfg(fig_gap2, max(280, len(df_gap_cat)*32), legend=False),
                        use_container_width=True)

    # ══════════════════════════════════════════════════════════
    #  SECCIÓN 5 — ELASTICIDAD PRECIO-VOLUMEN POR CATEGORÍA
    # ══════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### 📐 Sección 5 — Relación Precio vs Volumen (Elasticidad aproximada)")
    st.caption("¿Cuando el precio baja, el volumen sube? Identifica categorías con mayor sensibilidad al precio.")

    if not df_pev.empty and len(df_pev["Anio_mes"].unique()) >= 3:
        # Calcular correlación precio vs unidades por categoría
        elast_rows = []
        for cat, grp in df_pev.groupby("Categoria"):
            grp_agg = grp.groupby("Anio_mes").agg(
                Precio_ef=("Precio_ef","mean"),
                Unidades=("Unidades","sum")
            ).reset_index()
            if len(grp_agg) >= 3 and grp_agg["Precio_ef"].std() > 0 and grp_agg["Unidades"].std() > 0:
                corr = grp_agg["Precio_ef"].corr(grp_agg["Unidades"])
                elast_rows.append({
                    "Categoria": cat,
                    "Correlacion": corr,
                    "Precio_prom": grp_agg["Precio_ef"].mean(),
                    "Unidades_prom": grp_agg["Unidades"].mean(),
                    "Ingresos_tot": grp["Ingresos"].sum() if "Ingresos" in grp.columns else 0
                })
        df_elast = pd.DataFrame(elast_rows).sort_values("Correlacion")

        if not df_elast.empty:
            c1_s5, c2_s5 = st.columns([2, 1])
            with c1_s5:
                df_elast["Tipo"] = df_elast["Correlacion"].apply(
                    lambda x: "🔵 Elástica (precio↑ → vol↓)" if x < -0.3
                    else ("🟠 Inelástica (precio no afecta vol)" if x > 0.3
                    else "⚪ Sin correlación clara"))
                df_elast["lbl"] = df_elast["Correlacion"].apply(lambda x: f"{x:+.2f}")
                fig_el = px.bar(
                    df_elast, x="Correlacion", y="Categoria", orientation="h",
                    color="Correlacion",
                    color_continuous_scale=["#003B7A","#E5E7EB","#C8102E"],
                    color_continuous_midpoint=0,
                    text="lbl",
                    custom_data=["Tipo","Precio_prom","Unidades_prom"]
                )
                fig_el.update_traces(
                    textposition="outside", textfont_size=9.5, marker_line_width=0,
                    hovertemplate=(
                        "<b>%{y}</b><br>"
                        "Correlación precio-volumen: %{x:+.2f}<br>"
                        "%{customdata[0]}<br>"
                        "Precio prom: $%{customdata[1]:,.0f}<br>"
                        "Unidades prom: %{customdata[2]:,.0f}<extra></extra>"
                    )
                )
                fig_el.add_vline(x=0, line_color="#374151", line_width=1.5)
                fig_el.add_vline(x=-0.3, line_dash="dash", line_color=C_BLUE2, line_width=1,
                                 annotation_text="Umbral elástico", annotation_font_size=9)
                fig_el.add_vline(x=0.3, line_dash="dash", line_color=C_RED, line_width=1,
                                 annotation_text="Umbral inelástico", annotation_font_size=9)
                fig_el.update_layout(
                    yaxis=dict(autorange="reversed"),
                    xaxis=dict(range=[-1.1, 1.1], title="Correlación de Pearson (precio vs unidades)",
                               tickformat="+.1f"),
                    coloraxis_showscale=False
                )
                sec("Correlación Precio vs Volumen por Categoría — negativo = sensible al precio")
                st.plotly_chart(pcfg(fig_el, max(280, len(df_elast)*32), legend=False),
                                use_container_width=True)

            with c2_s5:
                st.markdown("""
                <div style="background:#F8FAFC;border-radius:10px;padding:1rem;margin-top:2rem;
                            border-left:3px solid #003B7A;font-size:.82rem;line-height:1.6">
                <b>Cómo leer este gráfico:</b><br><br>
                🔵 <b>Correlación negativa (azul)</b><br>
                La categoría es <b>elástica</b>: cuando el precio sube, el volumen baja.<br>
                → Considera subir precio con cuidado.<br><br>
                🔴 <b>Correlación positiva (roja)</b><br>
                La categoría es <b>inelástica</b> o tiene estacionalidad que mueve ambas variables.<br>
                → Precio tiene menos impacto en demanda.<br><br>
                ⚪ <b>Cercano a 0</b><br>
                Sin correlación clara. Puede haber otros factores (temporada, mix de tiendas).<br><br>
                <i>*Correlación de Pearson calculada entre precio efectivo mensual y unidades vendidas.</i>
                </div>""", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════
    #  SECCIÓN 6 — PENETRACIÓN PROMO vs SIN PROMO
    # ══════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### 📊 Sección 6 — Penetración Promo vs Sin Promo por Producto")
    st.caption("Para cada producto: qué % de sus ventas ocurrió bajo una promoción activa vs sin promo.")

    if df_pen_prod.empty:
        st.warning("⚠️ Recalcula el caché para generar este análisis (botón en el sidebar).")
    else:
        # KPIs globales de penetración
        _tot_cant     = df_pen_prod["Cant_total"].sum()
        _tot_promo    = df_pen_prod["Cant_promo"].sum()
        _tot_sin      = df_pen_prod["Cant_sin_promo"].sum()
        _pen_global   = _tot_promo / _tot_cant if _tot_cant > 0 else 0
        _ing_tot      = df_pen_prod["Ingresos"].sum()
        _ing_promo    = df_pen_prod["Ing_promo"].sum()
        _pen_ing_glob = _ing_promo / _ing_tot if _ing_tot > 0 else 0

        kp = st.columns(5)
        kpi(kp[0], f"{_pen_global*100:.1f}%",    "Penetración global (unidades)", "a")
        kpi(kp[1], f"{_pen_ing_glob*100:.1f}%",  "Penetración global (ingresos)", "a")
        kpi(kp[2], num(_tot_promo),               "Unidades en promo")
        kpi(kp[3], num(_tot_sin),                 "Unidades sin promo")
        kpi(kp[4], num(len(df_pen_prod)),         "Productos con datos")

        # ── Evolución mensual de penetración ──
        if not df_pen_mes.empty:
            _pm = df_pen_mes.copy()
            _pm["CicloAño"] = _pm["Anio_mes"].map(MES_CICLO).fillna("")
            fig_pen_ev = go.Figure()
            fig_pen_ev.add_trace(go.Bar(
                name="Sin Promo", x=_pm["Anio_mes"], y=_pm["Cant_sin_promo"],
                marker_color=C_NAVY, opacity=0.85, marker_line_width=0,
                hovertemplate="<b>%{x}</b><br>Sin promo: %{y:,.0f} un.<extra></extra>"
            ))
            fig_pen_ev.add_trace(go.Bar(
                name="Con Promo", x=_pm["Anio_mes"], y=_pm["Cant_promo"],
                marker_color=C_RED, opacity=0.85, marker_line_width=0,
                hovertemplate="<b>%{x}</b><br>Con promo: %{y:,.0f} un.<extra></extra>"
            ))
            fig_pen_ev.add_trace(go.Scatter(
                name="Penetración %", x=_pm["Anio_mes"], y=_pm["Penetracion"],
                mode="lines+markers+text",
                text=_pm["Penetracion"].apply(lambda x: f"{x*100:.1f}%"),
                textposition="top center", textfont=dict(size=9, color=C_AMBER),
                line=dict(color=C_AMBER, width=2.5), marker=dict(size=7),
                yaxis="y2",
                hovertemplate="<b>%{x}</b><br>Penetración: %{y:.1%}<extra></extra>"
            ))
            fig_pen_ev.update_layout(
                barmode="stack",
                yaxis=dict(title="Unidades vendidas", gridcolor="#F3F4F6", tickformat=",.0f"),
                yaxis2=dict(title="Penetración %", overlaying="y", side="right",
                            tickformat=".0%", showgrid=False),
                legend=dict(orientation="h", y=1.1, x=0, font_size=10)
            )
            sec("Evolución mensual — Unidades con promo (rojo) vs sin promo (azul) · línea = penetración %")
            st.plotly_chart(pcfg(fig_pen_ev, 320, legend=False), use_container_width=True)

        # ── Por categoría ──
        c1_s6, c2_s6 = st.columns(2)
        with c1_s6:
            if not df_pen_cat.empty:
                _pc = df_pen_cat.copy().sort_values("Penetracion", ascending=False)
                _pc["lbl"] = _pc.apply(
                    lambda r: f"{r['Penetracion']*100:.1f}%  ({num(r['Cant_promo'])} vs {num(r['Cant_sin_promo'])})", axis=1)
                fig_pen_cat = px.bar(
                    _pc, x="Penetracion", y="Categoria", orientation="h",
                    color="Penetracion",
                    color_continuous_scale=["#1A9E5C","#F5A623","#C8102E"],
                    text="lbl",
                    custom_data=["Cant_promo","Cant_sin_promo","Cant_total","Ingresos"]
                )
                fig_pen_cat.update_traces(
                    textposition="outside", textfont_size=9.5, marker_line_width=0,
                    hovertemplate=(
                        "<b>%{y}</b><br>"
                        "Penetración: %{x:.1%}<br>"
                        "Unidades en promo: %{customdata[0]:,.0f}<br>"
                        "Unidades sin promo: %{customdata[1]:,.0f}<br>"
                        "Total: %{customdata[2]:,.0f}<extra></extra>"
                    )
                )
                fig_pen_cat.update_layout(
                    yaxis=dict(autorange="reversed"),
                    xaxis=dict(tickformat=".0%", title="% unidades vendidas bajo promo"),
                    coloraxis_showscale=False
                )
                sec("Penetración Promo por Categoría — mayor % = más dependiente de promociones")
                st.plotly_chart(pcfg(fig_pen_cat, max(280, len(_pc)*32), legend=False),
                                use_container_width=True)

        with c2_s6:
            if not df_pen_cat_mes.empty and len(df_pen_mes) >= 2:
                piv_pen = df_pen_cat_mes.pivot(
                    index="Categoria", columns="Anio_mes", values="Penetracion").fillna(0)
                piv_pen = piv_pen.loc[piv_pen.mean(axis=1).sort_values(ascending=False).index]
                fig_pen_hm = px.imshow(
                    (piv_pen * 100).round(1),
                    color_continuous_scale=["#EEF2FF","#F5A623","#C8102E"],
                    zmin=0, zmax=100,
                    aspect="auto",
                    text_auto=".0f"
                )
                fig_pen_hm.update_traces(
                    textfont=dict(size=9),
                    hovertemplate="<b>%{y}</b><br>Mes: %{x}<br>Penetración: %{z:.1f}%%<extra></extra>"
                )
                fig_pen_hm.update_layout(
                    coloraxis_colorbar=dict(title="Penet. %", ticksuffix="%", len=0.6),
                    xaxis_title="", yaxis_title=""
                )
                sec("Heatmap Penetración % por Categoría × Mes")
                st.plotly_chart(pcfg(fig_pen_hm, max(280, len(piv_pen)*32), legend=False),
                                use_container_width=True)

        # ── Top productos más penetrados (ranking) ──
        st.divider()
        c3_s6, c4_s6 = st.columns([1, 3])
        with c3_s6:
            pen_min = st.slider("Penetración mín. %", 0, 100, 20, key="pen_min")
            pen_cat_f = st.selectbox("Categoría", ["(Todas)"] +
                                      sorted(df_pen_prod["Categoria"].unique().tolist()),
                                      key="pen_cat")
        with c4_s6:
            st.write("")

        _pp_f = df_pen_prod[df_pen_prod["Penetracion"] >= pen_min/100].copy()
        if pen_cat_f != "(Todas)":
            _pp_f = _pp_f[_pp_f["Categoria"] == pen_cat_f]

        if not _pp_f.empty:
            top_pen = _pp_f.nlargest(top_n, "Ingresos").copy()
            top_pen["lbl"] = top_pen["Penetracion"].apply(lambda x: f"{x*100:.1f}%")
            top_pen["Mg_promo"]    = (top_pen["Ing_promo"]    - top_pen["Costo_promo"]
                                      ) / top_pen["Ing_promo"].replace(0, np.nan)
            top_pen["Mg_sin_promo"]= (top_pen["Ing_sin_promo"] - top_pen["Costo_sin_promo"]
                                      ) / top_pen["Ing_sin_promo"].replace(0, np.nan)

            # Gráfico de barras agrupadas: cant promo vs sin promo
            fig_pp = go.Figure()
            fig_pp.add_trace(go.Bar(
                name="Sin Promo 🔵", x=top_pen["Cant_sin_promo"], y=top_pen["Producto"],
                orientation="h", marker_color=C_NAVY, opacity=0.85, marker_line_width=0,
                hovertemplate="<b>%{y}</b><br>Sin promo: %{x:,.0f} un.<extra></extra>"
            ))
            fig_pp.add_trace(go.Bar(
                name="Con Promo 🔴", x=top_pen["Cant_promo"], y=top_pen["Producto"],
                orientation="h", marker_color=C_RED, opacity=0.85, marker_line_width=0,
                text=top_pen["lbl"],
                textposition="outside", textfont_size=9.5,
                hovertemplate="<b>%{y}</b><br>Con promo: %{x:,.0f} un.<extra></extra>"
            ))
            fig_pp.update_layout(
                barmode="group",
                yaxis=dict(autorange="reversed"),
                xaxis=dict(title="Unidades vendidas", tickformat=",.0f"),
                legend=dict(orientation="h", y=1.08, x=0, font_size=10)
            )
            sec(f"Top {top_n} productos — Unidades con promo (rojo) vs sin promo (azul) · % = penetración")
            st.plotly_chart(pcfg(fig_pp, max(340, len(top_pen)*34), legend=False),
                            use_container_width=True)

            # Margen promo vs sin promo
            _mg_df = top_pen[["Producto","Mg_promo","Mg_sin_promo"]].dropna().copy()
            if not _mg_df.empty:
                fig_mg_pp = go.Figure()
                fig_mg_pp.add_trace(go.Bar(
                    name="Margen % SIN promo", x=_mg_df["Mg_sin_promo"], y=_mg_df["Producto"],
                    orientation="h", marker_color=C_NAVY, opacity=0.85, marker_line_width=0,
                    text=_mg_df["Mg_sin_promo"].apply(lambda x: f"{x*100:.1f}%" if pd.notna(x) else ""),
                    textposition="outside", textfont_size=9,
                    hovertemplate="<b>%{y}</b><br>Margen sin promo: %{x:.1%}<extra></extra>"
                ))
                fig_mg_pp.add_trace(go.Bar(
                    name="Margen % CON promo", x=_mg_df["Mg_promo"], y=_mg_df["Producto"],
                    orientation="h", marker_color=C_RED, opacity=0.85, marker_line_width=0,
                    text=_mg_df["Mg_promo"].apply(lambda x: f"{x*100:.1f}%" if pd.notna(x) else ""),
                    textposition="outside", textfont_size=9,
                    hovertemplate="<b>%{y}</b><br>Margen con promo: %{x:.1%}<extra></extra>"
                ))
                fig_mg_pp.update_layout(
                    barmode="group",
                    yaxis=dict(autorange="reversed"),
                    xaxis=dict(tickformat=".0%", title="Margen % sobre ingresos")
                )
                fig_mg_pp.add_vline(x=0, line_color=C_RED, line_width=1)
                sec("Margen % con promo (rojo) vs sin promo (azul) — ¿la promo destruye margen?")
                st.plotly_chart(pcfg(fig_mg_pp, max(320, len(_mg_df)*34), legend=False),
                                use_container_width=True)

        # ── Tabla detalle completa ──
        sec("Tabla Detalle — Penetración por Producto")
        _pp_show = df_pen_prod.copy().sort_values("Penetracion", ascending=False)
        if pen_cat_f != "(Todas)":
            _pp_show = _pp_show[_pp_show["Categoria"] == pen_cat_f]
        _pp_show["Penetración %"]   = _pp_show["Penetracion"].apply(lambda x: f"{x*100:.1f}%")
        _pp_show["Ingresos"]        = _pp_show["Ingresos"].apply(clp)
        _pp_show["Ing. Promo"]      = _pp_show["Ing_promo"].apply(clp)
        _pp_show["Ing. Sin Promo"]  = _pp_show["Ing_sin_promo"].apply(clp)
        _pp_show["Un. Promo"]       = _pp_show["Cant_promo"].apply(lambda x: f"{int(x):,}".replace(",","."))
        _pp_show["Un. Sin Promo"]   = _pp_show["Cant_sin_promo"].apply(lambda x: f"{int(x):,}".replace(",","."))
        _pp_show["Un. Total"]       = _pp_show["Cant_total"].apply(lambda x: f"{int(x):,}".replace(",","."))
        _pp_show["Alerta"] = _pp_show["Penetracion"].apply(
            lambda x: "🔴 Muy alta" if x > 0.7 else ("🟡 Alta" if x > 0.4 else ("🟢 Normal" if x > 0.1 else "⚪ Sin promo")))
        st.dataframe(
            _pp_show[["Alerta","Producto","Categoria","Penetración %",
                       "Un. Promo","Un. Sin Promo","Un. Total",
                       "Ing. Promo","Ing. Sin Promo","Ingresos"]],
            use_container_width=True, hide_index=True, height=380
        )

# ══════════════════════════════════════════════ TAB 7: ANÁLISIS PROMO
# ── Price Ladder (dentro de Cruces, antes de Análisis Promo) ──
# NOTA: se agrega como sección final de T[6] (Cruces)
with T[6]:
    if not df_price_ladder.empty:
        st.divider()
        st.markdown("#### 💲 Sección 7 — Price Ladder por Categoría")
        st.caption("Posicionamiento de precio de cada SKU dentro de su categoría. Detecta gaps, solapamiento y oportunidades.")

        _pl_cats = sorted(df_price_ladder["Categoria"].unique().tolist())
        _pl_cat_sel = st.selectbox("Categoría a analizar", _pl_cats, key="pl_cat")
        _pl = df_price_ladder[df_price_ladder["Categoria"] == _pl_cat_sel].copy()
        _pl = _pl.dropna(subset=["Precio_ef"]).sort_values("Precio_ef", ascending=False)

        if not _pl.empty:
            _pl["Precio_ef_fmt"] = _pl["Precio_ef"].apply(lambda x: f"${x:,.0f}".replace(",","."))
            _pl["Gap_lista"] = ((_pl["Precio_lista"] - _pl["Precio_ef"]) / _pl["Precio_lista"].replace(0, np.nan) * 100).round(1)

            fig_pl = go.Figure()
            # Precio lista (barra fondo)
            fig_pl.add_trace(go.Bar(
                name="Precio Lista", x=_pl["Precio_lista"], y=_pl["Producto"],
                orientation="h", marker_color="#CBD5E1", marker_line_width=0,
                hovertemplate="<b>%{y}</b><br>Precio lista: $%{x:,.0f}<extra></extra>"
            ))
            # Precio efectivo (barra encima)
            fig_pl.add_trace(go.Bar(
                name="Precio Efectivo", x=_pl["Precio_ef"], y=_pl["Producto"],
                orientation="h",
                marker=dict(
                    color=_pl["Mg_pct"],
                    colorscale=SEQ_RG,
                    colorbar=dict(title="Mg %", tickformat=".0%"),
                    line_width=0
                ),
                text=_pl["Precio_ef_fmt"],
                textposition="inside", textfont=dict(size=9, color="white"),
                insidetextanchor="middle",
                hovertemplate=(
                    "<b>%{y}</b><br>Precio ef.: $%{x:,.0f}<br>"
                    "Mg: %{marker.color:.1%}<extra></extra>"
                )
            ))
            fig_pl.update_layout(
                barmode="overlay",
                xaxis=dict(title="Precio ($)", tickprefix="$", tickformat=",.0f"),
                yaxis=dict(autorange="reversed"),
                legend=dict(orientation="h", y=1.08, font_size=10)
            )
            st.plotly_chart(pcfg(fig_pl, max(340, len(_pl)*30), legend=False),
                            use_container_width=True)

            # Tabla resumen
            _pl_show = _pl[["Producto","Precio_lista","Precio_ef","Gap_lista","Mg_pct","Unidades","Ingresos"]].copy()
            _pl_show["Precio lista"] = _pl_show["Precio_lista"].apply(lambda x: f"${x:,.0f}".replace(",","."))
            _pl_show["Precio ef."]   = _pl_show["Precio_ef"].apply(lambda x: f"${x:,.0f}".replace(",","."))
            _pl_show["Desc. %"]      = _pl_show["Gap_lista"].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "–")
            _pl_show["Margen %"]     = _pl_show["Mg_pct"].apply(lambda x: f"{x*100:.1f}%")
            _pl_show["Ingresos"]     = _pl_show["Ingresos"].apply(clp)
            _pl_show["Unidades"]     = _pl_show["Unidades"].apply(lambda x: f"{int(x):,}".replace(",","."))
            st.dataframe(
                _pl_show[["Producto","Precio lista","Precio ef.","Desc. %","Margen %","Unidades","Ingresos"]],
                use_container_width=True, hide_index=True, height=300
            )

with T[7]:
    st.markdown("### 📣 Análisis Promocional Completo")
    st.caption("Profundización en tickets, SKUs por promo, distribución por ciclo y categoría.")

    # ══════════════════════════════════════════════════════════
    #  SECCIÓN A — TICKET PROMEDIO CON PROMO vs SIN PROMO
    # ══════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### 🧾 Sección A — Ticket Promedio: Con Promo vs Sin Promo")
    st.caption("¿Las boletas con promoción tienen mayor o menor valor que las boletas sin promo?")

    if not df_tkt_mes.empty:
        _tm = df_tkt_mes.dropna(subset=["Tkt_con_promo","Tkt_sin_promo"]).copy()

        # KPIs agregados
        _tkt_cp  = (_tm["Ing_promo"].sum()     / _tm["Tkts_promo"].sum()     ) if _tm["Tkts_promo"].sum()     > 0 else 0
        _tkt_sp  = (_tm["Ing_sin_promo"].sum()  / _tm["Tkts_sin_promo"].sum()) if _tm["Tkts_sin_promo"].sum() > 0 else 0
        _tkt_tot = (_tm["Ing_total"].sum()       / _tm["Tkts_total"].sum()    ) if _tm["Tkts_total"].sum()     > 0 else 0
        _uplift  = (_tkt_cp - _tkt_sp) / _tkt_sp if _tkt_sp > 0 else 0

        kt = st.columns(4)
        kpi(kt[0], clp(_tkt_tot), "Ticket Promedio Global")
        kpi(kt[1], clp(_tkt_cp),  "Ticket CON Promo",    "g" if _tkt_cp > _tkt_tot else "a")
        kpi(kt[2], clp(_tkt_sp),  "Ticket SIN Promo",    "a")
        kpi(kt[3], f"{_uplift*100:+.1f}%", "Uplift de Ticket", "g" if _uplift > 0 else "r")

        c1_a, c2_a = st.columns(2)
        with c1_a:
            fig_tkt = go.Figure()
            fig_tkt.add_trace(go.Scatter(
                name="Sin Promo", x=_tm["Anio_mes"], y=_tm["Tkt_sin_promo"],
                mode="lines+markers", line=dict(color=C_NAVY, width=2.5),
                marker=dict(size=6),
                hovertemplate="<b>%{x}</b><br>Ticket sin promo: $%{y:,.0f}<extra></extra>"
            ))
            fig_tkt.add_trace(go.Scatter(
                name="Con Promo", x=_tm["Anio_mes"], y=_tm["Tkt_con_promo"],
                mode="lines+markers", line=dict(color=C_RED, width=2.5),
                marker=dict(size=6),
                fill="tonexty", fillcolor="rgba(200,16,46,0.07)",
                hovertemplate="<b>%{x}</b><br>Ticket con promo: $%{y:,.0f}<extra></extra>"
            ))
            fig_tkt.add_trace(go.Scatter(
                name="Promedio total", x=_tm["Anio_mes"], y=_tm["Tkt_promedio"],
                mode="lines", line=dict(color=C_AMBER, width=1.5, dash="dot"),
                hovertemplate="<b>%{x}</b><br>Ticket promedio: $%{y:,.0f}<extra></extra>"
            ))
            fig_tkt.update_yaxes(tickprefix="$", tickformat=",.0f", title="Valor de boleta ($)")
            fig_tkt.update_xaxes(title="Mes")
            fig_tkt.update_layout(legend=dict(orientation="h", y=1.1, x=0, font_size=10))
            sec("Evolución del Ticket Promedio — con promo (rojo) vs sin promo (azul)")
            st.plotly_chart(pcfg(fig_tkt, 310, legend=False), use_container_width=True)

        with c2_a:
            # Volumen de tickets con vs sin promo
            fig_tktvol = go.Figure()
            fig_tktvol.add_trace(go.Bar(
                name="Tkts sin promo", x=_tm["Anio_mes"], y=_tm["Tkts_sin_promo"],
                marker_color=C_NAVY, opacity=0.8, marker_line_width=0,
                hovertemplate="<b>%{x}</b><br>Boletas sin promo: %{y:,.0f}<extra></extra>"
            ))
            fig_tktvol.add_trace(go.Bar(
                name="Tkts con promo", x=_tm["Anio_mes"], y=_tm["Tkts_promo"],
                marker_color=C_RED, opacity=0.8, marker_line_width=0,
                hovertemplate="<b>%{x}</b><br>Boletas con promo: %{y:,.0f}<extra></extra>"
            ))
            fig_tktvol.update_layout(
                barmode="stack",
                yaxis=dict(title="N° boletas", tickformat=",.0f", gridcolor="#F3F4F6"),
                legend=dict(orientation="h", y=1.1, x=0, font_size=10)
            )
            sec("Volumen de Boletas — con promo (rojo) vs sin promo (azul)")
            st.plotly_chart(pcfg(fig_tktvol, 310, legend=False), use_container_width=True)

        # Uplift mensual
        _tm["Uplift"] = (_tm["Tkt_con_promo"] - _tm["Tkt_sin_promo"]) / _tm["Tkt_sin_promo"].replace(0, np.nan)
        _tm_u = _tm.dropna(subset=["Uplift"])
        if not _tm_u.empty:
            _tm_u["lbl"] = _tm_u["Uplift"].apply(lambda x: f"{x*100:+.1f}%")
            _tm_u["Color"] = _tm_u["Uplift"].apply(lambda x: C_GREEN if x >= 0 else C_RED)
            fig_upl = go.Figure(go.Bar(
                x=_tm_u["Anio_mes"], y=_tm_u["Uplift"],
                marker_color=_tm_u["Color"].tolist(),
                text=_tm_u["lbl"], textposition="outside", textfont_size=9,
                marker_line_width=0,
                hovertemplate="<b>%{x}</b><br>Uplift ticket: %{y:+.1%}<extra></extra>"
            ))
            fig_upl.add_hline(y=0, line_color="#374151", line_width=1.5)
            fig_upl.update_yaxes(tickformat="+.0%", title="Uplift %")
            sec("Uplift de Ticket — cuánto más (o menos) vale una boleta con promo vs sin promo")
            st.plotly_chart(pcfg(fig_upl, 240, legend=False), use_container_width=True)

    # ══════════════════════════════════════════════════════════
    #  SECCIÓN B — SKUs / CATEGORÍA POR PROMO (#3)
    # ══════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### 🏷️ Sección B — Ranking de Promos por Categoría y Desempeño")
    st.caption("Qué promos tuvieron salida real, en qué categorías, y cómo se comparan entre sí.")

    if not df_sku_promo.empty:
        # Filtros en línea
        cb1, cb2, cb3, cb4 = st.columns([2,2,1,2])
        with cb1:
            cat_b = st.selectbox("Categoría", ["(Todas)"] +
                                  sorted(df_sku_promo["Categoria"].unique().tolist()), key="cat_b")
        with cb2:
            orden_b = st.selectbox("Ordenar por", ["Ingresos","Unidades","ROI","Mg_pct","Desc_pct"], key="ord_b")
        with cb3:
            top_b = st.slider("Top N", 5, 50, 20, key="top_b")
        with cb4:
            prod_b = st.text_input("🔍 Producto en promo", placeholder="ej: Monster 473, Red Bull…", key="prod_b")

        _sp = df_sku_promo.copy()
        if cat_b != "(Todas)": _sp = _sp[_sp["Categoria"] == cat_b]
        if prod_b.strip():
            _sp = _sp[_sp["Promo"].str.contains(prod_b.strip(), case=False, na=False)]
        _sp = _sp.nlargest(top_b, orden_b).copy()

        if not _sp.empty:
            _sp["Semaforo"] = _sp.apply(lambda r:
                "🟢" if s(r["Mg_pct"]) >= 0.20 and s(r["ROI"]) >= 2
                else ("🟡" if s(r["Mg_pct"]) >= 0.10 or s(r["ROI"]) >= 1 else "🔴"), axis=1)

            # ── Preparar datos comunes de torta ──
            _pie_df = _sp.sort_values("Ingresos", ascending=False).copy()
            _pie_names = _pie_df["Promo"].tolist()
            _pie_ing   = _pie_df["Ingresos"].tolist()
            _pie_uni   = _pie_df["Unidades"].tolist()

            # "Sin Promo" solo cuando hay búsqueda de producto
            _ing_sin = 0
            _uni_sin = 0
            if prod_b.strip() and not df_pen_prod.empty:
                _match = df_pen_prod[df_pen_prod["Producto"].str.contains(
                    prod_b.strip(), case=False, na=False)]
                if not _match.empty:
                    _ing_sin = _match["Ing_sin_promo"].sum()
                    _uni_sin = _match["Cant_sin_promo"].sum() if "Cant_sin_promo" in _match.columns else 0

            _names_ing = _pie_names.copy()
            _vals_ing  = _pie_ing.copy()
            _names_uni = _pie_names.copy()
            _vals_uni  = _pie_uni.copy()

            if prod_b.strip():
                if _ing_sin > 0:
                    _names_ing.append(f"{prod_b.strip()} (sin promo)")
                    _vals_ing.append(_ing_sin)
                if _uni_sin > 0:
                    _names_uni.append(f"{prod_b.strip()} (sin promo)")
                    _vals_uni.append(_uni_sin)

            _n_pie = len(_pie_df)
            _colors_ing = [C_RED]*_n_pie + ([C_NAVY] if (prod_b.strip() and _ing_sin > 0) else [])
            _colors_uni = [C_RED]*_n_pie + ([C_NAVY] if (prod_b.strip() and _uni_sin > 0) else [])

            c1_b, c2_b = st.columns(2)

            with c1_b:
                fig_pie_ing = go.Figure(go.Pie(
                    labels=_names_ing, values=_vals_ing,
                    hole=0.42,
                    marker=dict(colors=_colors_ing, line=dict(color="white", width=1.5)),
                    texttemplate="%{label}<br><b>%{percent:.1%}</b>",
                    textfont_size=9,
                    hovertemplate="<b>%{label}</b><br>Ingresos: $%{value:,.0f}<br>%{percent:.1%}<extra></extra>",
                    sort=False
                ))
                _tot_ing = sum(_vals_ing)
                _lbl_ing = (f"<b>${_tot_ing/1e9:.2f}B</b>" if _tot_ing >= 1e9
                            else f"<b>${_tot_ing/1e6:.1f}M</b>" if _tot_ing >= 1e6
                            else f"<b>${_tot_ing/1e3:.0f}K</b>")
                fig_pie_ing.update_layout(
                    showlegend=False,
                    annotations=[dict(
                        text=f"{_lbl_ing}<br><span style='font-size:10px'>Total ingresos</span>",
                        x=0.5, y=0.5, font_size=14, showarrow=False,
                        font=dict(color=C_NAVY))]
                )
                _title_ing = (f"'{prod_b.strip()}' — Mix por Ingresos"
                              if prod_b.strip() else f"Top {top_b} Promos — Mix por Ingresos")
                sec(_title_ing)
                st.plotly_chart(pcfg(fig_pie_ing, 380, legend=False), use_container_width=True)

            with c2_b:
                fig_pie_uni = go.Figure(go.Pie(
                    labels=_names_uni, values=_vals_uni,
                    hole=0.42,
                    marker=dict(colors=_colors_uni, line=dict(color="white", width=1.5)),
                    texttemplate="%{label}<br><b>%{percent:.1%}</b>",
                    textfont_size=9,
                    hovertemplate="<b>%{label}</b><br>Unidades: %{value:,.0f}<br>%{percent:.1%}<extra></extra>",
                    sort=False
                ))
                _tot_uni = sum(_vals_uni)
                _lbl_uni = (f"<b>{_tot_uni/1e6:.1f}M</b>" if _tot_uni >= 1e6
                            else f"<b>{_tot_uni/1e3:.0f}K</b>" if _tot_uni >= 1e3
                            else f"<b>{int(_tot_uni):,}</b>")
                fig_pie_uni.update_layout(
                    showlegend=False,
                    annotations=[dict(
                        text=f"{_lbl_uni}<br><span style='font-size:10px'>Total unidades</span>",
                        x=0.5, y=0.5, font_size=14, showarrow=False,
                        font=dict(color=C_NAVY))]
                )
                _title_uni = (f"'{prod_b.strip()}' — Mix por Cantidad"
                              if prod_b.strip() else f"Top {top_b} Promos — Mix por Cantidad")
                sec(_title_uni)
                st.plotly_chart(pcfg(fig_pie_uni, 380, legend=False), use_container_width=True)

            # Tabla detalle
            sec("Detalle completo — Promos con salida real")
            _sp_t = _sp.copy()
            _sp_t["Ingresos"]  = _sp_t["Ingresos"].apply(clp)
            _sp_t["Unidades"]  = _sp_t["Unidades"].apply(lambda x: f"{int(x):,}".replace(",","."))
            _sp_t["Margen %"]  = _sp_t["Mg_pct"].apply(lambda x: f"{x*100:.1f}%")
            _sp_t["Desc %"]    = _sp_t["Desc_pct"].apply(lambda x: f"{x*100:.1f}%")
            _sp_t["ROI"]       = _sp_t["ROI"].apply(lambda x: f"{x:.2f}×" if pd.notna(x) else "–")
            _sp_t["Tiendas"]   = _sp_t["Tiendas"].apply(lambda x: f"{int(x)}")
            _sp_t["Período"]   = _sp_t["Primer_mes"] + " → " + _sp_t["Ultimo_mes"]
            st.dataframe(
                _sp_t[["Semaforo","Promo","Categoria","Ingresos","Unidades",
                        "Margen %","ROI","Desc %","Tiendas","Período"]],
                use_container_width=True, hide_index=True, height=340
            )

    # ══════════════════════════════════════════════════════════
    #  SECCIÓN C — DISTRIBUCIÓN POR CATEGORÍA × MES × PROMO
    # ══════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### 📅 Sección C — Distribución Promo por Categoría y Período")
    st.caption("Qué categorías concentran más actividad promocional en cada mes/ciclo.")

    if not df_promo_ciclo.empty:
        c1_c, c2_c = st.columns(2)
        with c1_c:
            # Heatmap: Ingresos promo por categoría × mes
            piv_pc = df_promo_ciclo.groupby(["Categoria","Anio_mes"])["Ingresos"].sum().reset_index()
            piv_pc = piv_pc.pivot(index="Categoria", columns="Anio_mes", values="Ingresos").fillna(0)
            piv_pc = piv_pc.loc[piv_pc.sum(axis=1).sort_values(ascending=False).index]

            def _fmt_heat(v):
                if v == 0: return ""
                if v >= 1e9: return f"${v/1e9:.1f}B"
                if v >= 1e6: return f"${v/1e6:.0f}M"
                return f"${v/1e3:.0f}K"

            lbl_pc = piv_pc.map(_fmt_heat).values
            fig_hpc = px.imshow(
                piv_pc, color_continuous_scale=["#EEF2F7","#93C5FD","#003B7A"],
                aspect="auto", zmin=0, text_auto=False
            )
            fig_hpc.update_traces(
                text=lbl_pc, texttemplate="%{text}",
                textfont=dict(size=9, color="white"),
                hovertemplate="<b>%{y}</b><br>%{x}<br>Ingresos promo: $%{z:,.0f}<extra></extra>"
            )
            fig_hpc.update_layout(
                coloraxis_colorbar=dict(title="Ingresos $", tickformat="$~s", len=0.6),
                xaxis_title="", yaxis_title=""
            )
            sec("Heatmap — Ingresos Promo por Categoría × Mes")
            st.plotly_chart(pcfg(fig_hpc, max(300, len(piv_pc)*34), legend=False),
                            use_container_width=True)

        with c2_c:
            # Barras apiladas: ingresos promo por categoría por mes
            _pc_bar = df_promo_ciclo.groupby(["Anio_mes","Categoria"])["Ingresos"].sum().reset_index()
            _pc_bar = _pc_bar.sort_values("Anio_mes")
            fig_pbar = px.bar(
                _pc_bar, x="Anio_mes", y="Ingresos",
                color="Categoria", color_discrete_sequence=CAT_COLORS,
                text_auto=False
            )
            fig_pbar.update_traces(marker_line_width=0)
            fig_pbar.update_layout(
                barmode="stack",
                xaxis_tickangle=-30,
                yaxis=dict(title="Ingresos promo ($)", tickprefix="$", tickformat="~s",
                           gridcolor="#F3F4F6"),
                legend=dict(orientation="v", x=1.01, font_size=9)
            )
            sec("Ingresos Promo Apilado por Categoría × Mes")
            st.plotly_chart(pcfg(fig_pbar, 340), use_container_width=True)

        # N° promos activas por mes
        _n_promos_mes = df_promo_ciclo.groupby("Anio_mes")["Promo"].nunique().reset_index()
        _n_promos_mes.columns = ["Anio_mes","N_promos"]
        _n_promos_mes["CicloAño"] = _n_promos_mes["Anio_mes"].map(MES_CICLO).fillna("")

        fig_np = go.Figure()
        fig_np.add_trace(go.Bar(
            x=_n_promos_mes["Anio_mes"], y=_n_promos_mes["N_promos"],
            marker_color=C_NAVY, opacity=0.85, marker_line_width=0,
            text=_n_promos_mes["N_promos"],
            textposition="outside", textfont_size=10,
            hovertemplate="<b>%{x}</b><br>Promos activas: %{y}<extra></extra>"
        ))
        fig_np.update_yaxes(title="N° promociones con venta real", gridcolor="#F3F4F6")
        sec("Promos Activas por Mes — cantidad de promos que generaron venta real")
        st.plotly_chart(pcfg(fig_np, 240, legend=False), use_container_width=True)

    # ══════════════════════════════════════════════════════════
    #  SECCIÓN D — CONTRIBUCIÓN AL MARGEN TOTAL
    # ══════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### 💼 Sección D — Contribución Promo al Margen Total del Negocio")
    st.caption("Cuánto del margen total del negocio proviene de ventas bajo promoción.")

    if not df_tkt_mes.empty and not df_sku_promo.empty:
        mg_promo_total = df_sku_promo["Ingresos"].sum() * df_sku_promo["Mg_pct"].mean() \
            if not df_sku_promo.empty else 0
        # Usar los datos de ventas ya calculados
        ing_promo_tot = df_tkt_mes["Ing_promo"].sum()
        ing_sin_tot   = df_tkt_mes["Ing_sin_promo"].sum()
        ing_total_d   = df_tkt_mes["Ing_total"].sum()

        c1_d, c2_d = st.columns(2)
        with c1_d:
            # Donut: mix ingresos promo vs sin promo
            _donut = pd.DataFrame({
                "Tipo":     ["Con Promo", "Sin Promo"],
                "Ingresos": [ing_promo_tot, ing_sin_tot]
            })
            fig_donut = px.pie(
                _donut, values="Ingresos", names="Tipo",
                color="Tipo",
                color_discrete_map={"Con Promo": C_RED, "Sin Promo": C_NAVY},
                hole=0.55
            )
            fig_donut.update_traces(
                texttemplate="%{label}<br><b>%{percent:.1%}</b>",
                textfont_size=12,
                hovertemplate="<b>%{label}</b><br>Ingresos: $%{value:,.0f} (%{percent:.1%})<extra></extra>"
            )
            fig_donut.update_layout(showlegend=False,
                                    annotations=[dict(text=f"<b>{clp(ing_total_d)}</b><br>Total",
                                                      x=0.5, y=0.5, font_size=13, showarrow=False)])
            sec("Split de Ingresos — Promo vs Base")
            st.plotly_chart(pcfg(fig_donut, 300, legend=False), use_container_width=True)

        with c2_d:
            # Contribución por categoría: margen promo vs total
            _contrib = df_sku_promo.groupby("Categoria").agg(
                Ing_promo=("Ingresos","sum"),
                Mg_pct_prom=("Mg_pct","mean")
            ).reset_index()
            _contrib["Margen_promo_est"] = _contrib["Ing_promo"] * _contrib["Mg_pct_prom"]
            _contrib["Contribucion"] = _contrib["Margen_promo_est"] / _contrib["Margen_promo_est"].sum()
            _contrib = _contrib.sort_values("Contribucion", ascending=False).head(12)
            _contrib["lbl"] = _contrib["Contribucion"].apply(lambda x: f"{x*100:.1f}%")

            fig_contrib = px.bar(
                _contrib, x="Contribucion", y="Categoria", orientation="h",
                color="Mg_pct_prom",
                color_continuous_scale=["#C8102E","#F5A623","#1A9E5C"],
                text="lbl",
                custom_data=["Ing_promo","Margen_promo_est"]
            )
            fig_contrib.update_traces(
                textposition="outside", textfont_size=10, marker_line_width=0,
                hovertemplate=(
                    "<b>%{y}</b><br>Contribución al margen promo: %{x:.1%}<br>"
                    "Ingresos promo: $%{customdata[0]:,.0f}<br>"
                    "Margen promo est.: $%{customdata[1]:,.0f}<extra></extra>"
                )
            )
            fig_contrib.update_layout(
                yaxis=dict(autorange="reversed"),
                xaxis=dict(tickformat=".0%"),
                coloraxis_colorbar=dict(title="Mg %", tickformat=".0%")
            )
            sec("Contribución al Margen Promo por Categoría")
            st.plotly_chart(pcfg(fig_contrib, 340, legend=False), use_container_width=True)

    # ══════════════════════════════════════════════════════════
    #  SECCIÓN E — CATÁLOGO DE PROMOS: TRANSACCIONES REALES Y SIN VENTA
    # ══════════════════════════════════════════════════════════
    st.divider()
    st.markdown("#### 📋 Sección E — Catálogo de Promos: Transacciones Reales y Cobertura")
    st.caption(
        "Usando el catálogo oficial de promos, normaliza unidades brutas → transacciones reales "
        "(ej: 1000 unidades de una promo 2x = 500 transacciones). "
        "También detecta promos planificadas que nunca tuvieron venta."
    )

    if df_catalogo.empty:
        st.info(
            "Catálogo de promos no cargado. Sube el archivo CSV con las columnas "
            "**Cod_promo, Nom_promo, Formato_promo, Unidades_por_promo** desde el sidebar."
        )
    elif df_sku_promo.empty:
        st.info("Sin datos de ventas promocionales para cruzar con el catálogo.")
    else:
        # ── Join: ventas reales con catálogo ──
        _cat = df_catalogo.copy()
        _ven = df_sku_promo[["Promo","Cod_promo","Categoria","Ingresos","Unidades",
                               "Mg_pct","Tiendas","Primer_mes","Ultimo_mes"]].copy()
        _ven["Cod_promo_up"] = _ven["Cod_promo"].str.strip().str.upper()
        _cat["Cod_promo_up"] = _cat["Cod_promo"].str.strip().str.upper()

        _join = _ven.merge(_cat[["Cod_promo_up","Formato_promo","Unidades_por_promo"]],
                           on="Cod_promo_up", how="left")
        _join["Unidades_por_promo"] = _join["Unidades_por_promo"].fillna(1).astype(int)
        _join["Transacciones_promo"] = (_join["Unidades"] / _join["Unidades_por_promo"]).round(0).astype(int)
        _join["Formato_promo"] = _join["Formato_promo"].fillna("Sin catálogo")

        # ── Promos sin venta (en catálogo pero no en ventas) ──
        _cod_con_venta = set(_ven["Cod_promo_up"].dropna().unique())
        _sin_venta = _cat[~_cat["Cod_promo_up"].isin(_cod_con_venta)].copy()

        # KPIs
        _n_cat   = len(_cat)
        _n_venta = _join["Cod_promo_up"].nunique()
        _n_sin   = len(_sin_venta)
        _cobert  = _n_venta / _n_cat if _n_cat > 0 else 0
        _tot_tran = _join["Transacciones_promo"].sum()
        _tot_uni  = _join["Unidades"].sum()
        _factor   = _tot_tran / _tot_uni if _tot_uni > 0 else 1

        ke = st.columns(5)
        kpi(ke[0], f"{_n_cat:,}", "Promos en catálogo")
        kpi(ke[1], f"{_n_venta:,}", "Promos con venta real", "g")
        kpi(ke[2], f"{_n_sin:,}", "Promos sin ninguna venta", "r" if _n_sin > 0 else "g")
        kpi(ke[3], f"{_cobert*100:.1f}%", "Cobertura del catálogo", "g" if _cobert >= 0.7 else "a")
        kpi(ke[4], f"{_tot_tran:,.0f}", "Transacciones promo reales")

        c1_e, c2_e = st.columns(2)

        with c1_e:
            # Tabla: promos con venta + transacciones reales normalizadas
            sec("Promos con Venta — Unidades Brutas vs Transacciones Reales")
            _norm_t = _join[["Promo","Categoria","Formato_promo","Unidades_por_promo",
                              "Unidades","Transacciones_promo","Ingresos","Mg_pct","Tiendas"]].copy()
            _norm_t = _norm_t.sort_values("Transacciones_promo", ascending=False)
            _norm_t["Ingresos"]            = _norm_t["Ingresos"].apply(clp)
            _norm_t["Margen %"]            = _norm_t["Mg_pct"].apply(lambda x: f"{x*100:.1f}%")
            _norm_t["Unidades brutas"]     = _norm_t["Unidades"].apply(lambda x: f"{int(x):,}".replace(",","."))
            _norm_t["Transacciones reales"]= _norm_t["Transacciones_promo"].apply(lambda x: f"{int(x):,}".replace(",","."))
            _norm_t["Unid/promo"]          = _norm_t["Unidades_por_promo"].astype(str) + "x"
            st.dataframe(
                _norm_t[["Promo","Categoria","Formato_promo","Unid/promo",
                          "Unidades brutas","Transacciones reales","Ingresos","Margen %","Tiendas"]],
                use_container_width=True, hide_index=True, height=360,
                column_config={
                    "Promo": st.column_config.TextColumn("Promoción", width=220),
                    "Formato_promo": st.column_config.TextColumn("Formato"),
                    "Unid/promo": st.column_config.TextColumn("Pack"),
                }
            )

        with c2_e:
            # Bar chart: top promos por transacciones reales
            _top_e = _join.nlargest(20, "Transacciones_promo").sort_values("Transacciones_promo")
            _top_e["lbl"] = _top_e.apply(
                lambda r: f"{int(r['Transacciones_promo']):,} ({r['Unidades_por_promo']}x)".replace(",","."), axis=1)
            fig_e1 = go.Figure(go.Bar(
                x=_top_e["Transacciones_promo"],
                y=_top_e["Promo"],
                orientation="h",
                marker_color=C_RED, marker_line_width=0, opacity=0.88,
                text=_top_e["lbl"],
                textposition="inside", textfont=dict(size=9, color="white"),
                hovertemplate="<b>%{y}</b><br>Transacciones reales: %{x:,.0f}<extra></extra>"
            ))
            fig_e1.update_layout(
                yaxis=dict(autorange="reversed"),
                xaxis=dict(title="Transacciones promo reales", tickformat=",.0f")
            )
            sec("Top 20 Promos — Transacciones Reales (unidades ÷ pack)")
            st.plotly_chart(pcfg(fig_e1, max(340, len(_top_e)*28), legend=False),
                            use_container_width=True)

        # ── Promos sin venta ──
        if not _sin_venta.empty:
            st.divider()
            sec(f"⚠️ {len(_sin_venta)} Promos del catálogo SIN ninguna venta registrada")
            _sv_t = _sin_venta[["Cod_promo","Nom_promo","Formato_promo","Unidades_por_promo"]].copy()
            _sv_t.columns = ["Código promo","Nombre","Formato","Pack"]
            _sv_t["Ciclo"] = _sv_t["Código promo"].apply(
                lambda c: "-".join(str(c).split("-")[:2]) if c else "")
            _sv_t = _sv_t.sort_values("Ciclo")
            st.dataframe(_sv_t, use_container_width=True, hide_index=True, height=300)

            # Distribución por ciclo de las promos sin venta
            _sv_ciclo = _sv_t.groupby("Ciclo").size().reset_index(name="N_sin_venta")
            _sv_ciclo = _sv_ciclo.sort_values("Ciclo")
            fig_sv = go.Figure(go.Bar(
                x=_sv_ciclo["Ciclo"], y=_sv_ciclo["N_sin_venta"],
                marker_color=C_AMBER, marker_line_width=0, opacity=0.9,
                text=_sv_ciclo["N_sin_venta"], textposition="outside", textfont_size=10,
                hovertemplate="<b>%{x}</b><br>Promos sin venta: %{y}<extra></extra>"
            ))
            fig_sv.update_yaxes(title="N° promos sin venta", gridcolor="#F3F4F6")
            sec("Promos sin venta por ciclo — cuántas promos del catálogo no se activaron")
            st.plotly_chart(pcfg(fig_sv, 220, legend=False), use_container_width=True)
        else:
            st.success("Todas las promos del catálogo tienen al menos una venta registrada.")

# ══════════════════════════════════════════════ TAB 8: YoY
with T[8]:
    st.markdown("### 📅 Comparativa Año vs Año (YoY)")
    st.caption("Compara los mismos meses entre distintos años para detectar tendencias reales vs estacionalidad.")

    if df_yoy_mes.empty:
        st.info("Sin datos suficientes para comparativa YoY.")
    else:
        años_disp = sorted(df_yoy_mes["Año"].unique().tolist())

        # ── KPIs YoY ──
        if len(años_disp) >= 2:
            _yr_last = años_disp[-1]
            _yr_prev = años_disp[-2]
            _ing_last = df_yoy_mes[df_yoy_mes["Año"]==_yr_last]["Ingresos"].sum()
            _ing_prev = df_yoy_mes[df_yoy_mes["Año"]==_yr_prev]["Ingresos"].sum()
            _mg_last  = df_yoy_mes[df_yoy_mes["Año"]==_yr_last]["Margen"].sum()
            _mg_prev  = df_yoy_mes[df_yoy_mes["Año"]==_yr_prev]["Margen"].sum()
            _un_last  = df_yoy_mes[df_yoy_mes["Año"]==_yr_last]["Unidades"].sum()
            _un_prev  = df_yoy_mes[df_yoy_mes["Año"]==_yr_prev]["Unidades"].sum()
            _var_ing  = (_ing_last/_ing_prev - 1) if _ing_prev else 0
            _var_mg   = (_mg_last/_mg_prev - 1)  if _mg_prev else 0
            _var_un   = (_un_last/_un_prev - 1)  if _un_prev else 0

            ky = st.columns(5)
            kpi(ky[0], clp(_ing_last), f"Ingresos {_yr_last}")
            kpi(ky[1], f"{_var_ing:+.1%}", f"vs {_yr_prev}", "g" if _var_ing>=0 else "r")
            kpi(ky[2], clp(_mg_last),  f"Margen {_yr_last}")
            kpi(ky[3], f"{_var_mg:+.1%}", f"Margen vs {_yr_prev}", "g" if _var_mg>=0 else "r")
            kpi(ky[4], f"{_var_un:+.1%}", "Unidades YoY", "g" if _var_un>=0 else "r")

        # ── Ingresos por mes comparado ──
        sec("Ingresos por Mes — comparativa entre años")
        _yoy = df_yoy_mes.copy()
        fig_yoy = px.bar(
            _yoy, x="Mes", y="Ingresos", color="Año",
            barmode="group",
            color_discrete_sequence=[C_NAVY, C_RED, C_AMBER, C_GREEN],
            text_auto=False,
            custom_data=["Año","Ingresos","Mg_pct"]
        )
        fig_yoy.update_traces(
            marker_line_width=0,
            hovertemplate="<b>Mes %{x} — %{customdata[0]}</b><br>Ingresos: $%{y:,.0f}<br>Mg %%: %{customdata[2]:.1%}<extra></extra>"
        )
        fig_yoy.update_layout(
            xaxis=dict(title="Mes", tickmode="linear"),
            yaxis=dict(title="Ingresos ($)", tickprefix="$", tickformat="~s", gridcolor="#F3F4F6"),
            legend=dict(orientation="h", y=1.08, font_size=10)
        )
        st.plotly_chart(pcfg(fig_yoy, 340, legend=False), use_container_width=True)

        # ── Margen % por mes y año ──
        c1_yoy, c2_yoy = st.columns(2)
        with c1_yoy:
            sec("Margen % por Mes — evolución YoY")
            fig_mg_yoy = px.line(
                _yoy, x="Mes", y="Mg_pct", color="Año",
                markers=True,
                color_discrete_sequence=[C_NAVY, C_RED, C_AMBER, C_GREEN],
                custom_data=["Año","Mg_pct"]
            )
            fig_mg_yoy.update_traces(
                line_width=2.5,
                hovertemplate="<b>Mes %{x} — %{customdata[0]}</b><br>Margen %%: %{customdata[1]:.1%}<extra></extra>"
            )
            fig_mg_yoy.update_layout(
                xaxis=dict(title="Mes"),
                yaxis=dict(title="Margen %", tickformat=".1%", gridcolor="#F3F4F6"),
                legend=dict(orientation="h", y=1.1, font_size=10)
            )
            st.plotly_chart(pcfg(fig_mg_yoy, 300, legend=False), use_container_width=True)

        with c2_yoy:
            sec("Mix Promo % por Mes — evolución YoY")
            fig_mix_yoy = px.line(
                _yoy, x="Mes", y="Mix_promo", color="Año",
                markers=True,
                color_discrete_sequence=[C_NAVY, C_RED, C_AMBER, C_GREEN],
                custom_data=["Año","Mix_promo"]
            )
            fig_mix_yoy.update_traces(
                line_width=2.5,
                hovertemplate="<b>Mes %{x} — %{customdata[0]}</b><br>Mix promo: %{customdata[1]:.1%}<extra></extra>"
            )
            fig_mix_yoy.update_layout(
                xaxis=dict(title="Mes"),
                yaxis=dict(title="Mix Promo %", tickformat=".1%", gridcolor="#F3F4F6"),
                legend=dict(orientation="h", y=1.1, font_size=10)
            )
            st.plotly_chart(pcfg(fig_mix_yoy, 300, legend=False), use_container_width=True)

        # ── Ingresos YoY por categoría ──
        if not df_yoy_cat.empty:
            st.divider()
            sec("Ingresos por Categoría — comparativa YoY")
            _yc = df_yoy_cat.copy()
            # Calcular variación para cada categoría
            if len(años_disp) >= 2:
                _yc_last = _yc[_yc["Año"]==_yr_last][["Categoria","Ingresos"]].rename(columns={"Ingresos":"Ing_last"})
                _yc_prev = _yc[_yc["Año"]==_yr_prev][["Categoria","Ingresos"]].rename(columns={"Ingresos":"Ing_prev"})
                _yc_var = _yc_last.merge(_yc_prev, on="Categoria", how="outer").fillna(0)
                _yc_var["Var_pct"] = (_yc_var["Ing_last"]/_yc_var["Ing_prev"].replace(0,np.nan) - 1)
                _yc_var = _yc_var.sort_values("Var_pct", ascending=False)
                _yc_var["lbl"] = _yc_var["Var_pct"].apply(lambda x: f"{x:+.1%}" if pd.notna(x) else "Nuevo")
                fig_yoy_cat = px.bar(
                    _yc_var, x="Var_pct", y="Categoria", orientation="h",
                    color="Var_pct",
                    color_continuous_scale=["#C8102E","#FFFFFF","#1A9E5C"],
                    color_continuous_midpoint=0,
                    text="lbl",
                    custom_data=["Ing_last","Ing_prev"]
                )
                fig_yoy_cat.update_traces(
                    textposition="outside", textfont_size=10, marker_line_width=0,
                    hovertemplate=(
                        "<b>%{y}</b><br>Variación: %{x:.1%}<br>"
                        f"Ingresos {_yr_last}: $%{{customdata[0]:,.0f}}<br>"
                        f"Ingresos {_yr_prev}: $%{{customdata[1]:,.0f}}<extra></extra>"
                    )
                )
                fig_yoy_cat.add_vline(x=0, line_color=C_GRAY, line_width=1)
                fig_yoy_cat.update_layout(
                    xaxis=dict(title=f"Variación % vs {_yr_prev}", tickformat=".0%"),
                    yaxis=dict(autorange="reversed"),
                    coloraxis_showscale=False
                )
                st.plotly_chart(pcfg(fig_yoy_cat, max(300, len(_yc_var)*30), legend=False),
                                use_container_width=True)

# ══════════════════════════════════════════════ TAB 9: RANKING SKU
with T[9]:
    st.markdown("### 🏆 Ranking Completo de SKUs")
    st.caption("Todos los productos rankeados por distintas métricas. Usa los filtros del sidebar para segmentar.")

    if df_sku_ranking.empty:
        st.info("Sin datos de productos.")
    else:
        # ── Controles ──
        r_col1, r_col2, r_col3 = st.columns(3)
        with r_col1:
            _rk_metric = st.selectbox("Ordenar por", ["Ingresos","Margen","Unidades","Mg_pct","Desc_pct"], key="rk_metric")
        with r_col2:
            _rk_cat = st.selectbox("Categoría", ["(Todas)"] + sorted(df_sku_ranking["Categoria"].unique().tolist()), key="rk_cat")
        with r_col3:
            _rk_n = st.slider("Top N", 10, 100, 30, key="rk_n")

        _rk = df_sku_ranking.copy()
        if _rk_cat != "(Todas)":
            _rk = _rk[_rk["Categoria"] == _rk_cat]
        _rk = _rk.sort_values(_rk_metric, ascending=(_rk_metric in ["Desc_pct"]))

        top_rk = _rk.head(_rk_n).copy()
        bot_rk = _rk.tail(_rk_n).copy().sort_values(_rk_metric, ascending=True)

        # ── KPIs ──
        kr = st.columns(4)
        kpi(kr[0], num(len(_rk)), "SKUs en selección")
        kpi(kr[1], clp(_rk["Ingresos"].sum()), "Ingresos totales")
        kpi(kr[2], pct(_rk["Mg_pct"].mean()), "Margen % promedio", "g")
        kpi(kr[3], pct(_rk["Desc_pct"].mean()), "Descuento % promedio", "a")

        # ── Top performers ──
        sec(f"Top {_rk_n} SKUs — mayor {_rk_metric}")
        fig_top = px.bar(
            top_rk, x=_rk_metric, y="Producto", orientation="h",
            color="Mg_pct", color_continuous_scale=SEQ_RG,
            custom_data=["Categoria","Ingresos","Mg_pct","Desc_pct"]
        )
        fig_top.update_traces(
            marker_line_width=0,
            hovertemplate=(
                "<b>%{y}</b><br>Cat.: %{customdata[0]}<br>"
                "Ingresos: $%{customdata[1]:,.0f}<br>"
                "Margen %%: %{customdata[2]:.1%}<br>"
                "Desc. %%: %{customdata[3]:.1%}<extra></extra>"
            )
        )
        fig_top.update_layout(
            yaxis=dict(autorange="reversed"),
            coloraxis_colorbar=dict(title="Mg %", tickformat=".0%")
        )
        st.plotly_chart(pcfg(fig_top, max(340, _rk_n*24), legend=False), use_container_width=True)

        # ── Bottom performers (menor margen) ──
        st.divider()
        sec(f"Bottom {_rk_n} SKUs — menor margen %")
        _bot_mg = _rk.nsmallest(_rk_n, "Mg_pct").copy()
        fig_bot = px.bar(
            _bot_mg, x="Mg_pct", y="Producto", orientation="h",
            color="Mg_pct", color_continuous_scale=["#C8102E","#F5A623","#1A9E5C"],
            custom_data=["Categoria","Ingresos","Desc_pct"]
        )
        fig_bot.update_traces(
            marker_line_width=0,
            hovertemplate=(
                "<b>%{y}</b><br>Cat.: %{customdata[0]}<br>"
                "Ingresos: $%{customdata[1]:,.0f}<br>"
                "Margen %%: %{x:.1%}<br>"
                "Desc. %%: %{customdata[2]:.1%}<extra></extra>"
            )
        )
        fig_bot.update_layout(
            yaxis=dict(autorange="reversed"),
            xaxis=dict(tickformat=".1%"),
            coloraxis_colorbar=dict(title="Mg %", tickformat=".0%")
        )
        st.plotly_chart(pcfg(fig_bot, max(340, _rk_n*24), legend=False), use_container_width=True)

        # ── Scatter: Ingresos vs Margen % ──
        st.divider()
        sec("Mapa de Posición — Ingresos vs Margen % (tamaño = unidades)")
        _rk_sc = _rk.head(100).dropna(subset=["Mg_pct","Ingresos"]).copy()
        _rk_med_ing = _rk_sc["Ingresos"].median()
        _rk_med_mg  = _rk_sc["Mg_pct"].median()
        fig_sc_rk = px.scatter(
            _rk_sc, x="Ingresos", y="Mg_pct",
            size=sz(_rk_sc["Unidades"]),
            color="Categoria", color_discrete_sequence=CAT_COLORS,
            hover_name="Producto",
            size_max=35, opacity=0.80,
            custom_data=["Categoria","Desc_pct","Precio_ef"]
        )
        fig_sc_rk.update_traces(
            hovertemplate=(
                "<b>%{hovertext}</b><br>Cat.: %{customdata[0]}<br>"
                "Ingresos: $%{x:,.0f}<br>Margen %%: %{y:.1%}<br>"
                "Desc. %%: %{customdata[1]:.1%}<br>Precio ef.: $%{customdata[2]:,.0f}<extra></extra>"
            )
        )
        fig_sc_rk.add_vline(x=_rk_med_ing, line_dash="dot", line_color=C_GRAY, opacity=0.5)
        fig_sc_rk.add_hline(y=_rk_med_mg,  line_dash="dot", line_color=C_GRAY, opacity=0.5)
        fig_sc_rk.update_layout(
            xaxis=dict(title="Ingresos ($)", tickprefix="$", tickformat="~s"),
            yaxis=dict(title="Margen %", tickformat=".1%"),
        )
        st.plotly_chart(pcfg(fig_sc_rk, 420), use_container_width=True)

        # ── Tabla completa ──
        st.divider()
        sec("Tabla Completa de SKUs")
        _rk_tbl = _rk.copy()
        _rk_tbl["#"] = range(1, len(_rk_tbl)+1)
        _rk_tbl["Semaforo"] = _rk_tbl["Mg_pct"].apply(
            lambda x: "🟢" if x > _rk["Mg_pct"].quantile(0.66) else ("🟡" if x > _rk["Mg_pct"].quantile(0.33) else "🔴"))
        _rk_tbl["Ingresos"]    = _rk_tbl["Ingresos"].apply(clp)
        _rk_tbl["Margen"]      = _rk_tbl["Margen"].apply(clp)
        _rk_tbl["Mg %"]        = _rk_tbl["Mg_pct"].apply(lambda x: f"{x*100:.1f}%")
        _rk_tbl["Precio ef."]  = _rk_tbl["Precio_ef"].apply(lambda x: f"${x:,.0f}".replace(",",".") if pd.notna(x) else "–")
        _rk_tbl["Desc. %"]     = _rk_tbl["Desc_pct"].apply(lambda x: f"{x*100:.1f}%")
        _rk_tbl["Unidades"]    = _rk_tbl["Unidades"].apply(lambda x: f"{int(x):,}".replace(",","."))
        st.dataframe(
            _rk_tbl[["#","Semaforo","Producto","Categoria","Ingresos","Margen","Mg %",
                      "Precio ef.","Desc. %","Unidades"]],
            use_container_width=True, hide_index=True, height=420
        )

# ══════════════════════════════════════════════ TAB 10: DATOS
with T[10]:
    sec("Exportar Datos")
    st.caption(f"{len(df_res):,} filas con el filtro actual")

    # Descripción del filtro activo para el Excel
    _meses_str = f"{_ms[0]} → {_ms[-1]}" if _ms else "Todos"
    _filtro_desc = f"Período: {_meses_str}  |  Categorías: {len(_cats)}  |  Generado: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"

    col_dl1, col_dl2, col_dl3 = st.columns(3)
    with col_dl1:
        csv_v = df_res.to_csv(index=False, sep=";", decimal=",").encode("utf-8-sig")
        st.download_button("⬇️ Ventas (CSV)", csv_v, "enex_ventas.csv", "text/csv",
                           use_container_width=True)
    with col_dl2:
        csv_p = pf.to_csv(index=False, sep=";", decimal=",").encode("utf-8-sig")
        st.download_button("⬇️ Productos (CSV)", csv_p, "enex_productos.csv", "text/csv",
                           use_container_width=True)
    with col_dl3:
        with st.spinner("Preparando Excel…"):
            _excel_bytes = generar_excel(D, _filtro_desc)
        st.download_button(
            "📊 Descargar Excel con Gráficos",
            _excel_bytes,
            "enex_pricing_intelligence.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )

    st.caption("El Excel incluye 7 hojas: Resumen, Categorías, SKU Ranking, Tiendas, Promociones, YoY y Datos.")

    df_show = df_res.copy()
    for col in ["Ingresos","Costo","Margen","Descuento"]:
        df_show[col] = df_show[col].apply(lambda x: f"${x:,.0f}".replace(",","."))
    df_show["Margen_pct"] = df_show["Margen_pct"].apply(lambda x: f"{x:.1f}%")
    st.dataframe(df_show, use_container_width=True, hide_index=True)
